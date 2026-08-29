# %%
# ==================================================
# MLB データ収集 → データマート 一括生成パイプライン  v2
#
# データソース: pybaseball (Baseball Savant / Statcast)
#
# ステップ順:
#   Step 1: games      試合データ取得    → raw/{date}/statcast_{date}.xlsx  (pybaseball)
#                                          or raw/{date}/statsapi_{date}.xlsx (statsapi速報)
#                                          → games/datamart/{date}.xlsx
#                                          → games/json/{date}.json
#   Step 2: highlights 活躍選手選出     → games/datamart/{date}.xlsx 更新
#                                          → games/json/{date}.json 更新
#   Step 3: datamart   データマートのみ再生成（活躍選手なし）
#                                          → games/datamart/{date}.xlsx
#                                          → games/json/{date}.json
# --steps の指定方法:
#   --steps games       Step1 のみ（RAW取得→datamart→JSON、活躍選手なし）
#   --steps highlights  Step2 のみ（RAW再利用→highlights.xlsx→datamart更新→JSON更新）
#   --steps games,highlights  Step1+2 連続実行（推奨）
#   --steps datamart    datamart再生成（活躍選手なし・RAW再利用）
#   --steps all         全ステップ実行（デフォルト）
#
# 出力フォルダ構成:
#   data/MLB/{YYYY}年/公式戦/
#     raw/{YYYY-MM-DD}/statcast_{date}.xlsx   ← pybaseball完全版
#     raw/{YYYY-MM-DD}/statsapi_{date}.xlsx   ← statsapi速報版
#     games/datamart/{date}.xlsx
#     games/json/{date}.json
#
# v2 追加指標（MLB/Statcast独自）:
#   [打者]
#     打球速度(avg/max/Hard-Hit%), 打球角度(avg), 飛距離(avg/max)
#     xBA, xwOBA, xSLG, wOBA, BABIP, ISO
#     バットスピード(avg), スイング軌跡長さ(avg), アタックアングル(avg)
#     Barrel%, EV95%(ハードヒット率)
#   [投手/球種]
#     回転数(avg/max), 変化量(横pfx_x/縦pfx_z cm), アーム角度
#     Extension(リリース距離), リリースポイント(高さ/横)
#     IVB/HB (Induced Break), 実効球速
#     xwOBA被打, wOBA被打
#   [活躍選手]
#     prompts/活躍選手_MLB.txt のシステムプロンプトを使用し
#     Claude API にスコアプレー・野手成績・投手成績を渡して生成
#     GEMINI_API_KEY 未設定時はスキップ
# ==================================================

# %%
# ==================================================
# Section 1. ライブラリのインポート
# ==================================================
import sys
import os
import json
import math
import logging
import argparse
from pathlib import Path

# スクリプト自身の場所を基準にしたルートディレクトリ
# scripts/ の親 = 野球ダッシュボード/ をプロジェクトルートとする
_SCRIPT_DIR = Path(__file__).resolve().parent.parent

import numpy as np
import pandas as pd
import datetime
from pybaseball import statcast, playerid_reverse_lookup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

try:
    import requests as _requests
    _REQUESTS_AVAILABLE = True
except ImportError:
    _REQUESTS_AVAILABLE = False

try:
    import statsapi as _statsapi
    _STATSAPI_AVAILABLE = True
except ImportError:
    _STATSAPI_AVAILABLE = False

try:
    import google.generativeai as genai
    from dotenv import load_dotenv
    _GENAI_AVAILABLE = True
except ImportError:
    _GENAI_AVAILABLE = False

# .env 読み込み（GEMINI_API_KEY / GEMINI_API_KEY）
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(dotenv_path=_SCRIPT_DIR / ".env")
except Exception:
    pass

logging.basicConfig(level=logging.WARNING, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# %%
# ==================================================
# Section 2. 設定
# ==================================================

TARGET_DATE   = "2026-04-18"
# data/MLB/{YYYY}年/{試合種別}/           ← 非公開（raw/datamart、gitでバックアップ管理のみ）
# docs/data/MLB/{YYYY}年/{試合種別}/      ← 公開（games/json、GitHub Pagesの公開元）
BASE_DATA_DIR   = str(_SCRIPT_DIR / "data" / "MLB")
BASE_PUBLIC_DIR = str(_SCRIPT_DIR / "docs" / "data" / "MLB")

def set_dirs(date: str = TARGET_DATE, game_type: str = "公式戦") -> None:
    global TARGET_DATE, RAW_DIR, GAMES_DM_DIR, GAMES_JSON_DIR
    TARGET_DATE = date
    year = date[:4]
    base        = os.path.join(BASE_DATA_DIR,   f"{year}年", game_type)
    base_public = os.path.join(BASE_PUBLIC_DIR, f"{year}年", game_type)
    RAW_DIR                 = os.path.join(base, "raw", date)
    GAMES_DM_DIR             = os.path.join(base, "games", "datamart")
    GAMES_JSON_DIR           = os.path.join(base_public, "games", "json")


set_dirs(TARGET_DATE)

# %%
# ==================================================
# Section 2b. 活躍選手 LLM 設定
# ==================================================

# ── Gemini ──────────────────────────────────────
# .env に GEMINI_API_KEY=xxx を記載するか環境変数で設定
GEMINI_MODEL             = "gemini-2.5-flash"
GEMINI_MAX_OUTPUT_TOKENS = 16384
GEMINI_TEMPERATURE       = 0
# プロンプトファイルのパス
HIGHLIGHTS_prompts_FILE = str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt")

# ── Claude（フォールバック）────────────────────
# .env に CLAUDE_API_KEY=xxx を記載するか環境変数で設定
CLAUDE_MODEL      = "claude-haiku-4-5-20251001"
CLAUDE_MAX_TOKENS = 2048

# ── LLM 優先順位 ────────────────────────────────
# "gemini" → Gemini を優先し、失敗時は Claude にフォールバック
# "claude" → Claude のみ使用
# "auto"   → 設定済みAPIキーを自動判定（Gemini優先）
LLM_PROVIDER = "gemini"

# %%
# ==================================================
# Section 2c. 単位変換設定
# ==================================================

# mph → km/h 変換係数（1 mph = 1.60934 km/h）
MPH_TO_KMH = 1.60934
# feet → m 変換係数（1 ft = 0.3048 m）
FT_TO_M    = 0.3048
# inches → cm 変換係数（1 inch = 2.54 cm）
IN_TO_CM   = 2.54

def mph2kmh(v, d=1):
    """mph → km/h。None/NaN は np.nan を返す"""
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return np.nan
        return round(float(v) * MPH_TO_KMH, d)
    except Exception:
        return np.nan

def ft2m(v, d=2):
    """feet → m。None/NaN は np.nan を返す"""
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return np.nan
        return round(float(v) * FT_TO_M, d)
    except Exception:
        return np.nan

# %%
# ==================================================
# Section 3. 定数マップ
# ==================================================

PITCH_TYPE_MAP = {
    "FF": "フォーシーム",    "FT": "ツーシーム",     "SI": "シンカー",
    "FC": "カットボール",    "SL": "スライダー",     "ST": "スイーパー",
    "CU": "カーブ",          "KC": "ナックルカーブ",  "CH": "チェンジアップ",
    "FS": "スプリット",      "FO": "フォーク",        "KN": "ナックル",
    "SC": "スクリュー",      "EP": "エフェス",         "CS": "スローカーブ",
    "SV": "スラーブ",         "PO": "ピッチアウト",    "IN": "申告敬遠",
    "FA": "ファストボール",  "UN": "不明",
}

# launch_speed_angle == 6 → Barrel
BARREL_CODE  = 6
# Hard-Hit: EV >= 95 mph
HARD_HIT_EV  = 95.0

# %%
# ==================================================
# Section 4. ユーティリティ
# ==================================================

def make_output_dirs():
    for d in [RAW_DIR, GAMES_DM_DIR, GAMES_JSON_DIR]:
        Path(d).mkdir(parents=True, exist_ok=True)

def _pct(n, d, decimals=1):
    try:
        if d == 0 or pd.isna(d) or pd.isna(n): return np.nan
        return round(float(n) / float(d) * 100, decimals)
    except Exception:
        return np.nan

def _round(v, d=1):
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)): return np.nan
        return round(float(v), d)
    except Exception:
        return np.nan

def _mean(series, d=1):
    s = pd.to_numeric(series, errors="coerce").dropna()
    return _round(s.mean(), d) if len(s) > 0 else np.nan

def _max_val(series, d=1):
    s = pd.to_numeric(series, errors="coerce").dropna()
    return _round(s.max(), d) if len(s) > 0 else np.nan

def _safe(v, default=np.nan):
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)): return default
        return v
    except Exception:
        return default

def innings_to_float(ip_str) -> float:
    if pd.isna(ip_str): return 0.0
    s = str(ip_str)
    if "." in s:
        whole, frac = s.split(".")
        return int(whole) + int(frac) / 3
    return float(s)

def float_to_innings_str(f: float) -> str:
    whole = int(f)
    frac  = round((f - whole) * 3)
    return str(whole) if frac == 0 else f"{whole}.{frac}"

def get_pitch_type_jp(code) -> str:
    if pd.isna(code) or str(code) == "": return "不明"
    return PITCH_TYPE_MAP.get(str(code).upper(), str(code))

def get_on_base_situation(on_1b, on_2b, on_3b) -> str:
    b1 = pd.notna(on_1b); b2 = pd.notna(on_2b); b3 = pd.notna(on_3b)
    if b1 and b2 and b3: return "満塁"
    if b1 and b2:        return "走者1・2塁"
    if b1 and b3:        return "走者1・3塁"
    if b2 and b3:        return "走者2・3塁"
    if b1:               return "走者1塁"
    if b2:               return "走者2塁"
    if b3:               return "走者3塁"
    return "走者なし"

def is_scoring_position(on_1b, on_2b, on_3b) -> bool:
    return pd.notna(on_2b) or pd.notna(on_3b)

def is_in_zone(zone) -> bool:
    try:
        return int(zone) in range(1, 10)
    except Exception:
        return False

def zone_position(zone):
    try:
        z = int(zone)
        h = {1:"In",2:"中央",3:"Out",4:"In",5:"中央",6:"Out",7:"In",8:"中央",9:"Out"}
        v = {1:"高め",2:"高め",3:"高め",4:"真中",5:"真中",6:"真中",7:"低め",8:"低め",9:"低め"}
        return h.get(z), v.get(z)
    except Exception:
        return None, None

def event_to_jp(e, bb_type="") -> str:
    e  = str(e)  if pd.notna(e)       else ""
    bt = str(bb_type) if pd.notna(bb_type) else ""
    if e == "single":       return "単"
    if e == "double":       return "二"
    if e == "triple":       return "三"
    if e == "home_run":     return "本"
    if e == "walk":         return "四"
    if e == "intent_walk":  return "申四"
    if e == "hit_by_pitch": return "死"
    if e in ("strikeout","strikeout_double_play"): return "三振"
    if e == "sac_fly":      return "犠飛"
    if e == "sac_bunt":     return "犠打"
    if bt == "ground_ball": return "ゴロ"
    if bt == "fly_ball":    return "飛"
    if bt == "popup":       return "内飛"
    if bt == "line_drive":  return "ライナー"
    return "その他"

# %%
# ==================================================
# Section 5. Statcastデータ取得・前処理
# ==================================================

def _maybe_upgrade_statsapi_cache(date: str) -> None:
    """
    statsapi速報版(statsapi_{date}.xlsx)が存在し翌日以降の場合、
    pybaseball(Statcast)で完全版(statcast_{date}.xlsx)に差し替える。
    """
    today  = datetime.date.today()
    target = datetime.date.fromisoformat(date)
    if (today - target).days <= 1:
        return

    statsapi_path = os.path.join(RAW_DIR, f"statsapi_{date}.xlsx")
    statcast_path = os.path.join(RAW_DIR, f"statcast_{date}.xlsx")

    if not os.path.exists(statsapi_path):
        return  # statsapi速報版なし
    if os.path.exists(statcast_path):
        return  # すでにStatcast完全版あり

    statcast_date = (target - datetime.timedelta(days=1)).isoformat()
    print(f"  自動差し替え中: statsapi速報 → Statcast ({statcast_date})")
    try:
        df = statcast(start_dt=statcast_date, end_dt=statcast_date)
        if df is not None and not df.empty:
            df["game_date"] = date
            df.to_excel(statcast_path, index=False, engine="openpyxl")
            print(f"  ✓ Statcast差し替え完了: {len(df)}行")
        else:
            logger.warning("  データなし → statsapi版を継続使用")
    except Exception as e:
        logger.warning(f"  取得失敗 → statsapi版を継続使用: {e}")



def fetch_statcast(date: str) -> pd.DataFrame:
    """
    データ取得戦略（時差考慮・xlsx対応版）:

    ファイル命名規則:
      statcast_{date}.xlsx  : pybaseball(Statcast)完全版
      statsapi_{date}.xlsx  : statsapi速報版（当日/翌日）

    優先順位:
      1. statcast_{date}.xlsx が存在 → 読み込む
      2. statsapi_{date}.xlsx が存在 → 読み込む
      3. pybaseball で date-1日 のStatcastを取得 → statcast_{date}.xlsx に保存
      4. statsapi フォールバック → statsapi_{date}.xlsx に保存
    """
    statcast_path = os.path.join(RAW_DIR, f"statcast_{date}.xlsx")
    statsapi_path = os.path.join(RAW_DIR, f"statsapi_{date}.xlsx")
    statcast_date = (
        datetime.date.fromisoformat(date) - datetime.timedelta(days=1)
    ).isoformat()

    if os.path.exists(statcast_path):
        print(f"  キャッシュ: {os.path.basename(statcast_path)}")
        return pd.read_excel(statcast_path, engine="openpyxl")

    if os.path.exists(statsapi_path):
        print(f"  キャッシュ: {os.path.basename(statsapi_path)} (statsapi速報版)")
        return pd.read_excel(statsapi_path, engine="openpyxl")

    print(f"  Statcast取得: {statcast_date}")
    sdf = _try_fetch_statcast(statcast_date, date)
    if sdf is not None and not sdf.empty:
        Path(RAW_DIR).mkdir(parents=True, exist_ok=True)
        sdf.to_excel(statcast_path, index=False, engine="openpyxl")
        print(f"  保存: {os.path.basename(statcast_path)} ({len(sdf)}行)")
        return sdf

    if not _STATSAPI_AVAILABLE:
        logger.warning("MLB-StatsAPI未インストール: pip install MLB-StatsAPI")
        return pd.DataFrame()

    print(f"  statsapi速報取得: {date}")
    df = _fetch_statsapi(date)
    if df is not None and not df.empty:
        df["_statsapi_source"] = True
        Path(RAW_DIR).mkdir(parents=True, exist_ok=True)
        df.to_excel(statsapi_path, index=False, engine="openpyxl")
        print(f"  保存: {os.path.basename(statsapi_path)} ({len(df)}行)")
        return df

    print(f"  [WARN] データ取得失敗: {date}")
    return pd.DataFrame()

def _try_fetch_statcast(statcast_date: str, store_date: str):
    """pybaseballでstatcast_dateを取得しgame_dateをstore_dateに書き換えて返す"""
    try:
        print(f"  pybaseball: {statcast_date} 取得中...")
        df = statcast(start_dt=statcast_date, end_dt=statcast_date)
        if df is None or df.empty:
            print(f"  [WARN] データなし: {statcast_date}")
            return None
        df["game_date"] = store_date
        print(f"  取得完了: {len(df)}投球")
        return df
    except Exception as e:
        print(f"  [WARN] pybaseball取得失敗: {e}")
        return None



# ── statsapi description → Statcast description マッピング ──────────
STATSAPI_DESC_MAP = {
    # ストライク系
    "Called Strike":             "called_strike",
    "Swinging Strike":           "swinging_strike",
    "Swinging Strike (Blocked)": "swinging_strike_blocked",
    "Foul Tip":                  "foul_tip",
    "Foul":                      "foul",
    "Foul Bunt":                 "foul_bunt",
    "Missed Bunt":               "swinging_strike",
    # ボール系
    "Ball":                      "ball",
    "Ball In Dirt":              "ball",
    "Intent Ball":               "ball",
    "Pitchout":                  "pitchout",
    "Automatic Ball":            "ball",
    # インプレー系
    "In play, no out":           "hit_into_play_no_out",
    "In play, run(s)":           "hit_into_play_score",
    "In play, out(s)":           "hit_into_play",
    "Hit By Pitch":              "hit_into_play",
}

# ── statsapi pitch type description → Statcast pitch_type コード ─────
STATSAPI_PITCH_CODE_MAP = {
    "4-Seam Fastball":    "FF",
    "2-Seam Fastball":    "FT",
    "Sinker":             "SI",
    "Cutter":             "FC",
    "Slider":             "SL",
    "Sweeper":            "ST",
    "Curveball":          "CU",
    "Knuckle Curve":      "KC",
    "Slurve":             "SV",
    "Slow Curve":         "CS",
    "Changeup":           "CH",
    "Split-Finger":       "FS",
    "Forkball":           "FO",
    "Knuckleball":        "KN",
    "Screwball":          "SC",
    "Eephus":             "EP",
    "Fastball":           "FA",
    "Pitch Out":          "PO",
    "Intentional Ball":   "IN",
    "Other":              "UN",
    "Automatic Ball":     "UN",
}


def _fetch_statsapi(date: str) -> pd.DataFrame:
    """
    statsapi.get('game', hydrate='playByPlay') から1球単位のDataFrameを生成する。

    取得データ（Statcast互換）:
      - 投手/打者/イニング/球種/球速/ゾーン/コース/回転数/変化量
      - スイング/空振り/インプレー判定
      - 打球速度/打球角度/飛距離（hitData）
      - 走者状況/カウント/打席結果
    
    時差補正: date の1日前の試合を取得しgame_dateはdateとして保存
    """
    if not _STATSAPI_AVAILABLE:
        return pd.DataFrame()

    finished_statuses = {"Final", "Game Over", "Completed Early"}

    # 時差補正: dateの1日前の試合を取得
    statsapi_date = (
        datetime.date.fromisoformat(date) - datetime.timedelta(days=1)
    ).isoformat()


    try:
        games = _statsapi.schedule(date=statsapi_date)
    except Exception as e:
        logger.warning(f"  ⚠ schedule取得失敗: {e}")
        return pd.DataFrame()

    if not games:

        return pd.DataFrame()

    all_rows = []

    for game in games:
        status  = game.get("status", "")
        game_pk = game.get("game_id")

        if status not in finished_statuses:

            continue

        # playByPlayデータ取得
        try:
            gdata = _statsapi.get("game", {
                "gamePk": game_pk,
                "hydrate": "playByPlay,boxscore",
            })
        except Exception as e:
            logger.warning(f"[statsapi] game取得失敗 game_pk={game_pk}: {e}")
            continue

        # チーム情報
        gamedata   = gdata.get("gameData", {})
        livedata   = gdata.get("liveData",  {})
        teams_info = gamedata.get("teams", {})
        home_abbr  = teams_info.get("home", {}).get("abbreviation",
                     game.get("home_name", ""))
        away_abbr  = teams_info.get("away", {}).get("abbreviation",
                     game.get("away_name", ""))
        home_score_final = int(game.get("home_score", 0) or 0)
        away_score_final = int(game.get("away_score", 0) or 0)

        # boxscoreから投手実成績マップを構築（pit_id → {er, r, pitches, ip_str, ...}）
        pit_stats_map = {}
        boxscore_data = livedata.get("boxscore", {})
        for side_key in ("home", "away"):
            side_data = boxscore_data.get("teams", {}).get(side_key, {})
            for pid_str, pdata in side_data.get("players", {}).items():
                pstats = pdata.get("stats", {}).get("pitching", {})
                if not pstats:
                    continue
                pid_int = pdata.get("person", {}).get("id")
                if pid_int:
                    pit_stats_map[pid_int] = {
                        "er":      int(pstats.get("earnedRuns",      0) or 0),
                        "r":       int(pstats.get("runs",            0) or 0),
                        "pitches": int(pstats.get("numberOfPitches", 0) or 0),
                        "ip_str":  pstats.get("inningsPitched", "0") or "0",
                        "k":       int(pstats.get("strikeOuts",      0) or 0),
                        "bb":      int(pstats.get("baseOnBalls",     0) or 0),
                        "h":       int(pstats.get("hits",            0) or 0),
                        "hr":      int(pstats.get("homeRuns",        0) or 0),
                        "hbp":     int(pstats.get("hitBatsmen",      0) or 0),
                    }

        all_plays = livedata.get("plays", {}).get("allPlays", [])


        for play in all_plays:
            about   = play.get("about", {})
            matchup = play.get("matchup", {})
            result  = play.get("result", {})
            # 打席結果のRBI（打点）: result.rbi がplayByPlayで取得可能
            pa_rbi      = int(result.get("rbi", 0) or 0)

            inning      = about.get("inning", 1)
            half        = about.get("halfInning", "top")  # "top" or "bottom"
            inning_topbot = "Top" if half == "top" else "Bot"
            at_bat_num  = about.get("atBatIndex", 0) + 1

            pitcher_info = matchup.get("pitcher", {})
            batter_info  = matchup.get("batter",  {})
            pit_id   = pitcher_info.get("id")
            pit_name = pitcher_info.get("fullName", "")
            bat_id   = batter_info.get("id")
            bat_name = batter_info.get("fullName", "")

            stand    = matchup.get("batSide",   {}).get("code", None)   # L/R/S
            p_throws = matchup.get("pitchHand", {}).get("code", None)   # L/R

            # 打席最終イベントからeventsカラムを取得
            final_event = result.get("eventType", None)  # "single","strikeout"等

            # 走者状況（打席開始時）
            pre_on_1b = play.get("matchup", {}).get("postOnFirst",  {})   # fallback
            runners_before = {r["movement"]["originBase"]: r["details"]["runner"]["id"]
                              for r in play.get("runners", [])
                              if r["movement"].get("originBase")}
            on_1b = runners_before.get("1B", None)
            on_2b = runners_before.get("2B", None)
            on_3b = runners_before.get("3B", None)

            bat_score_start = about.get("halfInning") and (
                livedata.get("linescore", {}).get("teams", {})
                .get("away" if half=="top" else "home", {})
                .get("runs", 0)
            )

            for event in play.get("playEvents", []):
                pitch_data  = event.get("pitchData",  {})
                hit_data    = event.get("hitData",    {})
                details     = event.get("details",    {})
                count       = event.get("count",      {})
                pitch_idx   = event.get("index", 0)

                # 投球データがないイベント（代走・守備交代等）はスキップ
                if not pitch_data and event.get("type") != "pitch":
                    continue

                # 球種
                type_info  = details.get("type", {})
                pitch_desc = type_info.get("description", "")
                pitch_code_api = type_info.get("code", "")
                # descriptionからStatcastコードにマッピング
                pitch_type = STATSAPI_PITCH_CODE_MAP.get(pitch_desc)
                if not pitch_type and pitch_code_api:
                    pitch_type = pitch_code_api  # APIコードをそのまま使用
                if not pitch_type:
                    pitch_type = None

                # 球速（mph）
                release_speed = pitch_data.get("startSpeed")
                effective_speed = pitch_data.get("endSpeed")  # 近似

                # コース（plate_x, plate_z）
                coords = pitch_data.get("coordinates", {})
                plate_x = coords.get("pX")
                plate_z = coords.get("pZ")

                # ゾーン（1-14）
                zone = pitch_data.get("zone")

                # ストライクゾーン
                sz_top = pitch_data.get("strikeZoneTop")
                sz_bot = pitch_data.get("strikeZoneBottom")

                # 変化量（inches → feet、Statcastはfeet単位）
                breaks = pitch_data.get("breaks", {})
                pfx_x = _safe_div(breaks.get("pfxX"), 12)    # inches→feet
                pfx_z = _safe_div(breaks.get("pfxZ"), 12)
                ivb   = breaks.get("breakVerticalInduced")   # inches
                hb    = breaks.get("breakHorizontal")        # inches
                spin_rate = breaks.get("spinRate")
                spin_axis = breaks.get("spinDirection")

                # リリースポイント
                rel = pitch_data.get("coordinates", {})
                release_pos_x = rel.get("x0")
                release_pos_z = rel.get("z0")
                release_pos_y = rel.get("y0")
                release_ext   = pitch_data.get("extension")

                # 打球データ（インプレー時のみ）
                launch_speed   = hit_data.get("launchSpeed")    # mph
                launch_angle   = hit_data.get("launchAngle")
                hit_distance   = hit_data.get("totalDistance")  # feet
                # hardness: "Hard"/"Medium"/"Soft"
                hardness = hit_data.get("hardness", "")
                # launch_speed_angle（Barrel判定: 6=Barrel）
                trajectory = hit_data.get("trajectory", "")

                # description → Statcast description
                raw_desc   = details.get("description", "")
                statcast_desc = STATSAPI_DESC_MAP.get(raw_desc, raw_desc.lower().replace(" ", "_"))

                # is_strike / is_in_play
                is_strike  = bool(details.get("isStrike", False))
                is_in_play = bool(details.get("isInPlay", False))
                is_out     = bool(details.get("isOut", False))

                # eventsは最終投球のみ付与
                is_last_pitch = event.get("index", -1) == len(play.get("playEvents", [])) - 1
                event_val = final_event if is_last_pitch else None

                # bb_type推定
                bb_type = None
                if is_in_play:
                    traj = trajectory.lower()
                    if "fly" in traj or "popup" in traj:
                        bb_type = "fly_ball"
                    elif "line" in traj:
                        bb_type = "line_drive"
                    elif "ground" in traj:
                        bb_type = "ground_ball"
                    elif "bunt" in traj:
                        bb_type = "ground_ball"

                # launch_speed_angle（Barrel=6）
                lsa = None
                if launch_speed and launch_angle:
                    if launch_speed >= 98 and 26 <= launch_angle <= 30:
                        lsa = 6  # Barrel
                    elif launch_speed >= 95:
                        lsa = 5  # Solid
                    elif launch_speed >= 80:
                        lsa = 4

                # カウント
                balls   = count.get("balls",   0)
                strikes = count.get("strikes", 0)

                row = {
                    # 基本情報
                    "game_pk":        game_pk,
                    "game_date":      date,
                    "home_team":      home_abbr,
                    "away_team":      away_abbr,
                    "inning":         inning,
                    "inning_topbot":  inning_topbot,
                    "at_bat_number":  at_bat_num,
                    # 選手
                    "batter":         bat_id,
                    "batter_name":    bat_name,
                    "pitcher":        pit_id,
                    "player_name":    pit_name,
                    "stand":          stand,
                    "p_throws":       p_throws,
                    # 投球結果
                    "events":         event_val,
                    "description":    statcast_desc,
                    "bb_type":        bb_type,
                    # 得点情報（post_bat_scoreに打席のRBIを設定→打点計算に使用）
                    "bat_score":      0,
                    "post_bat_score": pa_rbi if is_last_pitch else 0,
                    "home_score":     home_score_final,
                    "away_score":     away_score_final,
                    "post_home_score": home_score_final,
                    "post_away_score": away_score_final,
                    # 走者
                    "on_1b":          on_1b,
                    "on_2b":          on_2b,
                    "on_3b":          on_3b,
                    # ゾーン
                    "zone":           zone,
                    "type":           ("S" if is_strike else ("X" if is_in_play else "B")),
                    # Statcast: 投球データ
                    "pitch_type":         pitch_type,
                    "release_speed":      release_speed,
                    "effective_speed":    effective_speed,
                    "release_spin_rate":  spin_rate,
                    "spin_axis":          spin_axis,
                    "plate_x":            plate_x,
                    "plate_z":            plate_z,
                    "sz_top":             sz_top,
                    "sz_bot":             sz_bot,
                    "pfx_x":              pfx_x,
                    "pfx_z":              pfx_z,
                    "release_pos_x":      release_pos_x,
                    "release_pos_z":      release_pos_z,
                    "release_pos_y":      release_pos_y,
                    "release_extension":  release_ext,
                    "arm_angle":          float("nan"),
                    # IVB/HB（inches→cm: 1inch=2.54cm）
                    "api_break_z_with_gravity": _safe_mul(ivb, 1/2.54) if ivb else float("nan"),
                    "api_break_x_arm":          _safe_mul(hb,  1/2.54) if hb  else float("nan"),
                    # Statcast: 打球データ
                    "launch_speed":        launch_speed,
                    "launch_angle":        launch_angle,
                    "hit_distance_sc":     hit_distance,
                    "launch_speed_angle":  lsa,
                    # Statcast: 期待値（statsapiにはなし）
                    "estimated_ba_using_speedangle":   float("nan"),
                    "estimated_woba_using_speedangle": float("nan"),
                    "estimated_slg_using_speedangle":  float("nan"),
                    "woba_value":          float("nan"),
                    "woba_denom":          float("nan"),
                    "babip_value":         float("nan"),
                    "iso_value":           float("nan"),
                    # バットスイング（statsapiにはなし）
                    "bat_speed":           float("nan"),
                    "swing_length":        float("nan"),
                    "attack_angle":        float("nan"),
                    "swing_path_tilt":     float("nan"),
                    # statsapi投手実成績補助カラム（boxscoreから取得）
                    "_statsapi_source":    True,
                    "_statsapi_pitches":   0,
                    "_pit_er":     pit_stats_map.get(pit_id, {}).get("er",      0),
                    "_pit_r":      pit_stats_map.get(pit_id, {}).get("r",       0),
                    "_pit_pitches":pit_stats_map.get(pit_id, {}).get("pitches", 0),
                    "_pit_ip_str": pit_stats_map.get(pit_id, {}).get("ip_str",  "0"),
                    "_pit_k":      pit_stats_map.get(pit_id, {}).get("k",       0),
                    "_pit_bb":     pit_stats_map.get(pit_id, {}).get("bb",      0),
                    "_pit_h":      pit_stats_map.get(pit_id, {}).get("h",       0),
                    "_pit_hr":     pit_stats_map.get(pit_id, {}).get("hr",      0),
                    "_pit_hbp":    pit_stats_map.get(pit_id, {}).get("hbp",     0),
                }
                all_rows.append(row)

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    print(f"  statsapi: {len(df)}投球 ({df['game_pk'].nunique()}試合)")
    return df


def _safe_div(v, d):
    """安全な除算（None/NaN対応）"""
    try:
        if v is None: return float("nan")
        return float(v) / d
    except Exception:
        return float("nan")


def _safe_mul(v, m):
    """安全な乗算（None/NaN対応）"""
    try:
        if v is None: return float("nan")
        return float(v) * m
    except Exception:
        return float("nan")


def _parse_batter_summary(summary: str):
    """
    statsapi batting summary をパース。
    例: "2-4, HR, 2 RBI, BB, 3 K"
    戻り値: (h, ab, hr, bb, rbi, so, hbp)
    """
    import re
    if not summary:
        return 0, 0, 0, 0, 0, 0, 0

    def _count(pattern, s):
        m = re.search(rf"(\d+)\s*{pattern}\b", s)
        if m:
            return int(m.group(1))
        return len(re.findall(rf"\b{pattern}\b", s))

    m_hab = re.search(r"(\d+)-(\d+)", summary)
    h  = int(m_hab.group(1)) if m_hab else 0
    ab = int(m_hab.group(2)) if m_hab else 0
    hr  = _count("HR",  summary)
    bb  = _count("BB",  summary)
    m_rbi = re.search(r"(\d+)\s*RBI", summary)
    rbi = int(m_rbi.group(1)) if m_rbi else 0
    so  = _count("K",   summary)
    hbp = _count("HBP", summary)
    return h, ab, hr, bb, rbi, so, hbp


def _expand_pa_events(h, hr, bb, hbp, so, ab, pa_count):
    """打席結果をリストに展開する（pa_count個）"""
    events = []
    events += ["home_run"]     * min(hr, pa_count)
    rem = pa_count - len(events)
    events += ["single"]       * min(max(0, h - hr), rem)
    rem = pa_count - len(events)
    events += ["walk"]         * min(bb, rem)
    rem = pa_count - len(events)
    events += ["hit_by_pitch"] * min(hbp, rem)
    rem = pa_count - len(events)
    events += ["strikeout"]    * min(so, rem)
    rem = pa_count - len(events)
    events += ["field_out"]    * rem
    return events[:pa_count]


def preprocess(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    df = df.copy()
    df["game_date"] = pd.to_datetime(df["game_date"])

    numeric_cols = [
        "release_speed","release_spin_rate","plate_x","plate_z",
        "launch_speed","launch_angle","sz_top","sz_bot",
        "pfx_x","pfx_z","release_pos_x","release_pos_z","release_pos_y",
        "release_extension","spin_axis","arm_angle",
        "estimated_ba_using_speedangle","estimated_woba_using_speedangle",
        "estimated_slg_using_speedangle","woba_value","woba_denom",
        "babip_value","iso_value","hit_distance_sc",
        "bat_speed","swing_length","attack_angle","swing_path_tilt",
        "effective_speed","api_break_z_with_gravity","api_break_x_arm",
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df["pitch_type_jp"] = df["pitch_type"].apply(get_pitch_type_jp)
    df["is_in_zone"]    = df["zone"].apply(is_in_zone)

    swing_descs = {"swinging_strike","swinging_strike_blocked","foul","foul_tip","foul_bunt",
                   "hit_into_play","hit_into_play_no_out","hit_into_play_score"}
    whiff_descs = {"swinging_strike","swinging_strike_blocked","foul_tip"}
    df["is_swing"]   = df["description"].isin(swing_descs)
    df["is_whiff"]   = df["description"].isin(whiff_descs)
    df["is_contact"] = df["is_swing"] & ~df["is_whiff"]

    df["situation"]     = df.apply(lambda r: get_on_base_situation(
        r.get("on_1b"), r.get("on_2b"), r.get("on_3b")), axis=1)
    df["is_scoring_pos"] = df.apply(lambda r: is_scoring_position(
        r.get("on_1b"), r.get("on_2b"), r.get("on_3b")), axis=1)

    df[["zone_horiz","zone_vert"]] = df["zone"].apply(lambda z: pd.Series(zone_position(z)))

    # Hard-Hit / Barrel
    df["is_hard_hit"] = (df["launch_speed"] >= HARD_HIT_EV) if "launch_speed" in df.columns else False
    df["is_barrel"]   = (df["launch_speed_angle"] == BARREL_CODE) if "launch_speed_angle" in df.columns else False

    # pfx feet → cm (1 foot = 30.48 cm)
    if "pfx_x" in df.columns: df["pfx_x_cm"] = df["pfx_x"] * 30.48
    if "pfx_z" in df.columns: df["pfx_z_cm"] = df["pfx_z"] * 30.48

    # ── 打者名カラムを追加（player_name は投手名なので別途取得）──
    df = _add_batter_names(df)

    return df

def _add_batter_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Statcastの player_name は投手名のため、
    batter ID → 打者名 のマップを playerid_reverse_lookup で取得し
    batter_name カラムとして追加する。
    取得失敗時は batter ID の文字列をフォールバックとして使用。
    """
    if "batter" not in df.columns:
        df["batter_name"] = ""
        return df

    batter_ids = df["batter"].dropna().unique().tolist()
    batter_ids = [int(x) for x in batter_ids if str(x).isdigit() or isinstance(x, (int, float))]

    name_map = {}
    try:
        print(f"  打者名解決: {len(batter_ids)}名")
        lookup = playerid_reverse_lookup(batter_ids, key_type="mlbam")
        for _, row in lookup.iterrows():
            mlbam_id = row.get("key_mlbam")
            first = str(row.get("name_first", "")).strip()
            last  = str(row.get("name_last",  "")).strip()
            if mlbam_id and (first or last):
                # "Last, First" 形式（Statcastの表示に合わせる）
                name_map[int(mlbam_id)] = f"{last}, {first}" if first else last

    except Exception as e:
        print(f"  [WARN] 打者名取得失敗: {e}")

    df["batter_name"] = df["batter"].apply(
        lambda x: name_map.get(int(x), str(int(x)) if pd.notna(x) else "")
        if pd.notna(x) else ""
    )
    return df

# %%
# ==================================================
# Section 6. スイング指標
# ==================================================

def calc_swing_metrics(g: pd.DataFrame) -> dict:
    total    = len(g)
    in_zone  = g["is_in_zone"].sum()
    out_zone = total - in_zone
    swings   = g["is_swing"].sum()
    whiffs   = g["is_whiff"].sum()
    contacts = g["is_contact"].sum()
    z_sw     = (g["is_in_zone"] & g["is_swing"]).sum()
    o_sw     = (~g["is_in_zone"] & g["is_swing"]).sum()
    return {
        "投球数":               total,
        "ゾーン内投球数":       int(in_zone),
        "ゾーン外投球数":       int(out_zone),
        "SW数":                 int(swings),
        "ゾーン内SW数":         int(z_sw),
        "ゾーン外SW数":         int(o_sw),
        "コンタクト数":         int(contacts),
        "空振り数":             int(whiffs),
        "空振り率":             _pct(whiffs, total),   # 空振り ÷ 全投球数
        "Z-Swing%":             _pct(z_sw, in_zone),
        "O-Swing%":             _pct(o_sw, out_zone),
        "Contact%":             _pct(contacts, swings),
        "whiff%":               _pct(whiffs, total),   # 空振り ÷ 全投球数
        "ゾーン内スイング率":   _pct(z_sw, in_zone),
        "ゾーン外スイング率":   _pct(o_sw, out_zone),
        "ゾーン率":             _pct(in_zone, total),
        "ストライク率":         _pct((g["type"]=="S").sum(), total) if "type" in g.columns else np.nan,
    }

def calc_zone_pcts(g: pd.DataFrame) -> dict:
    iz = g[g["is_in_zone"]]
    n  = len(iz)
    if n == 0:
        return {"In%":np.nan,"中央%":np.nan,"Out%":np.nan,"高め%":np.nan,"真中%":np.nan,"低め%":np.nan}
    return {
        "In%":   _pct((iz["zone_horiz"]=="In").sum(),   n),
        "中央%": _pct((iz["zone_horiz"]=="中央").sum(), n),
        "Out%":  _pct((iz["zone_horiz"]=="Out").sum(),  n),
        "高め%": _pct((iz["zone_vert"] =="高め").sum(), n),
        "真中%": _pct((iz["zone_vert"] =="真中").sum(), n),
        "低め%": _pct((iz["zone_vert"] =="低め").sum(), n),
    }

# %%
# ==================================================
# Section 7. 打球指標（MLB独自含む）
# ==================================================

def calc_batted_stats(g: pd.DataFrame) -> dict:
    batted = g[g["bb_type"].notna() & (g["bb_type"] != "")]
    total  = len(batted)
    gb  = (batted["bb_type"] == "ground_ball").sum()
    ld  = (batted["bb_type"] == "line_drive").sum()
    fb  = (batted["bb_type"] == "fly_ball").sum()
    pu  = (batted["bb_type"] == "popup").sum()
    hr  = (batted["events"]  == "home_run").sum()
    h   = batted["events"].isin(["single","double","triple","home_run"]).sum()

    # EV / LA / 飛距離
    ev   = batted["launch_speed"].dropna()   if "launch_speed"   in batted.columns else pd.Series(dtype=float)
    la   = batted["launch_angle"].dropna()   if "launch_angle"   in batted.columns else pd.Series(dtype=float)
    dist = batted["hit_distance_sc"].dropna() if "hit_distance_sc" in batted.columns else pd.Series(dtype=float)

    hard_hit_n   = int((ev >= HARD_HIT_EV).sum()) if len(ev) > 0 else 0
    barrel_n     = int(batted["is_barrel"].sum()) if "is_barrel" in batted.columns else 0

    # xBA / xwOBA / xSLG / wOBA / BABIP / ISO (打席最終行から)
    def _col_mean(col, d=3):
        return _round(batted[col].dropna().mean(), d) if col in batted.columns else np.nan

    return {
        "全打球数":     total,
        "GB": int(gb), "LD": int(ld), "FB": int(fb), "IFFB": int(pu),
        "HR": int(hr), "H":  int(h),  "判断不可打球": 0,
        "GB%":   _pct(gb, total), "LD%": _pct(ld, total), "FB%":  _pct(fb, total),
        "IFFB%": _pct(pu, fb)    if fb  > 0 else np.nan,
        "HR%":   _pct(hr, fb)    if fb  > 0 else np.nan,
        # ── MLB独自: 打球 ──────────────────────────────────────
        "平均打球速度EV(km/h)": mph2kmh(_mean(ev)),
        "最高打球速度EV(km/h)": mph2kmh(_max_val(ev)),
        "Hard-Hit数":           hard_hit_n,
        "Hard-Hit%":            _pct(hard_hit_n, len(ev)) if len(ev) > 0 else np.nan,
        "平均打球角度LA(°)":    _mean(la),
        "平均飛距離(ft)":       _mean(dist),
        "最大飛距離(ft)":       _max_val(dist),
        "Barrel数":             barrel_n,
        "Barrel%":              _pct(barrel_n, total),
        # ── MLB独自: 期待値・価値 ──────────────────────────────
        "xBA":    _col_mean("estimated_ba_using_speedangle"),
        "xwOBA":  _col_mean("estimated_woba_using_speedangle"),
        "xSLG":   _col_mean("estimated_slg_using_speedangle"),
        "wOBA":   _col_mean("woba_value"),
        "BABIP":  _col_mean("babip_value"),
        "ISO":    _col_mean("iso_value"),
    }

# %%
# ==================================================
# Section 8. バットスイング指標（打者専用）
# ==================================================

def calc_bat_swing_stats(g: pd.DataFrame) -> dict:
    sw = g[g["is_swing"]]
    return {
        "平均バットスピード(km/h)": mph2kmh(_mean(sw["bat_speed"]))       if "bat_speed"       in sw.columns else np.nan,
        "平均スイング軌跡長(ft)":    _mean(sw["swing_length"])     if "swing_length"     in sw.columns else np.nan,
        "平均アタックアングル(°)":   _mean(sw["attack_angle"])     if "attack_angle"     in sw.columns else np.nan,
        "平均スイングパスチルト(°)": _mean(sw["swing_path_tilt"]) if "swing_path_tilt" in sw.columns else np.nan,
    }

# %%
# ==================================================
# Section 9. 投球変化量・リリース指標（投手・球種専用）
# ==================================================

def calc_pitch_movement_stats(g: pd.DataFrame) -> dict:
    # VAA（Vertical Approach Angle）: ホームプレート到達時の垂直進入角度（度）
    # マイナスが急角度。vz0/vy0/az/ayから計算
    def _calc_vaa(row):
        try:
            vy0 = float(row['vy0']); vz0 = float(row['vz0'])
            ay  = float(row['ay']);  az  = float(row['az'])
            y0  = float(row.get('release_pos_y', 55))
            t   = (-vy0 - np.sqrt(vy0**2 - 2*ay*(y0 - 1.417))) / ay
            vz_p = vz0 + az * t
            vy_p = vy0 + ay * t
            return np.degrees(np.arctan(vz_p / abs(vy_p)))
        except Exception:
            return np.nan

    vaa_vals = g.apply(_calc_vaa, axis=1) \
        if all(c in g.columns for c in ['vy0','vz0','ay','az']) \
        else pd.Series(dtype=float)

    # 回転効率（近似値）: |cos(spin_axis - 180)| × 100
    active_spin_vals = g['spin_axis'].apply(
        lambda x: abs(np.cos(np.radians(float(x) - 180))) * 100 if pd.notna(x) else np.nan
    ) if 'spin_axis' in g.columns else pd.Series(dtype=float)

    return {
        # 回転数・回転軸・回転効率
        "平均回転数(rpm)":    _mean(g["release_spin_rate"]) if "release_spin_rate" in g.columns else np.nan,
        "最高回転数(rpm)":    _max_val(g["release_spin_rate"]) if "release_spin_rate" in g.columns else np.nan,
        "回転軸(°)":          _round(_mean(g["spin_axis"]), 0) if "spin_axis" in g.columns else np.nan,
        "回転効率%(近似)":    _round(active_spin_vals.mean(), 1) if len(active_spin_vals.dropna()) > 0 else np.nan,
        # VAA
        "VAA(°)":             _round(vaa_vals.mean(), 2) if len(vaa_vals.dropna()) > 0 else np.nan,
        # 変化量 cm
        "横変化量(pfx_x cm)": _mean(g["pfx_x_cm"]) if "pfx_x_cm" in g.columns else np.nan,
        "縦変化量(pfx_z cm)": _mean(g["pfx_z_cm"]) if "pfx_z_cm" in g.columns else np.nan,
        # Induced Break (inches → cm)
        "IVB(cm)":    _round(g["api_break_z_with_gravity"].dropna().mean() * 2.54, 1)
                        if "api_break_z_with_gravity" in g.columns else np.nan,
        "HB arm(cm)": _round(g["api_break_x_arm"].dropna().mean() * 2.54, 1)
                        if "api_break_x_arm" in g.columns else np.nan,
        # リリースポイント
        "リリース高さ(ft)":   _mean(g["release_pos_z"])   if "release_pos_z"   in g.columns else np.nan,
        "リリース横位置(ft)": _mean(g["release_pos_x"])   if "release_pos_x"   in g.columns else np.nan,
        "Extension(m)":       _round(_mean(g["release_extension"]) * 0.3048, 2)
                                if "release_extension" in g.columns else np.nan,
        "アーム角度(°)":      _mean(g["arm_angle"])        if "arm_angle"        in g.columns else np.nan,
        "実効球速(km/h)":     mph2kmh(_mean(g["effective_speed"])) if "effective_speed" in g.columns else np.nan,
    }

# %%
# ==================================================
# Section 10. 打者成績計算（MLB指標付き）
# ==================================================

def calc_batter_pa_stats(g: pd.DataFrame) -> dict:
    pa_df = g.groupby("at_bat_number").first().reset_index()

    pa  = len(pa_df)
    AB_EVENTS = {
        "single","double","triple","home_run","strikeout","strikeout_double_play",
        "field_out","flyout","lineout","groundout","grounded_into_double_play",
        "force_out","double_play","fielders_choice","fielders_choice_out","other_out",
    }
    ab      = len(pa_df[pa_df["events"].isin(AB_EVENTS)])
    hits    = pa_df["events"].isin(["single","double","triple","home_run"]).sum()
    hr      = (pa_df["events"] == "home_run").sum()
    doubles = (pa_df["events"] == "double").sum()
    triples = (pa_df["events"] == "triple").sum()
    singles = hits - hr - doubles - triples
    xbh     = doubles + triples
    bb      = pa_df["events"].isin(["walk","intent_walk"]).sum()
    hbp     = (pa_df["events"] == "hit_by_pitch").sum()
    so      = pa_df["events"].isin(["strikeout","strikeout_double_play"]).sum()
    sac_f   = pa_df["events"].isin(["sac_fly","sac_fly_double_play"]).sum()
    sac_b   = (pa_df["events"] == "sac_bunt").sum()
    other   = pa_df["events"].isin(["fielders_choice","fielders_choice_out","catcher_interf"]).sum()
    fly_out = pa_df["bb_type"].isin(["fly_ball","popup"]).sum()
    line_out= (pa_df["bb_type"] == "line_drive").sum()
    gro_out = (pa_df["bb_type"] == "ground_ball").sum()

    ba   = _round(hits/ab, 3) if ab > 0 else np.nan
    tb   = singles + doubles*2 + triples*3 + hr*4
    slg  = _round(tb/ab, 3)   if ab > 0 else np.nan
    obpd = ab + bb + hbp + sac_f
    obp  = _round((hits+bb+hbp)/obpd, 3) if obpd > 0 else np.nan
    ops  = _round(_safe(obp,0) + _safe(slg,0), 3)
    rbi  = max(0, int((pa_df["post_bat_score"].fillna(0)-pa_df["bat_score"].fillna(0)).clip(lower=0).sum()))

    ab_result_str = ",".join(
        event_to_jp(r["events"], r.get("bb_type",""))
        for _, r in pa_df.iterrows() if pd.notna(r.get("events"))
    )

    sw  = calc_swing_metrics(g)
    bd  = calc_batted_stats(g)
    zp  = calc_zone_pcts(g)
    bat = calc_bat_swing_stats(g)

    # xBA/xwOBA (打席最終行の平均)
    xba   = _round(pa_df["estimated_ba_using_speedangle"].dropna().mean(),   3) if "estimated_ba_using_speedangle"   in pa_df.columns else np.nan
    xwoba = _round(pa_df["estimated_woba_using_speedangle"].dropna().mean(), 3) if "estimated_woba_using_speedangle" in pa_df.columns else np.nan

    return {
        "打席別結果": ab_result_str,
        "OPS": ops, "出塁率": obp, "長打率": slg, "打率": ba,
        "打点": int(rbi), "盗塁": 0,
        "打席": pa, "打数": ab,
        "安打": int(hits), "本塁打": int(hr), "長打": int(xbh), "単打": int(singles),
        "四球": int(bb), "死球": int(hbp), "三振": int(so),
        "フライアウト(犠牲フライ含む)": int(fly_out),
        "ライナーアウト": int(line_out), "ゴロアウト": int(gro_out),
        "その他(犠打、失策、野選)": int(sac_b+other),
        "本塁打割合":        _pct(hr,      hits) if hits > 0 else np.nan,
        "長打割合":          _pct(xbh,     hits) if hits > 0 else np.nan,
        "単打割合":          _pct(singles, hits) if hits > 0 else np.nan,
        "フライアウト割合":  _pct(fly_out,  ab),
        "ライナーアウト割合":_pct(line_out, ab),
        "ゴロアウト割合":    _pct(gro_out,  ab),
        "K%": _pct(so, pa), "BB%": _pct(bb, pa),
        # スイング
        "球数":           sw["投球数"],
        "ゾーン内SW数":   sw["ゾーン内SW数"],   "ゾーン内投球数": sw["ゾーン内投球数"],
        "ゾーン外SW数":   sw["ゾーン外SW数"],   "ゾーン外投球数": sw["ゾーン外投球数"],
        "SW数":           sw["SW数"],
        "コンタクト数":   sw["コンタクト数"],   "空振り数":       sw["空振り数"],
        "Z-Swing%":       sw["Z-Swing%"],        "O-Swing%":       sw["O-Swing%"],
        "Contact%":       sw["Contact%"],         "whiff%":         sw["whiff%"],
        # ── MLB独自: 打球 ──────────────────────────────────────
        "平均打球速度EV(km/h)":  bd["平均打球速度EV(km/h)"],
        "最高打球速度EV(km/h)":  bd["最高打球速度EV(km/h)"],
        "Hard-Hit%":            bd["Hard-Hit%"],
        "平均打球角度LA(°)":    bd["平均打球角度LA(°)"],
        "平均飛距離(ft)":       bd["平均飛距離(ft)"],
        "最大飛距離(ft)":       bd["最大飛距離(ft)"],
        "Barrel数":             bd["Barrel数"],
        "Barrel%":              bd["Barrel%"],
        "xBA":   xba,
        "xwOBA": xwoba,
        # ── MLB独自: バットスイング ────────────────────────────
        "平均バットスピード(km/h)":   bat["平均バットスピード(km/h)"],
        "平均スイング軌跡長(ft)":    bat["平均スイング軌跡長(ft)"],
        "平均アタックアングル(°)":   bat["平均アタックアングル(°)"],
        "平均スイングパスチルト(°)": bat["平均スイングパスチルト(°)"],
    }

# %%
# ==================================================
# Section 11. 投手成績計算（MLB指標付き）
# ==================================================

def calc_pitcher_stats(g: pd.DataFrame) -> dict:
    pa_df = g.groupby("at_bat_number").first().reset_index()

    # statsapiデータの場合は _statsapi_pitches の合計を投球数とする
    if "_statsapi_pitches" in g.columns and g["_statsapi_pitches"].sum() > 0:
        pitches = int(g["_statsapi_pitches"].sum())
    else:
        pitches = len(g)
    tbf     = len(pa_df)
    h    = pa_df["events"].isin(["single","double","triple","home_run"]).sum()
    hr   = (pa_df["events"] == "home_run").sum()
    bb   = pa_df["events"].isin(["walk","intent_walk"]).sum()
    hbp  = (pa_df["events"] == "hit_by_pitch").sum()
    so   = pa_df["events"].isin(["strikeout","strikeout_double_play"]).sum()
    runs = (pa_df["post_bat_score"].fillna(0)-pa_df["bat_score"].fillna(0)).clip(lower=0).sum()

    outs_r = pa_df["events"].isin([
        "field_out","flyout","lineout","groundout","grounded_into_double_play",
        "force_out","double_play","strikeout","strikeout_double_play",
        "sac_fly","sac_fly_double_play","sac_bunt","other_out","fielders_choice_out",
    ]).sum()
    dp = pa_df["events"].isin(["grounded_into_double_play","double_play",
                                "strikeout_double_play","sac_fly_double_play"]).sum()
    ip_str = float_to_innings_str((int(outs_r)+int(dp)) / 3)

    sw = calc_swing_metrics(g)
    bd = calc_batted_stats(g)
    zp = calc_zone_pcts(g)
    mv = calc_pitch_movement_stats(g)

    xwoba_against = _round(pa_df["estimated_woba_using_speedangle"].dropna().mean(), 3) \
                    if "estimated_woba_using_speedangle" in pa_df.columns else np.nan
    woba_against  = _round(pa_df["woba_value"].dropna().mean(), 3) \
                    if "woba_value" in pa_df.columns else np.nan

    BATTED_KEYS = ["全打球数","GB","LD","FB","IFFB","判断不可打球","HR","H",
                   "GB%","LD%","FB%","IFFB%","HR%",
                   "平均打球速度EV(km/h)","最高打球速度EV(km/h)","Hard-Hit%",
                   "Barrel%","xBA","xwOBA","xSLG","wOBA","BABIP","ISO"]
    SWING_KEYS  = ["空振り率","ゾーン内スイング率","ゾーン外スイング率","ゾーン率","ストライク率",
                   "ゾーン内SW数","ゾーン内投球数","ゾーン外SW数","ゾーン外投球数","SW数","コンタクト数","空振り数"]

    # _pit_r/_pit_er補助カラムがあればそちらを優先（statsapiデータの実値）
    if "_pit_r" in g.columns and g["_pit_r"].sum() > 0:
        runs = int(g["_pit_r"].iloc[0])
    if "_pit_er" in g.columns and g["_pit_er"].sum() > 0:
        earned = int(g["_pit_er"].iloc[0])
    else:
        earned = int(runs)

    return {
        "投球回": ip_str, "投球数": pitches, "対戦打者数": int(tbf),
        "失点": int(runs), "自責点": earned,
        "被安打": int(h), "被本塁打": int(hr),
        "与四球": int(bb), "与死球": int(hbp), "奪三振": int(so),
        "K%": _pct(so, tbf), "BB%": _pct(bb, tbf),
        "K-BB%": _round(_safe(_pct(so,tbf),0) - _safe(_pct(bb,tbf),0), 1),
        **{k: bd[k] for k in BATTED_KEYS if k in bd},
        **{k: sw[k] for k in SWING_KEYS},
        **zp,
        # ── MLB独自: 球速・変化量 ──────────────────────────────
        **mv,
        "xwOBA被打": xwoba_against,
        "wOBA被打":  woba_against,
    }

# %%
# ==================================================
# Section 12. 球種別集計（MLB指標付き）
# ==================================================

def calc_pitch_type_stats(g: pd.DataFrame, pitch_col: str = "pitch_type") -> pd.DataFrame:
    rows = []
    total_pitches = len(g)
    for code, pg in g.groupby(pitch_col):
        sw = calc_swing_metrics(pg)
        bd = calc_batted_stats(pg)
        zp = calc_zone_pcts(pg)
        mv = calc_pitch_movement_stats(pg)
        n  = len(pg)
        rows.append({
            "球種名": get_pitch_type_jp(code), "球種コード": str(code),
            "投球数": n, "投球割合%": _pct(n, total_pitches),
            "平均球速": mph2kmh(_mean(pg["release_speed"])), "最高球速": mph2kmh(_max_val(pg["release_speed"])),
            **{k: bd[k] for k in ["全打球数","GB","LD","FB","IFFB","判断不可打球","HR","H",
                                   "GB%","LD%","FB%","IFFB%","HR%",
                                   "平均打球速度EV(km/h)","Hard-Hit%","Barrel%","xwOBA"] if k in bd},
            **{k: sw[k] for k in ["空振り率","ゾーン内スイング率","ゾーン外スイング率","ゾーン率",
                                   "ストライク率","ゾーン内SW数","ゾーン内投球数","ゾーン外SW数",
                                   "ゾーン外投球数","SW数","コンタクト数","空振り数"]},
            **zp,
            **mv,
        })
    return pd.DataFrame(rows)

# %%
# ==================================================
# Section 13. 状況別ヘルパー
# ==================================================

def _iter_situations(g: pd.DataFrame):
    yield "走者なし",  g[g["situation"] == "走者なし"]
    yield "走者あり",  g[g["situation"] != "走者なし"]
    yield "得点圏",    g[g["is_scoring_pos"] == True]
    yield "非得点圏",  g[g["is_scoring_pos"] == False]
    for lbl in ["走者1塁","走者2塁","走者3塁","走者1・2塁","走者1・3塁","走者2・3塁","満塁"]:
        yield lbl, g[g["situation"] == lbl]

# %%
# ==================================================
# Section 14. 試合概要シート
# ==================================================

def build_game_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for gid, gdf in df.groupby("game_pk"):
        row0 = gdf.iloc[0]
        def inning_scores(side):
            sub = gdf[gdf["inning_topbot"] == side]
            out = []
            for inn in sorted(sub["inning"].unique()):
                idf = sub[sub["inning"] == inn]
                r = (idf.groupby("at_bat_number").first()["post_bat_score"].fillna(0)
                     - idf.groupby("at_bat_number").first()["bat_score"].fillna(0)
                    ).clip(lower=0).sum()
                out.append(str(int(r)))
            return ",".join(out)

        home_score = int(_safe(gdf[gdf["inning_topbot"]=="Bot"]["post_home_score"].max(), 0))
        away_score = int(_safe(gdf[gdf["inning_topbot"]=="Top"]["post_away_score"].max(), 0))
        home_hits  = gdf[(gdf["inning_topbot"]=="Bot") & gdf["events"].isin(["single","double","triple","home_run"])]["at_bat_number"].nunique()
        away_hits  = gdf[(gdf["inning_topbot"]=="Top") & gdf["events"].isin(["single","double","triple","home_run"])]["at_bat_number"].nunique()

        rows.append({
            "試合ID": str(gid), "試合日": str(row0["game_date"])[:10],
            "ホームチーム": row0["home_team"], "アウェイチーム": row0["away_team"],
            "ホーム得点": home_score, "アウェイ得点": away_score,
            "球場": "", "開始時間": "", "試合時間": "", "観客数": "",
            "ホームイニング得点": inning_scores("Bot"),
            "アウェイイニング得点": inning_scores("Top"),
            "ホーム安打": int(home_hits), "アウェイ安打": int(away_hits),
            "ホームエラー": 0, "アウェイエラー": 0,
            "MVP選手名": "", "MVP成績": "", "ハイライト": "",
        })
    return pd.DataFrame(rows)

# %%
# ==================================================
# Section 15. 活躍選手（Claude API + prompts/活躍選手_MLB.txt）
# ==================================================

# ──────────────────────────────────────────────────
# 15-a. データ整形ヘルパー（LLMへ渡すテキスト生成）
# ──────────────────────────────────────────────────

def _build_score_plays(df: pd.DataFrame) -> str:
    """
    スコアプレー: 得点が動いた打席を時系列で抽出
    出力例:
      [Game 700001 NYY vs BOS]
      Top 3rd: Judge (NYY) - home_run (2 RBI) → NYY 0, BOS 2
      Bot 5th: Devers (BOS) - single (1 RBI) → NYY 2, BOS 3
    """
    lines = []
    for gid, gdf in df.groupby("game_pk"):
        home = gdf.iloc[0]["home_team"]
        away = gdf.iloc[0]["away_team"]
        lines.append(f"\n[Game {gid}: {away} @ {home}]")

        pa_last = (gdf.sort_values(["inning","inning_topbot","at_bat_number"])
                      .groupby("at_bat_number").first().reset_index())

        for _, row in pa_last.iterrows():
            rbi = int(max(0, (row.get("post_bat_score",0) or 0) - (row.get("bat_score",0) or 0)))
            if rbi == 0 and row.get("events") not in ("home_run",): 
                # 得点が動いた打席、またはHRのみ抽出
                continue
            side   = row.get("inning_topbot","")
            inn    = int(row.get("inning", 0))
            name   = row.get("batter_name") or row.get("player_name","Unknown")
            event  = row.get("events","")
            h_score= int(row.get("post_home_score",0) or 0)
            a_score= int(row.get("post_away_score",0) or 0)
            bat_team = away if side == "Top" else home
            inning_label = f"{'Top' if side=='Top' else 'Bot'} {inn}th"
            rbi_str = f" ({rbi} RBI)" if rbi > 0 else ""
            lines.append(f"  {inning_label}: {name} ({bat_team}) - {event}{rbi_str} → {away} {a_score}, {home} {h_score}")
    return "\n".join(lines) if lines else "(no scoring plays)"


def _build_batter_summary(df: pd.DataFrame) -> str:
    """
    野手成績サマリー: 試合ごと・打者ごとの打席結果
    出力例:
      [Game 700001: BOS @ NYY]
      NYY (Home):
        Judge      : 4AB 2H 1HR 3RBI  BB:1  K:1
        ...
      BOS (Away):
        ...
    """
    lines = []
    AB_EVENTS = {
        "single","double","triple","home_run","strikeout","strikeout_double_play",
        "field_out","flyout","lineout","groundout","grounded_into_double_play",
        "force_out","double_play","fielders_choice","fielders_choice_out","other_out",
    }
    for gid, gdf in df.groupby("game_pk"):
        home = gdf.iloc[0]["home_team"]
        away = gdf.iloc[0]["away_team"]
        lines.append(f"\n[Game {gid}: {away} @ {home}]")

        for side_label, side_code in [(f"{home} (Home)","Bot"),(f"{away} (Away)","Top")]:
            lines.append(f"  {side_label}:")
            side_df = gdf[gdf["inning_topbot"] == side_code]
            for bid, bg in side_df.groupby("batter"):
                pa_last = bg.groupby("at_bat_number").first()
                name= bg.iloc[0].get("batter_name") or bg.iloc[0].get("player_name", str(bid))
                pa  = len(pa_last)
                ab  = pa_last["events"].isin(AB_EVENTS).sum()
                h   = pa_last["events"].isin(["single","double","triple","home_run"]).sum()
                hr  = (pa_last["events"] == "home_run").sum()
                rbi = max(0, int((pa_last["post_bat_score"].fillna(0)-pa_last["bat_score"].fillna(0)).clip(lower=0).sum()))
                bb  = pa_last["events"].isin(["walk","intent_walk"]).sum()
                k   = pa_last["events"].isin(["strikeout","strikeout_double_play"]).sum()
                sb  = pa_last["events"].isin(["single","double","triple"]).sum()  # extra bases

                # 打席結果文字列
                results = [event_to_jp(r["events"], r.get("bb_type",""))
                           for _, r in pa_last.iterrows() if pd.notna(r.get("events"))]
                result_str = "/".join(results)

                stat = f"{pa}PA {ab}AB {h}H"
                if hr:  stat += f" {hr}HR"
                if rbi: stat += f" {rbi}RBI"
                if bb:  stat += f" BB:{bb}"
                if k:   stat += f" K:{k}"

                # HR打った相手投手名（先頭1本のみ）
                hr_rows = pa_last[pa_last["events"]=="home_run"]
                if not hr_rows.empty:
                    hr_pitcher_id = hr_rows.iloc[0].get("pitcher")
                    hr_pitcher_name = df[df["pitcher"]==hr_pitcher_id]["player_name"].iloc[0] \
                                      if hr_pitcher_id and not df[df["pitcher"]==hr_pitcher_id].empty else "?"
                    stat += f" [HR off {hr_pitcher_name}]"

                lines.append(f"    {name:<20s}: {stat}  ({result_str})")
    return "\n".join(lines)


def _build_pitcher_summary(df: pd.DataFrame) -> str:
    """
    投手成績サマリー: 試合ごと・投手ごとの登板結果
    出力例:
      [Game 700001: BOS @ NYY]
      NYY (Home - Pitchers):
        Cole (SP): 7.0IP 1ER 10K 2BB  vs BOS
    """
    lines = []
    OUTS_EVENTS = {
        "field_out","flyout","lineout","groundout","grounded_into_double_play",
        "force_out","double_play","strikeout","strikeout_double_play",
        "sac_fly","sac_fly_double_play","sac_bunt","other_out","fielders_choice_out",
    }
    DP_EVENTS = {"grounded_into_double_play","double_play","strikeout_double_play","sac_fly_double_play"}

    for gid, gdf in df.groupby("game_pk"):
        home = gdf.iloc[0]["home_team"]
        away = gdf.iloc[0]["away_team"]
        lines.append(f"\n[Game {gid}: {away} @ {home}]")

        # 投手は「表の攻撃=ホーム投手」「裏=アウェイ投手」
        for side_label, pit_side, opp in [
            (f"{home} (Home Pitchers)", "Top", away),
            (f"{away} (Away Pitchers)", "Bot", home),
        ]:
            lines.append(f"  {side_label}:")
            pit_df = gdf[gdf["inning_topbot"] == pit_side]
            for pid, pg in pit_df.groupby("pitcher"):
                name    = pg.iloc[0].get("player_name", str(pid))
                pa_last = pg.groupby("at_bat_number").first()
                pitches = len(pg)
                tbf     = len(pa_last)
                h   = pa_last["events"].isin(["single","double","triple","home_run"]).sum()
                hr  = (pa_last["events"] == "home_run").sum()
                bb  = pa_last["events"].isin(["walk","intent_walk"]).sum()
                hbp = (pa_last["events"] == "hit_by_pitch").sum()
                k   = pa_last["events"].isin(["strikeout","strikeout_double_play"]).sum()
                er  = max(0, int((pa_last["post_bat_score"].fillna(0)-pa_last["bat_score"].fillna(0)).clip(lower=0).sum()))
                outs_r = int(pa_last["events"].isin(OUTS_EVENTS).sum())
                dp     = int(pa_last["events"].isin(DP_EVENTS).sum())
                ip     = float_to_innings_str((outs_r+dp)/3)
                role   = "SP" if pg["inning"].min()==1 else "RP"

                stat = f"{ip}IP {er}ER {k}K {bb}BB {h}H"
                if hr: stat += f" {hr}HR"
                stat += f" {pitches}P  vs {opp}"
                lines.append(f"    {name:<20s} ({role}): {stat}")
    return "\n".join(lines)


# ──────────────────────────────────────────────────
# 15-b. プロンプト読み込みヘルパー
# ──────────────────────────────────────────────────

def _load_system_prompts(prompts_path: str = str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt")) -> str:
    """
    prompts/活躍選手_MLB.txt からシステムプロンプトを読み込む。
    ファイルは以下いずれかの形式に対応:
      (A) system_prompts = \"\"\"...\"\"\"\n  （Pythonコード形式）
      (B) プレーンテキスト
    """
    if not os.path.exists(prompts_path):
        logger.warning(f"プロンプトファイルが見つかりません: {prompts_path}")
        return ""
    with open(prompts_path, encoding="utf-8") as f:
        content = f.read()
    # Pythonコード形式 (system_prompts = """...""") をパース
    import re
    m = re.search(r'system_prompts\s*=\s*"""(.*?)"""', content, re.DOTALL)
    if m:
        return m.group(1).strip()
    m = re.search(r"system_prompts\s*=\s*'''(.*?)'''", content, re.DOTALL)
    if m:
        return m.group(1).strip()
    # プレーンテキストとして使用
    return content.strip()


# ──────────────────────────────────────────────────
# 15-c. Claude API 呼び出し
# ──────────────────────────────────────────────────

def _call_gemini(system_prompts: str, user_message: str) -> str:
    """Gemini API を呼び出す。失敗時は "" を返す。"""
    import os as _os
    api_key = _os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        logger.warning("GEMINI_API_KEY が未設定です。")
        return ""
    if not _GENAI_AVAILABLE:
        logger.warning("google-generativeai 未インストール。pip install google-generativeai")
        return ""
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(
            model_name=GEMINI_MODEL,
            system_instruction=system_prompts,
            generation_config=genai.GenerationConfig(
                temperature=GEMINI_TEMPERATURE,
                max_output_tokens=GEMINI_MAX_OUTPUT_TOKENS,
            ),
        )
        response = model.generate_content(user_message)
        return response.text.strip()
    except Exception as e:
        print(f"  [ERROR] Gemini API エラー: {type(e).__name__}: {e}")
        logger.error(f"Gemini API エラー: {e}")
        return ""


def _call_claude(system_prompts: str, user_message: str) -> str:
    """Claude API を呼び出す。失敗時は "" を返す。"""
    import os as _os
    api_key = _os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        logger.warning("GEMINI_API_KEY が未設定です。")
        return ""
    if not _REQUESTS_AVAILABLE:
        logger.warning("requests 未インストール。pip install requests")
        return ""
    headers = {
        "x-api-key":         api_key,
        "anthropic-version": "2023-06-01",
        "content-type":      "application/json",
    }
    payload = {
        "model":      CLAUDE_MODEL,
        "max_tokens": CLAUDE_MAX_TOKENS,
        "system":     system_prompts,
        "messages":   [{"role": "user", "content": user_message}],
    }
    try:
        resp = _requests.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers, json=payload, timeout=60,
        )
        resp.raise_for_status()
        return resp.json()["content"][0]["text"].strip()
    except Exception as e:
        logger.error(f"Claude API エラー: {e}")
        return ""


def _call_llm(system_prompts: str, user_message: str) -> str:
    """
    LLM_PROVIDER 設定に従い Gemini / Claude を呼び出す。
      "gemini" : Gemini 優先、失敗時 Claude にフォールバック
      "claude" : Claude のみ
      "auto"   : GEMINI_API_KEY があれば Gemini、なければ Claude
    """
    import os as _os
    provider = LLM_PROVIDER.lower()

    use_gemini = (
        provider == "gemini"
        or (provider == "auto" and bool(_os.environ.get("GEMINI_API_KEY")))
    )

    if use_gemini:
        result = _call_gemini(system_prompts, user_message)
        if result:
            return result
        print("  [WARN] Gemini 失敗。活躍選手選出をスキップします。")
        logger.warning("Gemini 失敗。活躍選手選出をスキップします。")
        return ""
    else:
        return _call_claude(system_prompts, user_message)


# ──────────────────────────────────────────────────
# 15-d. メイン: build_highlights
# ──────────────────────────────────────────────────

def build_highlights(
    df: pd.DataFrame,
    date: str,
    prompts_path: str = str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt"),
) -> pd.DataFrame:
    """
    活躍選手シート生成。

    フロー:
      1. Statcastデータから「スコアプレー」「野手成績」「投手成績」のテキストを組み立て
      2. prompts/活躍選手_MLB.txt のシステムプロンプトを読み込み
      3. Claude API に渡して活躍選手コメントを生成
      4. LLMの出力をパースして DataFrame に変換

    GEMINI_API_KEY / GEMINI_API_KEY が未設定の場合は空のDataFrameを返す。

    出力カラム: date / no / player / team / role / detail
    """
    system_prompts = _load_system_prompts(prompts_path)
    if not system_prompts:
        logger.warning("システムプロンプトが読み込めませんでした。活躍選手をスキップします。")
        return pd.DataFrame(columns=["date","no","player","player_en","team","role","detail"])

    # ── データテキスト組み立て ──
    score_plays     = _build_score_plays(df)
    batter_summary  = _build_batter_summary(df)
    pitcher_summary = _build_pitcher_summary(df)

    user_message = f"""以下は {date} のMLB試合データです。

=== スコアプレー ===
{score_plays}

=== 野手成績 ===
{batter_summary}

=== 投手成績 ===
{pitcher_summary}

上記データをもとに、指定の形式で活躍選手を選出・解説してください。

【重要】各エントリに `player_en` フィールドを追加し、選手名の英語表記（姓, 名 の形式。例: "mullins, cedric"）を必ず入れてください。"""

    provider_label = "Gemini" if LLM_PROVIDER != "claude" else "Claude"
    print(f"  {provider_label} API で活躍選手選出中...")
    llm_output = _call_llm(system_prompts, user_message)

    if not llm_output:
        return pd.DataFrame(columns=["date","no","player","player_en","team","role","detail"])

    # ── LLM出力のパース ──
    rows = _parse_highlights_output(llm_output, date)
    print(f"  ✓ 活躍選手: {len(rows)}名を選出")
    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["date","no","player","player_en","team","role","detail"])


def _parse_highlights_output(text: str, date: str) -> list[dict]:
    """
    LLM出力テキスト → DataFrame行リスト。

    Geminiは前置きテキスト付きでJSONを返すことがあるため
    テキスト中のJSONリスト部分を抽出して解析する。
    """
    import re
    rows = []

    # ── Step1: コードブロック除去 ──
    text = re.sub(r"```json", "", text, flags=re.IGNORECASE)
    text = re.sub(r"```",     "", text)
    text = text.strip()

    # ── Step2: JSON部分を抽出（[ ... ] の最外部を探す）──
    def _extract_json_array(s: str) -> str:
        start = s.find("[")
        if start == -1:
            return s
        depth = 0
        for i, ch in enumerate(s[start:], start):
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    return s[start:i+1]
        return s[start:]  # 閉じ括弧がない場合は末尾まで

    json_text = _extract_json_array(text)

    # ── Step3: パース ──
    try:
        data = json.loads(json_text)
    except json.JSONDecodeError:
        # 不完全なJSONを末尾から補完して再試行
        try:
            fixed = json_text.rstrip().rstrip(",")
            if not fixed.endswith("}"):
                fixed += "}"
            if not fixed.endswith("]"):
                fixed += "]"
            data = json.loads(fixed)
        except json.JSONDecodeError as e:
            logger.error(f"  ⚠ 活躍選手JSONパース失敗: {e}")
            logger.error(f"  出力先頭200文字: {text[:200]}")
            return []

    if not isinstance(data, list):
        logger.warning("  ⚠ LLM出力がリスト形式ではありません")
        return []

    for item in data:
        if not isinstance(item, dict):
            continue
        detail = item.get("detail", "")
        # roleがない場合は詳細コメントから自動判定
        if item.get("role"):
            role = item["role"]
        elif re.search(r"(奪三振|好投|無失点|イニング|被安打|先発|リリーフ|セーブ|完封|失点)", detail):
            role = "投手"
        else:
            role = "打者"

        rows.append({
            "date":      date,
            "no":        item.get("no", len(rows) + 1),
            "player":    item.get("player", ""),
            "player_en": item.get("player_en", ""),
            "team":      item.get("team", ""),
            "role":      role,
            "detail":    detail,
        })

    return rows

# %%
# ==================================================
# Section 16. 試合別各シート
# ==================================================

def _role(g, df, gid):
    role = "先発" if g["inning"].min()==1 else "中継ぎ"
    if g["inning"].max()==df[df["game_pk"]==gid]["inning"].max() and role!="先発":
        role = "抑え"
    return role

def build_game_batter_stats(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (gid, bid), g in df.groupby(["game_pk","batter"]):
        row0 = g.iloc[0]
        ha   = "home" if row0.get("inning_topbot","")=="Bot" else "away"
        name = str(row0.get("batter_name") or row0.get("player_name", str(bid)))
        team = row0.get("home_team","") if ha=="home" else row0.get("away_team","")
        stats= calc_batter_pa_stats(g)
        rows.append({"試合ID":str(gid),"試合日":str(row0["game_date"])[:10],
                     "選手名":name,"チーム":team,"ホーム/アウェイ":ha,
                     "打順":0,"守備位置":"",**stats})
    return pd.DataFrame(rows)

def build_game_pitcher_stats(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    is_statsapi = "_statsapi_source" in df.columns and df["_statsapi_source"].any()
    for (gid, pid), g in df.groupby(["game_pk","pitcher"]):
        row0 = g.iloc[0]
        ha   = "home" if row0.get("inning_topbot","")=="Top" else "away"
        name = str(row0.get("player_name", pid))
        team = row0.get("home_team","") if ha=="home" else row0.get("away_team","")

        if is_statsapi:
            # statsapiデータ: 補助カラムから正確な成績を取得
            stats = _calc_pitcher_stats_from_statsapi(g)
        else:
            stats = calc_pitcher_stats(g)

        rows.append({"試合ID":str(gid),"試合日":str(row0["game_date"])[:10],
                     "選手名":name,"チーム":team,"ホーム/アウェイ":ha,
                     "役割":_role(g,df,gid),"勝敗成績":"",**stats})
    return pd.DataFrame(rows)


def _calc_pitcher_stats_from_statsapi(g: pd.DataFrame) -> dict:
    """
    statsapiデータから投手成績を集計する。
    補助カラム _pit_* に格納された実際の成績を使用。

    [修正] g は投球行ベース（1行=1投球）のため、TBF を len(g) で求めるのは誤り。
           at_bat_number でグループして打席数を正しく算出する。
           また、スイング指標・ゾーン指標も投球行から計算できるため NaN を返さない。
    """
    row0    = g.iloc[0]
    ip_str  = str(row0.get("_pit_ip_str", "0") or "0")
    pitches = int(row0.get("_pit_pitches", 0) or 0)

    # ── TBF: 投球行ではなく打席単位で数える（旧: len(g) は投球数なので誤り）──
    pa_df = g.groupby("at_bat_number").first().reset_index()
    tbf   = len(pa_df)

    k       = int(row0.get("_pit_k",   0) or 0)
    bb      = int(row0.get("_pit_bb",  0) or 0)
    h       = int(row0.get("_pit_h",   0) or 0)
    hr      = int(row0.get("_pit_hr",  0) or 0)
    hbp     = int(row0.get("_pit_hbp", 0) or 0)
    er      = int(row0.get("_pit_er",  0) or 0)
    runs    = int(row0.get("_pit_r",   0) or 0)

    # 打球種別（投球行の bb_type から集計）
    gb     = g["bb_type"].eq("ground_ball").sum()
    ld     = g["bb_type"].eq("line_drive").sum()
    fb     = g["bb_type"].eq("fly_ball").sum()
    batted = int(gb + ld + fb)

    # スイング・ゾーン指標（投球行ベースで正しく計算できる）
    sw = calc_swing_metrics(g)
    zp = calc_zone_pcts(g)
    mv = calc_pitch_movement_stats(g)

    NaN = float("nan")
    return {
        "投球回":     ip_str,
        "投球数":     pitches if pitches > 0 else len(g),
        "対戦打者数": tbf,
        "失点":       runs,
        "自責点":     er,
        "被安打":     h,
        "被本塁打":   hr,
        "与四球":     bb,
        "与死球":     hbp,
        "奪三振":     k,
        "K%":         _pct(k, tbf),
        "BB%":        _pct(bb, tbf),
        "K-BB%":      _round(_safe(_pct(k,tbf),0) - _safe(_pct(bb,tbf),0), 1),
        "全打球数":   batted,
        "GB":  int(gb), "LD": int(ld), "FB": int(fb),
        "IFFB": 0, "判断不可打球": 0, "HR": hr, "H": h,
        "GB%": _pct(gb, batted), "LD%": _pct(ld, batted), "FB%": _pct(fb, batted),
        "IFFB%": NaN, "HR%": NaN,
        "平均打球速度EV(km/h)": NaN, "最高打球速度EV(km/h)": NaN,
        "Hard-Hit%": NaN, "Barrel%": NaN,
        "xBA": NaN, "xwOBA": NaN, "xSLG": NaN, "wOBA": NaN, "BABIP": NaN, "ISO": NaN,
        "空振り率":           sw["空振り率"],
        "ゾーン内スイング率": sw["ゾーン内スイング率"],
        "ゾーン外スイング率": sw["ゾーン外スイング率"],
        "ゾーン率":           sw["ゾーン率"],
        "ストライク率":       sw["ストライク率"],
        "ゾーン内SW数":   sw["ゾーン内SW数"],
        "ゾーン内投球数": sw["ゾーン内投球数"],
        "ゾーン外SW数":   sw["ゾーン外SW数"],
        "ゾーン外投球数": sw["ゾーン外投球数"],
        "SW数":         sw["SW数"],
        "コンタクト数": sw["コンタクト数"],
        "空振り数":     sw["空振り数"],
        **zp,
        **mv,
        "xwOBA被打": NaN, "wOBA被打": NaN,
    }

def build_game_pitcher_lr(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    # stand列が全NaN/None（statsapiデータ等）の場合はスキップ
    if "stand" not in df.columns or df["stand"].isna().all():
        return pd.DataFrame()
    for (gid, pid), g in df.groupby(["game_pk","pitcher"]):
        g_st = g[g["stand"].notna()]
        if g_st.empty:
            continue
        row0 = g.iloc[0]
        ha   = "home" if row0.get("inning_topbot","")=="Top" else "away"
        name = str(row0.get("player_name", pid))
        team = row0.get("home_team","") if ha=="home" else row0.get("away_team","")
        role = _role(g, df, gid)
        for hand, hg in g_st.groupby("stand"):
            lr = "左" if hand=="L" else "右"
            s  = calc_pitcher_stats(hg)
            s.pop("投球回", None)
            rows.append({"試合ID":str(gid),"試合日":str(row0["game_date"])[:10],
                         "選手名":name,"チーム":team,"ホーム/アウェイ":ha,"役割":role,"対打者":lr,**s})
    return pd.DataFrame(rows)

def build_game_pitch_mix(df: pd.DataFrame) -> pd.DataFrame:
    # pitch_typeがすべてNoneの場合はスキップ（球種データなし）
    if "pitch_type" not in df.columns or df["pitch_type"].isna().all():
        return pd.DataFrame()
    rows = []
    for (gid, pid), g in df.groupby(["game_pk","pitcher"]):
        row0 = g.iloc[0]
        ha   = "home" if row0.get("inning_topbot","")=="Top" else "away"
        name = str(row0.get("player_name", pid))
        team = row0.get("home_team","") if ha=="home" else row0.get("away_team","")
        role = _role(g, df, gid)
        for _, pr in calc_pitch_type_stats(g).iterrows():
            rows.append({"試合ID":str(gid),"試合日":str(row0["game_date"])[:10],
                         "選手名":name,"チーム":team,"ホーム/アウェイ":ha,"役割":role,**pr.to_dict()})
    return pd.DataFrame(rows)

def build_game_pitch_mix_lr(df: pd.DataFrame) -> pd.DataFrame:
    # pitch_typeがすべてNoneの場合はスキップ（球種データなし）
    if "pitch_type" not in df.columns or df["pitch_type"].isna().all():
        return pd.DataFrame()
    rows = []
    for (gid, pid), g in df.groupby(["game_pk","pitcher"]):
        row0 = g.iloc[0]
        ha   = "home" if row0.get("inning_topbot","")=="Top" else "away"
        name = str(row0.get("player_name", pid))
        team = row0.get("home_team","") if ha=="home" else row0.get("away_team","")
        role = _role(g, df, gid)
        for hand, hg in g.groupby("stand"):
            lr = "左" if hand=="L" else "右"
            for _, pr in calc_pitch_type_stats(hg).iterrows():
                rows.append({"試合ID":str(gid),"試合日":str(row0["game_date"])[:10],
                             "選手名":name,"チーム":team,"ホーム/アウェイ":ha,
                             "役割":role,"対打者":lr,**pr.to_dict()})
    return pd.DataFrame(rows)

# %%
# ==================================================
# Section 17. games datamart
# ==================================================

def run_games_datamart(
    df: pd.DataFrame,
    date: str,
    prompts_path: str = str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt"),
    with_highlights: bool = True,
) -> str:
    """
    games datamart (xlsx) + JSON を生成する。

    with_highlights=True  : 活躍選手を Claude API で生成して含める（Step games/highlights）
    with_highlights=False : RAWキャッシュがあればそれを使用、なければ空で生成（Step datamart）
    """
    out_path = os.path.join(GAMES_DM_DIR, f"{date}.xlsx")
    print(f"  datamart生成中...")

    if with_highlights:
        hl_df = build_highlights(df, date, prompts_path=prompts_path)
    else:
        # RAWフォルダに活躍選手キャッシュがあれば読み込む
        date_nodash = date.replace("-", "")
        hl_raw_path = os.path.join(RAW_DIR, f"highlights_{date_nodash}.xlsx")
        if os.path.exists(hl_raw_path):
            try:
                hl_df = pd.read_excel(hl_raw_path, engine="openpyxl")
                print(f"  キャッシュ: {os.path.basename(hl_raw_path)} ({len(hl_df)}件)")
            except Exception as e:
                print(f"  [WARN] highlights RAW読み込み失敗: {e}")
                hl_df = pd.DataFrame(columns=["date","no","player","player_en","team","role","detail"])
        else:
            hl_df = pd.DataFrame(columns=["date","no","player","player_en","team","role","detail"])

    try:
        sheets = {
            "試合概要":              build_game_summary(df),
            "試合別打者成績":        build_game_batter_stats(df),
            "試合別投手成績":        build_game_pitcher_stats(df),
            "試合別投手成績_左右別": build_game_pitcher_lr(df),
            "試合別投球配球":        build_game_pitch_mix(df),
            "試合別投球配球_左右別": build_game_pitch_mix_lr(df),
            "活躍選手":              hl_df,
        }
    except Exception as e:
        import traceback
        print(f"  [ERROR] シート生成失敗: {e}")
        traceback.print_exc()
        return ""

    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for sname, sdf in sheets.items():
                if sdf is not None and not sdf.empty:
                    sdf.to_excel(writer, sheet_name=sname, index=False)
                    _apply_style(writer.sheets[sname])
                else:
                    # 空シートも最低1行（ヘッダーのみ）書き込む
                    empty = pd.DataFrame(columns=sdf.columns if sdf is not None else [])
                    empty.to_excel(writer, sheet_name=sname, index=False)
        print(f"  ✓ datamart: {os.path.basename(out_path)} ({os.path.getsize(out_path)//1024}KB)")
    except Exception as e:
        import traceback
        print(f"  [ERROR] Excel書き込み失敗: {e}")
        traceback.print_exc()
        return ""

    try:
        write_games_json(out_path, date, df_statcast=df)
    except Exception as e:
        import traceback
        print(f"  [ERROR] JSON生成失敗: {e}")
        traceback.print_exc()

    return out_path


def run_highlights_only(
    df: pd.DataFrame,
    date: str,
    prompts_path: str = str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt"),
) -> str:
    """
    Step highlights: 既存 datamart xlsx の「活躍選手」シートのみ上書きして JSON を再生成する。
    xlsx が存在しない場合は run_games_datamart を with_highlights=True で呼ぶ。
    """
    out_path = os.path.join(GAMES_DM_DIR, f"{date}.xlsx")
    if not os.path.exists(out_path):
        print("  [WARN] datamartが存在しません。gamesから実行します。")
        return run_games_datamart(df, date, prompts_path=prompts_path, with_highlights=True)


    hl_df = build_highlights(df, date, prompts_path=prompts_path)

    # ── raw/{date}/highlights_{date}.xlsx に保存 ──
    date_nodash = date.replace("-", "")
    hl_raw_path = os.path.join(RAW_DIR, f"highlights_{date_nodash}.xlsx")
    try:
        Path(RAW_DIR).mkdir(parents=True, exist_ok=True)
        hl_df.to_excel(hl_raw_path, index=False, engine="openpyxl")
        print(f"  保存: {os.path.basename(hl_raw_path)} ({len(hl_df)}件)")
    except Exception as e:
        print(f"  [WARN] highlights RAW保存失敗: {e}")

    # openpyxl で既存ファイルを開き、活躍選手シートのみ差し替え
    import openpyxl as _opx
    wb = _opx.load_workbook(out_path)
    if "活躍選手" in wb.sheetnames:
        del wb["活躍選手"]
    ws = wb.create_sheet("活躍選手")
    # ヘッダー書き込み
    ws.append(list(hl_df.columns))
    for _, row in hl_df.iterrows():
        ws.append(list(row))
    _apply_style(ws)
    wb.save(out_path)
    wb.close()
    print(f"  ✓ 活躍選手シート更新: {len(hl_df)}件")

    # JSON再生成（df_statcast を渡してlocs/cbsを保持）
    write_games_json(out_path, date, df_statcast=df)
    return out_path

# %%
# ==================================================
# Section 17b. games JSON 生成（ダッシュボード用）
# ==================================================

def _fv(v, d=1):
    """float変換。NaN/None は None を返す"""
    try:
        f = float(v)
        return None if (f != f) else round(f, d)   # NaN check
    except Exception:
        return None

def _iv(v):
    """int変換。NaN/None は 0"""
    try:
        f = float(v)
        return 0 if (f != f) else int(f)
    except Exception:
        return 0

def _nv(v, default=""):
    if v is None or (isinstance(v, float) and v != v):
        return default
    return v


def _build_mlb_locs_and_cbs(df: pd.DataFrame) -> tuple[dict, dict]:
    """
    Statcast DataFrame から pitch_locs と cbs_idx を生成する。

    pitch_locs : {(game_pk, player_name, pitch_type, hand): [[cx,cy,flag,in_zone,rtype,is_strike,balls,strikes], ...]}
    cbs_idx    : {(game_pk, player_name, pitch_type, hand): {countKey: {c,sw,fo,lo,ba,ou,hi}}}

    座標正規化:
      cx = plate_x / MLB_PLATE_HALF_W  （右打者外角=+1）
      cy = (plate_z - sz_center) / sz_half  （高め=+1）
      flip_x: p_throws XOR stand が左 → 投手視点に統一
    """
    MLB_PLATE_HALF_W = (17 / 12) / 2   # ホームプレート半幅(ft)
    MLB_SZ_TOP_AVG   = 3.5
    MLB_SZ_BOT_AVG   = 1.5

    pitch_locs: dict = {}
    cbs_idx:    dict = {}

    if df.empty:
        return pitch_locs, cbs_idx

    throws_map: dict = {}
    if "p_throws" in df.columns and "player_name" in df.columns:
        throws_map = df.groupby("player_name")["p_throws"].first().to_dict()

    for _, row in df.iterrows():
        gid   = str(row.get("game_pk",  ""))
        pname = str(row.get("player_name", ""))
        pt    = str(row.get("pitch_type", "") or "")
        stand = str(row.get("stand", "") or "")
        hand  = "R" if stand == "R" else ("L" if stand == "L" else "ALL")

        # ── cbs 集計（座標なし投球も含む） ──
        balls   = row.get("balls",   None)
        strikes = row.get("strikes", None)
        if balls is None or strikes is None:
            continue
        try:
            balls = int(balls); strikes = int(strikes)
        except (ValueError, TypeError):
            continue
        ckey = f"{min(balls,3)}-{min(strikes,2)}"

        desc     = str(row.get("description", "") or "")
        event    = str(row.get("events", "")       or "")
        is_swstr = desc in {"swinging_strike","swinging_strike_blocked","foul_tip"}
        is_foul  = desc in {"foul","foul_bunt"} and not is_swstr
        is_look  = desc in {"called_strike"}
        is_ball  = desc in {"ball","blocked_ball","intent_ball","pitchout","automatic_ball"}
        is_inplay = desc.startswith("hit_into_play")
        is_hit   = event in {"single","double","triple","home_run"}

        for ck in [(gid, pname, pt, "ALL"), (gid, pname, pt, hand)]:
            if ck not in cbs_idx: cbs_idx[ck] = {}
            if ckey not in cbs_idx[ck]: cbs_idx[ck][ckey] = {"c":0,"sw":0,"fo":0,"lo":0,"ba":0,"ou":0,"hi":0}
            d2 = cbs_idx[ck][ckey]
            d2["c"]  += 1
            d2["sw"] += int(is_swstr)
            d2["fo"] += int(is_foul)
            d2["lo"] += int(is_look)
            d2["ba"] += int(is_ball)
            d2["ou"] += int(is_inplay and not is_hit)
            d2["hi"] += int(is_inplay and is_hit)

        # ── pitch_locs（座標あり投球のみ） ──
        px = row.get("plate_x", None)
        pz = row.get("plate_z", None)
        if px is None or pz is None:
            continue
        try:
            px = float(px); pz = float(pz)
        except (ValueError, TypeError):
            continue
        if math.isnan(px) or math.isnan(pz):
            continue

        try:
            sz_top = float(row.get("sz_top") or MLB_SZ_TOP_AVG)
            sz_bot = float(row.get("sz_bot") or MLB_SZ_BOT_AVG)
            if math.isnan(sz_top): sz_top = MLB_SZ_TOP_AVG
            if math.isnan(sz_bot): sz_bot = MLB_SZ_BOT_AVG
        except (ValueError, TypeError):
            sz_top, sz_bot = MLB_SZ_TOP_AVG, MLB_SZ_BOT_AVG

        sz_center = (sz_top + sz_bot) / 2
        sz_half   = max((sz_top - sz_bot) / 2, 0.01)

        cx = px / MLB_PLATE_HALF_W
        cy = (pz - sz_center) / sz_half

        p_throws = throws_map.get(pname, row.get("p_throws", "") or "")
        flip_x   = (p_throws == "L") ^ (stand == "L")
        if flip_x: cx = -cx
        cx = round(float(cx), 3)
        cy = round(float(cy), 3)

        is_swing    = desc in {"swinging_strike","swinging_strike_blocked","foul_tip",
                                "foul","foul_bunt","hit_into_play",
                                "hit_into_play_no_out","hit_into_play_score"}
        in_zone     = bool(row.get("is_in_zone", False))
        flag        = 1 if is_swstr else (0 if is_swing else -1)
        is_strike_f = 1 if (desc in {"called_strike","swinging_strike","swinging_strike_blocked",
                                      "foul_tip","foul","foul_bunt","automatic_strike"}
                             or is_inplay) else 0

        bb_type = str(row.get("bb_type", "") or "")
        # 安打を最優先で判定（bb_typeより先に確認）
        if   event == "home_run":                              rtype = 5
        elif event in {"double", "triple"}:                    rtype = 4
        elif event == "single":                                rtype = 3
        elif is_inplay and bb_type == "ground_ball":          rtype = 1
        elif is_inplay and bb_type in {"fly_ball", "popup"}:  rtype = 2
        elif is_inplay and bb_type == "line_drive":           rtype = 2
        else:                                                  rtype = 0

        # xwOBA（打球結果の期待値）
        xwoba_val = row.get("estimated_woba_using_speedangle", None)
        try:
            xwoba_val = round(float(xwoba_val), 3) if xwoba_val is not None and str(xwoba_val) not in ('', 'nan', 'None') else None
        except (ValueError, TypeError):
            xwoba_val = None

        entry = [cx, cy, flag, 1 if in_zone else 0, rtype, is_strike_f, balls, strikes, xwoba_val]
        for lk in [(gid, pname, pt, "ALL"), (gid, pname, pt, hand)]:
            if lk not in pitch_locs: pitch_locs[lk] = []
            pitch_locs[lk].append(entry)

    return pitch_locs, cbs_idx


def _build_game_json(dm_path: str, date: str,
                     pitch_locs: dict | None = None,
                     cbs_idx:    dict | None = None,
                     hand_map:   dict | None = None) -> dict:
    """
    games/datamart/{date}.xlsx → ダッシュボード用 DATA dict

    ダッシュボードが期待する構造:
    {
      "YYYY-MM-DD": [ { game_obj }, ... ],
      "highlights": { "YYYY-MM-DD": [ { no, player, team, role, detail }, ... ] },
      "_game_type": "公式戦",
      "_league":    "mlb"
    }

    game_obj:
    {
      gameId, home, away, homeScore, awayScore,
      stadium, time, gameTime, att, status,
      innings: { home: [...], away: [...] },
      rhe: { home: [R,H,E], away: [R,H,E] },
      pitchers: { home: [pitcher_obj, ...], away: [...] },
      batters:  { home: [batter_obj, ...], away: [...] },
      highlight
    }

    pitcher_obj (MLB独自指標付き):
    {
      name, role, result, ip, pitches, h, bb, hbp, k, r, er,
      kpct, bbpct, kbbpct, gbpct, swstr, oSwing, strike, zone,
      mix: [{ name, key, count, pct, vel, maxVel, swstr, oSwing,
              zone, strike, gbpct, hits, hr,
              avgSpin, ivb, hb, ext }],
      mixVsR, mixVsL,
      pitStatVsR, pitStatVsL,
      season: { ip, era, kpct, bbpct, kbbpct }
    }

    batter_obj (MLB独自指標付き):
    {
      order, name, pos, abs,
      ops, obp, slg, h, hr, bb, k, pa, ab, rbi,
      avgEV, hardHitPct, barrelPct, xba, xwoba,
      batSpd, attackAngle
    }
    """
    import openpyxl as _opx
    wb = _opx.load_workbook(dm_path, read_only=True)

    def _rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        try:
            headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        except StopIteration:
            return []
        return [dict(zip(headers, [c for c in row])) for row in ws.iter_rows(min_row=2, values_only=True)]

    # ── インデックス構築 ──
    mix_idx    = {}   # (game_id, name) → [rows]
    mix_lr_idx = {}   # (game_id, name, lr) → [rows]
    pit_lr_idx = {}   # (game_id, name, lr) → row
    pit_idx    = {}   # game_id → {home:[], away:[]}
    bat_idx    = {}   # game_id → {home:[], away:[]}
    locs_idx   = pitch_locs or {}   # (game_id, name, pitch_type, hand) → [entry]
    _cbs_idx   = cbs_idx    or {}   # (game_id, name, pitch_type, hand) → {countKey: {...}}
    hand_idx: dict = hand_map or {}  # (game_id, name) → "R" or "L"

    for r in _rows("試合別投球配球"):
        k = (str(r["試合ID"]), str(r["選手名"]))
        mix_idx.setdefault(k, []).append(r)

    for r in _rows("試合別投球配球_左右別"):
        k = (str(r["試合ID"]), str(r["選手名"]), str(r.get("対打者","")))
        mix_lr_idx.setdefault(k, []).append(r)

    for r in _rows("試合別投手成績_左右別"):
        k = (str(r["試合ID"]), str(r["選手名"]), str(r.get("対打者","")))
        pit_lr_idx[k] = r

    for r in _rows("試合別投手成績"):
        gid = str(r["試合ID"])
        if gid not in pit_idx:
            pit_idx[gid] = {"home": [], "away": []}
        side = str(r.get("ホーム/アウェイ", "home"))
        pit_idx[gid][side].append(r)

    for r in _rows("試合別打者成績"):
        gid = str(r["試合ID"])
        if gid not in bat_idx:
            bat_idx[gid] = {"home": [], "away": []}
        side = str(r.get("ホーム/アウェイ", "home"))
        bat_idx[gid][side].append(r)

    def _mix_obj(r, game_id: str = "", pitcher_name: str = "", bat_hand: str = "ALL"):
        pt  = _nv(r.get("球種コード"), "")
        obj = {
            "name":       _nv(r.get("球種名"), ""),
            "key":        pt,
            "count":      _iv(r.get("投球数")),
            "pct":        _fv(r.get("投球割合%")),
            "vel":        _fv(r.get("平均球速")),
            "maxVel":     _fv(r.get("最高球速")),
            "swstr":      _fv(r.get("空振り率")),
            "oSwing":     _fv(r.get("ゾーン外スイング率")),
            "zone":       _fv(r.get("ゾーン率")),
            "strike":     _fv(r.get("ストライク率")),
            "gbpct":      _fv(r.get("GB%")),
            "hits":       _iv(r.get("H")),
            "hr":         _iv(r.get("HR")),
            "xwoba":      _fv(r.get("xwOBA"), d=3),
            # MLB独自
            "avgSpin":    _fv(r.get("平均回転数(rpm)"), d=0),
            "spinAxis":   _fv(r.get("回転軸(°)"), d=0),
            "activeSpin": _fv(r.get("回転効率%(近似)")),
            "vaa":        _fv(r.get("VAA(°)")),
            "ivb":        _fv(r.get("IVB(cm)")),
            "hb":         _fv(r.get("HB arm(cm)")),
            "ext":        _fv(r.get("Extension(m)")),
        }
        # locs（コース分布）
        key_lr  = (game_id, pitcher_name, pt, bat_hand)
        key_all = (game_id, pitcher_name, pt, "ALL")
        locs = locs_idx.get(key_lr) or locs_idx.get(key_all) or []
        if locs:
            obj["locs"] = locs
        # cbs（カウント別集計）
        cbs = _cbs_idx.get(key_lr) or _cbs_idx.get(key_all)
        if cbs:
            obj["cbs"] = cbs
        return obj

    def _lr_stat(r):
        if not r:
            return None
        return {
            "pitches": _iv(r.get("投球数")),
            "tbf":     _iv(r.get("対戦打者数")),
            "k":       _iv(r.get("奪三振")),
            "bb":      _iv(r.get("与四球")),
            "h":       _iv(r.get("被安打")),
            "kpct":    _fv(r.get("K%")),
            "bbpct":   _fv(r.get("BB%")),
            "kbbpct":  _fv(r.get("K-BB%")),
            "swstr":   _fv(r.get("空振り率")),
            "oSwing":  _fv(r.get("ゾーン外スイング率")),
            "zone":    _fv(r.get("ゾーン率")),
            "strike":  _fv(r.get("ストライク率")),
            "gbpct":   _fv(r.get("GB%")),
        }

    def _build_pitcher(r):
        gid  = str(r["試合ID"])
        name = str(r["選手名"])
        mix  = sorted(mix_idx.get((gid, name), []), key=lambda x: -_iv(x.get("投球数")))
        return {
            "name":    name,
            "role":    _nv(r.get("役割"), ""),
            "hand":    hand_idx.get((gid, name), "R"),  # 利き手 (R/L)
            "result":  _nv(r.get("勝敗成績"), ""),
            "ip":      _nv(r.get("投球回"), ""),
            "pitches": _iv(r.get("投球数")),
            "h":       _iv(r.get("被安打")),
            "bb":      _iv(r.get("与四球")),
            "hbp":     _iv(r.get("与死球")),
            "k":       _iv(r.get("奪三振")),
            "r":       _iv(r.get("失点")),
            "er":      _iv(r.get("自責点")),
            "tbf":     _iv(r.get("対戦打者数")),
            "kpct":    _fv(r.get("K%")),
            "bbpct":   _fv(r.get("BB%")),
            "kbbpct":  _fv(r.get("K-BB%")),
            "gbpct":   _fv(r.get("GB%")),
            "swstr":   _fv(r.get("空振り率")),
            "oSwing":  _fv(r.get("ゾーン外スイング率")),
            "strike":  _fv(r.get("ストライク率")),
            "zone":    _fv(r.get("ゾーン率")),
            # MLB独自
            "avgEV":       _fv(r.get("平均打球速度EV(km/h)")),
            "hardHitPct":  _fv(r.get("Hard-Hit%")),
            "barrelPct":   _fv(r.get("Barrel%")),
            "xwoba":       _fv(r.get("xwOBA被打"), d=3),
            "avgSpin":     _fv(r.get("平均回転数(rpm)"), d=0),
            "spinAxis":    _fv(r.get("回転軸(°)"), d=0),
            "activeSpin":  _fv(r.get("回転効率%(近似)")),
            "vaa":         _fv(r.get("VAA(°)")),
            "extension":   _fv(r.get("Extension(m)")),
            # 球種 mix（locs/cbs付与のためgame_id/pitcher_name/handを渡す）
            "mix":    [_mix_obj(m, gid, name, "ALL") for m in mix],
            "mixVsR": [_mix_obj(m, gid, name, "R") for m in sorted(
                mix_lr_idx.get((gid, name, "右"), []), key=lambda x: -_iv(x.get("投球数")))],
            "mixVsL": [_mix_obj(m, gid, name, "L") for m in sorted(
                mix_lr_idx.get((gid, name, "左"), []), key=lambda x: -_iv(x.get("投球数")))],
            "pitStatVsR": _lr_stat(pit_lr_idx.get((gid, name, "右"))),
            "pitStatVsL": _lr_stat(pit_lr_idx.get((gid, name, "左"))),
            "season": {},
        }

    def _build_batter(r):
        abs_str = _nv(r.get("打席別結果"), "")
        abs_list = [a.strip() for a in abs_str.split(",") if a.strip()] if abs_str else []
        return {
            "order":  _iv(r.get("打順")),
            "name":   _nv(r.get("選手名"), ""),
            "pos":    _nv(r.get("守備位置"), ""),
            "abs":    abs_list,
            "ops":    _fv(r.get("OPS"), d=3),
            "obp":    _fv(r.get("出塁率"), d=3),
            "slg":    _fv(r.get("長打率"), d=3),
            "h":      _iv(r.get("安打")),
            "hr":     _iv(r.get("本塁打")),
            "bb":     _iv(r.get("四球")),
            "k":      _iv(r.get("三振")),
            "pa":     _iv(r.get("打席")),
            "ab":     _iv(r.get("打数")),
            "rbi":    _iv(r.get("打点")),
            "sb":     _iv(r.get("盗塁")),
            # MLB独自
            "avgEV":       _fv(r.get("平均打球速度EV(km/h)")),
            "maxEV":       _fv(r.get("最高打球速度EV(km/h)")),
            "hardHitPct":  _fv(r.get("Hard-Hit%")),
            "barrelPct":   _fv(r.get("Barrel%")),
            "xba":         _fv(r.get("xBA"), d=3),
            "xwoba":       _fv(r.get("xwOBA"), d=3),
            "batSpd":      _fv(r.get("平均バットスピード(km/h)")),
            "attackAngle": _fv(r.get("平均アタックアングル(°)")),
        }

    def _parse_inn(s):
        if not s:
            return []
        return [_iv(x) for x in str(s).split(",")]

    DATA = {}
    for g in _rows("試合概要"):
        gid      = str(g["試合ID"])
        date_str = str(g["試合日"])[:10]

        pit_sides = pit_idx.get(gid, {"home": [], "away": []})
        bat_sides = bat_idx.get(gid, {"home": [], "away": []})

        game_obj = {
            "gameId":    gid,
            "home":      _nv(g.get("ホームチーム"), ""),
            "away":      _nv(g.get("アウェイチーム"), ""),
            "homeScore": _iv(g.get("ホーム得点")),
            "awayScore": _iv(g.get("アウェイ得点")),
            "stadium":   _nv(g.get("球場"), ""),
            "time":      _nv(g.get("開始時間"), ""),
            "gameTime":  _nv(g.get("試合時間"), ""),
            "att":       _nv(g.get("観客数"), ""),
            "status":    "",
            "innings": {
                "home": _parse_inn(g.get("ホームイニング得点")),
                "away": _parse_inn(g.get("アウェイイニング得点")),
            },
            "rhe": {
                "home": [_iv(g.get("ホーム得点")),  _iv(g.get("ホーム安打")),  _iv(g.get("ホームエラー"))],
                "away": [_iv(g.get("アウェイ得点")), _iv(g.get("アウェイ安打")), _iv(g.get("アウェイエラー"))],
            },
            "pitchers": {
                "home": [_build_pitcher(r) for r in pit_sides["home"]],
                "away": [_build_pitcher(r) for r in pit_sides["away"]],
            },
            "batters": {
                "home": [_build_batter(r) for r in sorted(bat_sides["home"], key=lambda x: _iv(x.get("打順")))],
                "away": [_build_batter(r) for r in sorted(bat_sides["away"], key=lambda x: _iv(x.get("打順")))],
            },
            "highlight": _nv(g.get("ハイライト"), ""),
        }
        DATA.setdefault(date_str, []).append(game_obj)

    # 活躍選手
    hl_by_date = {}
    for r in _rows("活躍選手"):
        d = str(r.get("date", ""))[:10]
        if not d:
            continue
        hl_by_date.setdefault(d, []).append({
            "no":        _iv(r.get("no")),
            "player":    _nv(r.get("player"),    ""),
            "player_en": _nv(r.get("player_en"), ""),
            "team":      _nv(r.get("team"),       ""),
            "role":      _nv(r.get("role"),       ""),
            "detail":    _nv(r.get("detail"),     ""),
        })
    DATA["highlights"]  = hl_by_date
    DATA["_game_type"]  = "公式戦"
    DATA["_league"]     = "mlb"

    wb.close()
    return DATA


def write_games_json(dm_path: str, date: str,
                     df_statcast: "pd.DataFrame | None" = None) -> str:
    """
    games/datamart/{date}.xlsx を読んで games/json/{date}.json を出力し、
    index.json を更新する。
    df_statcast を渡すとコース分布（locs）とカウント別（cbs）をJSONに含める。
    """
    print(f"  JSON生成中...")

    pitch_locs, cbs_idx = None, None
    hand_map: dict = {}  # (game_pk_str, player_name) → "R" or "L"
    if df_statcast is not None and not df_statcast.empty:
        try:
            print(f"  locs/cbs生成中...")
            pitch_locs, cbs_idx = _build_mlb_locs_and_cbs(df_statcast)
            print(f"  locs: {len(pitch_locs)}キー  cbs: {len(cbs_idx)}キー")
            # 投手利き手マップを構築
            if "player_name" in df_statcast.columns and "p_throws" in df_statcast.columns and "game_pk" in df_statcast.columns:
                for _, row in df_statcast[["game_pk","player_name","p_throws"]].drop_duplicates().iterrows():
                    gid = str(row["game_pk"])
                    pname = str(row["player_name"])
                    if (gid, pname) not in hand_map:
                        hand_map[(gid, pname)] = str(row["p_throws"] or "R")
                print(f"  hand_map: {len(hand_map)}エントリ")
        except Exception as e:
            import traceback
            print(f"  [WARN] locs/cbs生成失敗: {e}")
            traceback.print_exc()

    data = _build_game_json(dm_path, date, pitch_locs=pitch_locs, cbs_idx=cbs_idx, hand_map=hand_map)

    # {date}.json 出力
    json_path = os.path.join(GAMES_JSON_DIR, f"{date}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  ✓ JSON: {os.path.basename(json_path)} ({os.path.getsize(json_path)//1024}KB)")

    # index.json 更新
    _update_index_json(GAMES_JSON_DIR)
    return json_path


def _update_index_json(json_dir: str) -> None:
    """json_dir 内の YYYY-MM-DD.json を列挙して index.json を再生成"""
    files = sorted(
        f for f in os.listdir(json_dir)
        if f.endswith(".json") and f != "index.json" and not f.startswith("season_")
    )
    index_path = os.path.join(json_dir, "index.json")
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump({"files": files}, f, ensure_ascii=False, indent=2)
    print(f"  ✓ index.json: {len(files)}件")



# %%
# ==================================================
# Section 20. Excelスタイル（指標種別で色分け）
# ==================================================

# MLB独自指標カラムの色
_COL_COLORS = {
    # 打球系 → ダークグリーン
    "平均打球速度EV(km/h)":"1B4332","最高打球速度EV(km/h)":"1B4332",
    "Hard-Hit%":"1B4332","Hard-Hit数":"1B4332",
    "平均打球角度LA(°)":"1B4332","平均飛距離(ft)":"1B4332",
    "最大飛距離(ft)":"1B4332","Barrel%":"1B4332","Barrel数":"1B4332",
    # 期待値系 → ダークネイビー
    "xBA":"1A3A5C","xwOBA":"1A3A5C","xSLG":"1A3A5C",
    "wOBA":"1A3A5C","BABIP":"1A3A5C","ISO":"1A3A5C",
    "xwOBA被打":"1A3A5C","wOBA被打":"1A3A5C",
    # バットスイング系 → ダークオレンジ
    "平均バットスピード(km/h)":"78350A","平均スイング軌跡長(ft)":"78350A",
    "平均アタックアングル(°)":"78350A","平均スイングパスチルト(°)":"78350A",
    # 投球変化量・リリース系 → ダークパープル
    "平均回転数(rpm)":"3B0764","最高回転数(rpm)":"3B0764","回転軸(°)":"3B0764",
    "横変化量(pfx_x cm)":"3B0764","縦変化量(pfx_z cm)":"3B0764",
    "IVB(cm)":"3B0764","HB arm(cm)":"3B0764",
    "リリース高さ(ft)":"3B0764","リリース横位置(ft)":"3B0764",
    "Extension(m)":"3B0764","アーム角度(°)":"3B0764","実効球速(km/h)":"3B0764",
}
_DEFAULT_BG = "1F2937"

def _apply_style(ws):
    for cell in ws[1]:
        bg = _COL_COLORS.get(str(cell.value or ""), _DEFAULT_BG)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 38
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 28)

# %%
# ==================================================
# Section 22. メイン実行
# ==================================================

def _parse_date_arg(date_str: str) -> list:
    """
    --date をパースして日付リストを返す。
      単日:  "2026-04-10"            → ["2026-04-10"]
      範囲:  "2026-04-08:2026-04-10" → ["2026-04-08","2026-04-09","2026-04-10"]
    """
    if ":" in date_str:
        start_str, end_str = date_str.split(":", 1)
        start = datetime.date.fromisoformat(start_str.strip())
        end   = datetime.date.fromisoformat(end_str.strip())
        if start > end:
            raise ValueError(f"開始日が終了日より後です: {start} > {end}")
        days = (end - start).days + 1
        return [(start + datetime.timedelta(days=i)).isoformat() for i in range(days)]
    return [date_str.strip()]


def main():
    parser = argparse.ArgumentParser(
        description="MLB Statcast データマート生成 v2",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--date",
        default=TARGET_DATE,
        help=(
            "対象日。単日または範囲で指定。\n"
            "  単日: YYYY-MM-DD             例) 2026-04-10\n"
            "  範囲: YYYY-MM-DD:YYYY-MM-DD   例) 2026-04-08:2026-04-10"
        ),
    )
    parser.add_argument("--game-type",    default="公式戦",     help="試合種別")
    parser.add_argument("--season-start", default=None,        help="シーズン集計開始日 YYYY-MM-DD")
    parser.add_argument("--prompts-path",  default=str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt"),
                        help="活躍選手プロンプトファイルパス")
    parser.add_argument(
        "--steps",
        nargs="+",
        default=["all"],
        metavar="STEP",
        help=(
            "実行ステップ（スペース区切りで複数指定可）:\n"
            "  all        全ステップ実行（デフォルト: games→highlights）\n"
            "  games      Step1: RAW取得→datamart→JSON（活躍選手なし）\n"
            "  highlights Step2: 活躍選手選出→highlights.xlsx→datamart更新→JSON更新\n"
            "  datamart   datamart再生成（活躍選手なし・RAW再利用）\n"
            "  例) --steps games highlights"
        ),
    )
    args = parser.parse_args()

    try:
        date_list = _parse_date_arg(args.date)
    except ValueError as e:
        logger.error(f"--date の形式エラー: {e}"); return

    # ── ステップ解析 ──
    raw_steps = [s.strip().lower() for s in args.steps]
    run_all        = "all"        in raw_steps
    run_games      = run_all or "games"      in raw_steps
    run_highlights = run_all or "highlights" in raw_steps
    run_datamart   = run_all or "datamart"   in raw_steps
    # games は datamart を内包するので重複実行しない
    # highlights は games 完了後に単独で使う想定
    if run_games and run_datamart:
        print("  [INFO] games が datamart を内包 → datamart はスキップ")
        run_datamart = False
    # gamesとhighlightsは別ステップ: games後にhighlightsも実行する（--steps all含む）

    label = date_list[0] if len(date_list)==1 else f"{date_list[0]} 〜 {date_list[-1]}"
    # ステップ名を日本語に
    STEP_NAMES = {
        "games":      "Step1 RAWデータ取得 → datamart → JSON",
        "highlights": "Step2 活躍選手選出 → datamart更新 → JSON更新",
        "datamart":   "datamart再生成 → JSON",
        "all":        "全ステップ",
    }
    steps_label = " → ".join(STEP_NAMES.get(s, s) for s in raw_steps)

    print("=" * 55)
    print(f"  MLB データパイプライン開始")
    print(f"  日付  : {label}  ({len(date_list)}日分)")
    print(f"  ステップ: {steps_label}")
    print("=" * 55)

    results = {}

    # ── 日付ループ ──
    for date in date_list:
        if len(date_list) > 1:
            print(f"\n{'─' * 55}")
            print(f"▶ {date}")
            print(f"{'─' * 55}")
        else:
            print(f"\n▶ {date}")

        set_dirs(date, args.game_type)
        make_output_dirs()

        needs_raw = run_games or run_datamart or run_highlights

        # statsapi速報キャッシュ → Statcast自動差し替えチェック
        _maybe_upgrade_statsapi_cache(date)

        # RAW読み込み / Statcast取得
        df = pd.DataFrame()
        if needs_raw:
            print(f"\n--- RAWデータ取得 ---")
            raw_df = fetch_statcast(date)
            if raw_df.empty:
                print(f"  [WARN] データなし → スキップ")
                continue
            df = preprocess(raw_df)
            n_games = df["game_pk"].nunique()
            print(f"  取得完了: {len(df)}投球 / {n_games}試合")

        # ── Step games ──
        if run_games:
            print(f"\n--- {STEP_NAMES['games']} ---")
            path = run_games_datamart(df, date,
                                      prompts_path=args.prompts_path,
                                      with_highlights=False)
            if path:
                results.setdefault(date, {})["datamart"] = path
                print(f"  完了: {os.path.basename(path)}")

        # ── Step highlights ──
        if run_highlights:
            print(f"\n--- {STEP_NAMES['highlights']} ---")
            run_highlights_only(df, date, prompts_path=args.prompts_path)
            print(f"  完了")

        # ── Step datamart ──
        if run_datamart:
            print(f"\n--- {STEP_NAMES['datamart']} ---")
            path = run_games_datamart(df, date,
                                      prompts_path=args.prompts_path,
                                      with_highlights=False)
            if path:
                results.setdefault(date, {})["datamart"] = path
                print(f"  完了: {os.path.basename(path)}")

    # ── 完了サマリー ──
    print("\n" + "=" * 55)
    print("✅ 完了!")
    for date, r in results.items():
        dm = r.get("datamart", "")
        if dm:
            kb = os.path.getsize(dm) // 1024 if os.path.exists(dm) else 0
            json_path = os.path.join(
                os.path.dirname(dm).replace("datamart", "json"),
                f"{date}.json"
            )
            json_kb = os.path.getsize(json_path) // 1024 if os.path.exists(json_path) else 0
            print(f"  {date}  datamart: {kb}KB  /  JSON: {json_kb}KB")
    print("=" * 55)

if __name__ == "__main__":
    main()

# %%
# ==================================================
# Jupyter / Colab での手動実行例:
# ==================================================
#
# from mlb_run import *
#
# set_dirs("2026-04-10", "公式戦")
# make_output_dirs()
#
# raw_df = fetch_statcast("2026-04-10")
# df = preprocess(raw_df)
#
# # games datamart（活躍選手: Claude API + prompts/活躍選手_MLB.txt）
# # GEMINI_API_KEY を環境変数にセット済みであること
# run_games_datamart(df, "2026-04-10", prompts_path=str(_SCRIPT_DIR / "prompts" / "活躍選手_MLB.txt"))