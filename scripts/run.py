# %%
# ==================================================
# NPB データ収集 → データマート 一括生成パイプライン
#
# ステップ順:
#   Step 1: games      試合データ取得    → raw/{date}/all_games_{date}.xlsx
#   Step 2: pitch      投球データ取得    → raw/{date}/daily_pitch_data_{date}.xlsx
#   Step 3: highlights 活躍選手選出     → raw/{date}/highlights_{date}.xlsx
#   Step 4: datamart   データマート&JSON作成
#             datamart → games/datamart/{date}.xlsx
#             json     → games/json/{date}.json
#
# 空欄のまま残した指標と理由:
#   - 天候                    : スクレイピングデータに含まれない
#   - ホーム/アウェイ勝敗成績  : 取得困難（仕様シート注釈通り）
#   - 被打球速度(hard_hit)    : 打球速度データなし
#   - 対右打者/対左打者スプリット: 打者の左右データが全NaN → 計算不可
# ==================================================

# %%
# ==================================================
# Section 0. セットアップ（Colabなど、初回のみ実行）
# ==================================================
# !pip install requests beautifulsoup4 openpyxl --break-system-packages

# %%
# ==================================================
# Section 1. ライブラリのインポート
# ==================================================
import sys
import re
import os
import time
import json
import math
from collections import defaultdict
from pathlib import Path

# スクリプト自身の場所を基準にしたルートディレクトリ
# → どこから実行しても prompts/.env/data へのパスが正しく解決される
# scripts/ の親 = 野球ダッシュボード/ をプロジェクトルートとする
_SCRIPT_DIR = Path(__file__).resolve().parent.parent

import io
import logging
import datetime

import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np

# Gemini（活躍選手ステップで使用。未インストールの場合はスキップ）
try:
    import google.generativeai as genai
    from dotenv import load_dotenv
    _GENAI_AVAILABLE = True
except ImportError:
    _GENAI_AVAILABLE = False

# %%
# ==================================================
# Section 2. 設定（TARGET_DATEだけ変えれば全工程に反映）
# ==================================================

TARGET_DATE = "2026-04-19"

BASE_URL      = "https://baseball.yahoo.co.jp/npb"
SCHEDULE_URL  = f"{BASE_URL}/schedule/first/all?date={TARGET_DATE}"
FARM_URL      = f"{BASE_URL}/schedule/farm/all?date={TARGET_DATE}"
HEADERS       = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# ── 試合種別マップ ──
# 1軍
ICHI_GAME_TYPES = {
    "レギュラーシーズン": "all",        # リーグ戦＋交流戦をまとめて "all" で取得
    "オープン戦":         "open",
    "ポストシーズン":     "climax",     # CS・日本シリーズ
}
# 2軍
NI_GAME_TYPES = {
    "公式戦":             "official",
    "春季教育リーグ":     "spring",
    "フレッシュオールスター": "fresh_allstar",
    "日本選手権":         "nippon",
    "フェニックスリーグ": "phoenix",
}

def _game_type_schedule_path(league: str, game_type: str) -> str:
    """league + game_type → Yahoo Baseball スケジュールパス（first/all 等）"""
    if league == "ichi":
        slug = ICHI_GAME_TYPES.get(game_type, "all")
        return f"first/{slug}"
    else:
        slug = NI_GAME_TYPES.get(game_type, "official")
        return f"farm/{slug}"

# フォルダ構成:
#   data/プロ野球/{YYYY}年/1軍|2軍/{試合種別}/   ← raw/datamartなど非公開の作業データ（gitでバックアップ管理はするが、Pagesでは公開しない）
#     raw/{YYYYMMDD}/
#     games/datamart/
#     players/datamart/投手|野手/
#   docs/data/プロ野球/{YYYY}年/1軍|2軍/{試合種別}/  ← ダッシュボードが実際にfetchする公開データ（GitHub Pagesの公開元）
#     games/json/
#     season/
#     players/json/投手|野手/
BASE_DATA_DIR   = str(_SCRIPT_DIR / "data" / "プロ野球")
BASE_PUBLIC_DIR = str(_SCRIPT_DIR / "docs" / "data" / "プロ野球")

# 実行時に set_league_dirs() で上書きされる（デフォルト: 1軍レギュラーシーズン）
_year          = TARGET_DATE[:4]
_game_type     = "レギュラーシーズン"
_type_base        = os.path.join(BASE_DATA_DIR,   f"{_year}年", "1軍", _game_type)
_type_base_public = os.path.join(BASE_PUBLIC_DIR, f"{_year}年", "1軍", _game_type)
RAW_DIR             = os.path.join(_type_base, "raw", TARGET_DATE)
GAMES_DM_DIR        = os.path.join(_type_base, "games", "datamart")
GAMES_JSON_DIR      = os.path.join(_type_base_public, "games", "json")
SEASON_DIR          = os.path.join(_type_base_public, "season")


# 後方互換
DATAMART_DIR = GAMES_DM_DIR
JSON_DIR     = GAMES_JSON_DIR

_current_league    = "ichi"
_current_game_type = "レギュラーシーズン"

def set_league_dirs(league: str, date: str = TARGET_DATE, game_type: str | None = None) -> None:
    """
    league / game_type に応じてグローバルなディレクトリ変数を切り替える。
    league:    "ichi"（1軍） / "ni"（2軍）
    game_type: 1軍→"レギュラーシーズン"/"オープン戦"/"ポストシーズン"
               2軍→"公式戦"/"春季教育リーグ"/"フレッシュオールスター"/"日本選手権"/"フェニックスリーグ"
    """
    global RAW_DIR, GAMES_DM_DIR, GAMES_JSON_DIR, SEASON_DIR, DATAMART_DIR, JSON_DIR, \
           SCHEDULE_URL, TARGET_DATE, _current_league, _current_game_type
    _current_league = league
    league_label = "1軍" if league == "ichi" else "2軍"
    TARGET_DATE  = date
    year         = date[:4]

    if game_type is None:
        game_type = "レギュラーシーズン" if league == "ichi" else "公式戦"
    _current_game_type = game_type

    sched_path = _game_type_schedule_path(league, game_type)
    SCHEDULE_URL = f"{BASE_URL}/schedule/{sched_path}?date={date}"

    # data/プロ野球/{YYYY}年/1軍|2軍/{試合種別}/       ← 非公開（raw/datamart）
    # docs/data/プロ野球/{YYYY}年/1軍|2軍/{試合種別}/  ← 公開（games/json・season、Pagesの公開元）
    type_base        = os.path.join(BASE_DATA_DIR,   f"{year}年", league_label, game_type)
    type_base_public = os.path.join(BASE_PUBLIC_DIR, f"{year}年", league_label, game_type)
    RAW_DIR          = os.path.join(type_base, "raw", date)
    GAMES_DM_DIR     = os.path.join(type_base, "games", "datamart")
    GAMES_JSON_DIR   = os.path.join(type_base_public, "games", "json")
    SEASON_DIR       = os.path.join(type_base_public, "season")
    DATAMART_DIR     = GAMES_DM_DIR
    JSON_DIR         = GAMES_JSON_DIR

    import sys as _sys
    _mod = _sys.modules[__name__]
    _mod.SCHEDULE_URL            = SCHEDULE_URL
    _mod.TARGET_DATE             = TARGET_DATE
    _mod.RAW_DIR                 = RAW_DIR
    _mod.GAMES_DM_DIR            = GAMES_DM_DIR
    _mod.GAMES_JSON_DIR          = GAMES_JSON_DIR
    _mod.SEASON_DIR              = SEASON_DIR
    _mod.DATAMART_DIR            = GAMES_DM_DIR
    _mod.JSON_DIR                = GAMES_JSON_DIR
    _mod._current_league         = league
    _mod._current_game_type      = game_type

# Step 4: 活躍選手 Gemini 設定
# GEMINI_API_KEY: .env ファイルに GEMINI_API_KEY=xxx を記載するか環境変数で設定
GEMINI_MODEL             = "gemini-2.5-flash"
GEMINI_MAX_OUTPUT_TOKENS = 25000
GEMINI_TEMPERATURE       = 0
# プロンプト・スキーマファイルのパス（プロジェクトルート相対）
HIGHLIGHTS_PROMPT_FILE = str(_SCRIPT_DIR / "prompts" / "活躍選手.txt")
HIGHLIGHTS_SCHEMA_FILE = str(_SCRIPT_DIR / "schemas" / "活躍選手_schema.txt")


# チーム名辞書
# 1軍チームIDマップ
ICHI_TEAM_NAMES = {
    "1": "巨人", "2": "ヤクルト", "3": "DeNA", "4": "中日",
    "5": "阪神", "6": "広島", "7": "西武", "8": "日本ハム",
    "9": "ロッテ", "11": "オリックス", "12": "ソフトバンク", "376": "楽天",
}
# 2軍チームIDマップ
NI_TEAM_NAMES = {
    "23": "日本ハム", "24": "西武",  "25": "DeNA",    "26": "巨人",   "27": "ロッテ",
    "28": "ヤクルト",  # Yahoo2軍ID
    "29": "オリックス", "30": "ソフトバンク", "32": "広島", "33": "中日",
    "34": "阪神",  "35": "日本ハム", "36": "ヤクルト", "377": "楽天",
    "23879": "ハヤテ",  # ハヤテ223（2軍独立参加チーム）
}
# 統合（チームID解決用、スクレイピング時に参照）
TEAM_NAMES = {**ICHI_TEAM_NAMES, **NI_TEAM_NAMES}

# ランナー状態辞書
RUNNER_DICT = {
    "b000": "なし", "b100": "1塁", "b010": "2塁", "b001": "3塁",
    "b110": "1・2塁", "b101": "1・3塁", "b011": "2・3塁", "b111": "満塁",
}

# 投球判定カテゴリ辞書
BALL_CAT_DICT = {
    "ball1": "ストライク/ファウル系", "ball2": "ボール系", "ball3": "アウト系",
    "ball4": "出塁/ヒット系", "ball5": "犠打/犠飛系",
}

# コース座標のストライクゾーン境界（Yahoo Baseball 投球チャート座標系）
# ストライクゾーン境界（Yahoo Baseball 投球チャート座標系）
# ※ 実データ検証済み: Top[14.40, 53.64] Left[10.08, 41.52]
ZONE_TOP_MIN, ZONE_TOP_MAX   = 14.40, 53.64
ZONE_LEFT_MIN, ZONE_LEFT_MAX = 10.08, 41.52

# 9分割境界（実データ検証済み: le1=20.16, le2=31.04, te1=27.20, te2=39.32）
ZONE_LE1 = 20.16   # Left: Out/中央 境界（14×1.44px）
ZONE_LE2 = 31.04   # Left: 中央/In 境界（ゾーン幅2/3点）
ZONE_TE1 = 27.20   # Top:  高め/真中 境界（検証済み）
ZONE_TE2 = 39.32   # Top:  真中/低め 境界（検証済み）

# Heart/Shadow/Chase/Waste 境界
HEART_L_MIN = ZONE_LE1;  HEART_L_MAX = ZONE_LE2
HEART_T_MIN = ZONE_TE1;  HEART_T_MAX = ZONE_TE2
SHADOW_EXT  = 5.0   # ゾーン境界からの Shadow 拡張幅
CHASE_EXT   = 10.0  # Shadow 境界からの Chase 拡張幅

# %%
# ==================================================
# Section 3. 共通ユーティリティ
# ==================================================

def make_output_dir() -> None:
    for d in [RAW_DIR, GAMES_DM_DIR, GAMES_JSON_DIR, SEASON_DIR]:
        os.makedirs(d, exist_ok=True)

def get_soup(url: str, retries: int = 5, backoff: float = 2.0) -> BeautifulSoup | None:
    """GETしてBeautifulSoupを返す。失敗時は指数バックオフでリトライ。404はリトライしない。"""
    for attempt in range(retries):
        try:
            res = requests.get(url, headers=HEADERS, timeout=15)
            # 404 はリトライ不要（URLが存在しない）
            if res.status_code == 404:
                return None
            res.raise_for_status()
            return BeautifulSoup(res.content, "html.parser")
        except Exception as e:
            wait = backoff ** attempt  # 1秒 → 2秒 → 4秒
            if attempt < retries - 1:
                print(f"\n通信エラー ({url}): {e}  → {wait:.0f}秒後にリトライ ({attempt+1}/{retries-1}回)")
                time.sleep(wait)
            else:
                print(f"\n通信エラー ({url}): {e}  → リトライ上限に達しました")
    return None

def get_text(element, selector: str, default: str = "") -> str:
    tag = element.select_one(selector)
    return tag.text.strip() if tag else default

def clean_text(text: str) -> str:
    return text.strip().replace("\n", "").replace("  ", " ")

def clean_column_name(name: str) -> str:
    return name.replace(" ", "").replace("　", "")

def clean_str(val) -> str:
    """NaN / nan / - は空文字に変換"""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s in ("nan", "NaN", "-") else s

def make_seq_id(prefix: str, n: int, width: int = 5) -> list[str]:
    return [f"{prefix}{str(i+1).zfill(width)}" for i in range(n)]

def parse_vel(val) -> float | None:
    if pd.isna(val): return None
    m = re.search(r"([\d\.]+)", str(val))
    return float(m.group(1)) if m else None

def to_num(val, default=np.nan):
    try: return float(str(val).strip())
    except (ValueError, TypeError): return default

def innings_to_float(val) -> float | None:
    if pd.isna(val): return None
    try:
        s = str(val).strip()
        if "." in s:
            w, f = s.split(".", 1)
            return int(w) + int(f) / 3
        return float(s)
    except Exception: return None

def build_innings_str(df_sb: pd.DataFrame, game_id, team: str) -> str:
    row = df_sb[(df_sb["試合ID"] == game_id) & (df_sb["チーム"] == team)]
    if row.empty: return ""
    inn_cols = [c for c in df_sb.columns if re.match(r"^\d+回$", str(c))]
    return ",".join(row.iloc[0][inn_cols].fillna(0).astype(str).tolist())

def get_side(game_id, team, home_team_map: dict) -> str:
    home = home_team_map.get(game_id, "")
    # TEAM_NAMES 逆引きで別名も照合（2軍で同チームに複数IDがある場合）
    if team == home:
        return "home"
    # 同じチーム名でも表記揺れに対応
    if team and home and team.replace(" ", "").replace("　", "") == home.replace(" ", "").replace("　", ""):
        return "home"
    return "away"

def get_game_ids() -> list[str]:
    print(f"[{TARGET_DATE}] のスケジュールから試合IDを抽出中...")
    soup = get_soup(SCHEDULE_URL)
    if not soup: return []
    gm_card = soup.select_one("#gm_card")
    if not gm_card: return []
    game_ids = []
    for a in gm_card.find_all("a", href=True):
        m = re.search(r'/npb/game/(\d+)/', a["href"])
        if m: game_ids.append(m.group(1))
    return list(dict.fromkeys(game_ids))

# %%
# ==================================================
# Section 4. Step 1 ── 試合データ取得
#   出力: all_games_{TARGET_DATE}.xlsx
# ==================================================

def _parse_game_info(soup, game_id: str) -> dict:
    info = {"試合ID": game_id}

    round_info = soup.select_one(".bb-gameRound")
    if round_info:
        info["試合情報"] = " ".join(
            tag.text.strip() for tag in round_info.select("li, p, div") if tag.text.strip()
        )

    desc_left = soup.select_one(".bb-gameDescription__left")
    if desc_left:
        raw = desc_left.get_text(strip=True)
        tm = re.search(r'(\d{1,2}:\d{2})', raw)
        if tm:
            info["開始時間"] = tm.group(1)
            info["球場"]    = raw[tm.end():]
        else:
            info["開始時間"] = ""
            info["球場"]    = raw

    # チーム名取得（後攻=HOME, 先攻=AWAY）
    teams = soup.select(".bb-gameTeam__name")
    if len(teams) >= 2:
        # 「後攻」「先攻」テキストで確認（bb-gameTeam__role等）
        roles = [el.text.strip() for el in soup.select(".bb-gameTeam__role, .bb-gameDescription__team--role")]
        if "後攻" in roles and "先攻" in roles:
            # 後攻=HOME, 先攻=AWAY の並び順で取得
            team_sections = soup.select(".bb-gameTeam")
            home_name, away_name = "", ""
            for sec in team_sections:
                role_el = sec.select_one(".bb-gameTeam__role, .bb-gameDescription__team--role")
                name_el = sec.select_one(".bb-gameTeam__name")
                if role_el and name_el:
                    role_txt = role_el.text.strip()
                    name_txt = name_el.text.strip()
                    if "後攻" in role_txt:
                        home_name = name_txt
                    elif "先攻" in role_txt:
                        away_name = name_txt
            if home_name and away_name:
                info["ホームチーム"]   = home_name
                info["アウェイチーム"] = away_name
            else:
                # フォールバック: teams[0]=後攻=HOME
                info["ホームチーム"]   = teams[0].text.strip()
                info["アウェイチーム"] = teams[1].text.strip()
        else:
            # 従来通り teams[0]=ホーム
            info["ホームチーム"]   = teams[0].text.strip()
            info["アウェイチーム"] = teams[1].text.strip()

    scores = soup.select(".bb-gameTeam__score .bb-gameCard__detail span")
    if len(scores) >= 3:
        info["ホーム得点"]   = scores[0].text.strip()
        info["アウェイ得点"] = scores[2].text.strip()

    state = soup.select_one(".bb-gameCard__state")
    if state: info["試合状態"] = state.text.strip()

    for th in soup.select("table.bb-tableLeft th"):
        if "観客数" in th.text or "試合時間" in th.text:
            td = th.find_next_sibling("td")
            if td: info[th.text.strip()] = td.text.strip()

    # 試合中止チェック：専用のステータス要素のみで判定（ページ全体テキストは誤検知が多いため不使用）
    cancel_el = soup.select_one(".bb-gameCard__state, .bb-gameDescription__state, .bb-gameStatus")
    cancel_text = cancel_el.text.strip() if cancel_el else ""
    if "中止" in cancel_text or "ノーゲーム" in cancel_text:
        info["試合状態"] = cancel_text

    recap = soup.select_one("#async-recap .bb-paragraph")
    if recap: info["戦評"] = recap.text.strip()

    mep_sec = soup.select_one("#mep")
    if mep_sec:
        mep_player = mep_sec.select_one(".bb-tableTeamHead__player")
        if mep_player: info["エキサイティングプレーヤー"] = mep_player.text.strip()
        for tr in mep_sec.select("table.bb-tableLeft tr"):
            th_tag = tr.select_one("th")
            td_tag = tr.select_one("td")
            if th_tag and td_tag:
                info[f"MEP_{th_tag.text.strip()}"] = td_tag.text.strip().replace("\n", " ").replace("  ", " ")

    return info


def _parse_score_board(soup, game_id: str) -> list[dict]:
    rows = []
    board_table = soup.select_one("#ing_brd")
    if not board_table: return rows

    for tr in board_table.select("tbody tr.bb-gameScoreTable__row"):
        row_data = [td.text.strip() for td in tr.select("td")]
        if not row_data: continue
        team    = row_data[0]
        innings = row_data[1:-3]
        inning_list = [int(x) if x.isdigit() else 0 for x in innings]
        row_dict = {"試合ID": game_id, "チーム": team}
        for i, inn in enumerate(innings, 1):
            row_dict[f"{i}回"] = inn
        row_dict["計"] = row_data[-3]
        row_dict["安"] = row_data[-2]
        row_dict["失"] = row_data[-1]
        row_dict["イニング得点リスト"] = str(inning_list)
        rows.append(row_dict)
    return rows


def _parse_score_plays(soup, game_id: str) -> list[dict]:
    rows = []
    scor_ply = soup.select_one("#scor_ply")
    if not scor_ply: return rows

    for tr in scor_ply.select("tbody tr"):
        inning = get_text(tr, "th")
        td = tr.select_one("td")
        if not td: continue
        rows.append({
            "試合ID": game_id, "イニング": inning,
            "チーム": get_text(td, ".bb-gameTable__team"),
            "打順":   get_text(td, ".bb-gameTable__order"),
            "打者":   get_text(td, ".bb-gameTable__player"),
            "状況":   get_text(td, ".bb-gameTable__state"),
            "詳細":   " / ".join(p.text.strip().replace("\n", " ") for p in td.select(".bb-gameTable__summary")),
        })
    return rows


def _parse_starters(soup, game_id: str) -> list[dict]:
    rows = []
    strt_mem = soup.select_one("#strt_mem")
    if not strt_mem: return rows

    for team_sec in strt_mem.select(".bb-splits__item"):
        team_name = get_text(team_sec, "h1")
        for table in team_sec.select("table"):
            for tr in table.select("tbody tr.bb-splitsTable__row"):
                tds = [td.text.strip().replace("\n", "") for td in tr.select("td")]
                if len(tds) >= 6:
                    rows.append({
                        "試合ID": game_id, "チーム": team_name,
                        "打順/投手": tds[0], "位置": tds[1], "選手名": tds[2],
                        "投/打": tds[3], "防御率/打率": tds[4], "調子": tds[5],
                    })
    return rows


def _parse_batter_stats(soup_stats, game_id: str, home_team: str = ""):
    batter_list, headers = [], []
    for table in soup_stats.select(".bb-statsTable"):
        if not table.select("tr.bb-statsTable__row--total"): continue
        team_class = [c for c in table.get("class", []) if "npbTeam" in c]
        if not team_class: continue
        import re as _re
        m = _re.search(r'npbTeam(\d+)', team_class[0])
        team_id   = m.group(1) if m else team_class[0]
        team_name = TEAM_NAMES.get(team_id, f"Team_{team_id}")
        side = "home" if (home_team and team_name == home_team) else "away"
        if not headers:
            headers = ["試合ID", "チーム", "ホーム/アウェイ"] + [th.text.strip() for th in table.select("thead th")]
        for tr in table.select("tbody tr.bb-statsTable__row"):
            row_data = [game_id, team_name, side]
            for td in tr.select("td"):
                text   = clean_text(td.text)
                detail = td.select_one(".bb-statsTable__dataDetail")
                if detail: text = detail.text.strip()
                row_data.append(text)
            if len(row_data) == len(headers):
                batter_list.append(row_data)
    return headers, batter_list


def _parse_pitcher_stats(soup_stats, game_id: str, home_team: str = ""):
    pitcher_list, headers = [], []
    for sec in soup_stats.select(".bb-scoreTable"):
        team_class = [c for c in sec.get("class", []) if "npbTeam" in c]
        if not team_class: continue
        import re as _re
        m = _re.search(r'npbTeam(\d+)', team_class[0])
        team_id   = m.group(1) if m else team_class[0]
        team_name = TEAM_NAMES.get(team_id, f"Team_{team_id}")
        raw_headers = [th.text.strip() for th in sec.select("thead th")]
        col_names   = ["勝敗成績" if h == "" else h for h in raw_headers]
        if not headers:
            # 先頭に「役割」列を追加
            headers = ["試合ID", "チーム", "ホーム/アウェイ", "役割"] + col_names
        side = "home" if (home_team and team_name == home_team) else "away"
        rows_in_sec = sec.select("tbody tr")
        for i, tr in enumerate(rows_in_sec):
            role = "先発" if i == 0 else "中継ぎ"
            row_data = [game_id, team_name, side, role] + [clean_text(td.text) for td in tr.select("td")]
            if len(row_data) == len(headers):
                pitcher_list.append(row_data)
    return headers, pitcher_list


def scrape_game_data(game_id: str) -> dict | None:
    top_url   = f"{BASE_URL}/game/{game_id}/top"
    stats_url = f"{BASE_URL}/game/{game_id}/stats"

    print(f"  -> 試合データ取得中: {game_id}")
    soup_top = get_soup(top_url)
    if not soup_top:
        print(f"     [スキップ] 試合情報が取得できませんでした: {game_id}")
        return None

    game_info = _parse_game_info(soup_top, game_id)
    # 試合中止の場合は基本情報のみ保存してスタメン・成績はスキップ
    is_cancelled = "中止" in str(game_info.get("試合状態", ""))
    if is_cancelled:
        print(f"     [中止] 試合ID {game_id} は試合中止のため成績をスキップ")

    game_data = {
        "試合基本情報": pd.DataFrame([game_info]),
        "スコアボード":    pd.DataFrame() if is_cancelled else pd.DataFrame(_parse_score_board(soup_top, game_id)),
        "スコアプレー詳細": pd.DataFrame() if is_cancelled else pd.DataFrame(_parse_score_plays(soup_top, game_id)),
        "スタメン":        pd.DataFrame(_parse_starters(soup_top, game_id)),  # 中止でもスタメンは取得試みる
    }

    # ホームチームを取得（試合情報から）
    home_team = game_info.get("ホームチーム", "")

    time.sleep(1)
    soup_stats = get_soup(stats_url)
    if soup_stats:
        bat_headers, bat_rows = _parse_batter_stats(soup_stats, game_id, home_team)
        game_data["打撃成績"] = pd.DataFrame(bat_rows, columns=bat_headers) if bat_rows else pd.DataFrame()
        pit_headers, pit_rows = _parse_pitcher_stats(soup_stats, game_id, home_team)
        game_data["投手成績"] = pd.DataFrame(pit_rows, columns=pit_headers) if pit_rows else pd.DataFrame()

    return game_data


def run_game_scraper() -> str:
    """試合データを取得して all_games_{TARGET_DATE}.xlsx に保存。パスを返す。"""
    game_ids = get_game_ids()
    if not game_ids:
        print(f"[{TARGET_DATE}] の試合は見つかりませんでした。")
        return ""

    print(f"合計 {len(game_ids)} 試合のIDを取得しました: {game_ids}")
    print("-" * 40)

    all_data = {sheet: pd.DataFrame() for sheet in
                ["試合基本情報", "スコアボード", "スコアプレー詳細", "スタメン", "打撃成績", "投手成績"]}

    for gid in game_ids:
        single = scrape_game_data(gid)
        if single:
            for sheet_name, df in single.items():
                all_data[sheet_name] = pd.concat([all_data[sheet_name], df], ignore_index=True)
        time.sleep(2)

    make_output_dir()
    output_path = os.path.join(RAW_DIR, f"all_games_{TARGET_DATE}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in all_data.items():
            if not df.empty:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"完了: '{output_path}'")
    return output_path


# ファーム用の関数は league 切り替えで統合済み


# %%
# ==================================================
# Section 5. シーズン成績スクレイピング（内部ヘルパー）
#   run_datamart() から呼ばれ、DataFrameを直接返す（ファイル出力なし）
# ==================================================

# npb.jp 2軍チームコードマップ
NPB_FARM_TEAM_CODES = {
    "g": "巨人", "t": "阪神", "db": "DeNA", "d": "中日",
    "c": "広島", "s": "ヤクルト",
    "h": "ソフトバンク", "f": "日本ハム", "b": "オリックス",
    "e": "楽天", "l": "西武", "m": "ロッテ",
}
NPB_BASE = "https://npb.jp"


def _parse_situation(soup) -> dict:
    inning_full = get_text(soup, "h4.live em")
    inning_num  = re.search(r'\d+回', inning_full)
    top_bottom  = "表" if "表" in inning_full else "裏" if "裏" in inning_full else ""

    score_text, attack_team = "", ""
    score_div = soup.select_one(".score table")
    if score_div:
        rows = score_div.select("tr")
        if len(rows) >= 2:
            t1_name = get_text(rows[0], ".nm")
            s1 = rows[0].select("td")[-1].text.strip()
            t2_name = get_text(rows[1], ".nm")
            s2 = rows[1].select("td")[-1].text.strip()
            score_text = f"{t1_name} {s1}-{s2} {t2_name}"
            nm0 = rows[0].select_one(".nm")
            nm1 = rows[1].select_one(".nm")
            if nm0 and "act" in nm0.get("class", []): attack_team = t1_name
            elif nm1 and "act" in nm1.get("class", []): attack_team = t2_name

    out_count    = soup.select_one("p.o b").text.count("●") if soup.select_one("p.o b") else 0
    runner_state = "なし"
    base_div     = soup.select_one("#base")
    if base_div:
        for c in base_div.get("class", []):
            if c in RUNNER_DICT: runner_state = RUNNER_DICT[c]; break

    return {
        "イニング": inning_num.group(0) if inning_num else "",
        "表/裏": top_bottom, "スコア": score_text,
        "攻撃チーム": attack_team, "アウト数": out_count, "ランナー": runner_state,
    }


def _parse_pitcher_info(soup) -> dict:
    info = {"投手名": "", "投左右": "", "投手_対戦打者数": "", "投手_防御率": ""}
    pit_div = soup.select_one("#pit")
    if not pit_div: return info
    info["投手名"]       = get_text(pit_div, ".nm a")
    info["投左右"]       = get_text(pit_div, ".dominantHand")
    p_scores = pit_div.select("tr.score td")
    if len(p_scores) >= 3:
        info["投手_対戦打者数"] = p_scores[1].text.strip()
        info["投手_防御率"]     = p_scores[2].text.strip()
    return info


def _parse_batter_info(soup) -> dict:
    info = {"打者名": "", "打左右": "", "打者_今季成績": "", "打者_本日の履歴": ""}
    bat_div = soup.select_one("#batter")
    if not bat_div: return info
    info["打者名"]       = get_text(bat_div, ".nm a")
    info["打左右"]       = get_text(bat_div, ".dominantHand")
    info["打者_今季成績"] = get_text(bat_div, "td.rate")
    anda_td = bat_div.select_one("td.anda")
    if anda_td:
        info["打者_本日の履歴"] = "、".join(s.text.strip() for s in anda_td.select("span") if s.text.strip())
    return info


def _parse_courses(soup) -> dict:
    courses = {}
    chart = soup.select_one(".bb-allocationChart")
    if not chart: return courses
    for ball in chart.select("span.bb-icon__ballCircle"):
        style   = ball.get("style", "")
        cat_key = next((bc.split("--")[1] for bc in ball.get("class", []) if "ballCircle--ball" in bc), "")
        num_tag = ball.select_one("span.bb-icon__number")
        num     = num_tag.text.strip() if num_tag else ""
        if num:
            top_m  = re.search(r'top:([\d\.]+)px', style)
            left_m = re.search(r'left:([\d\.]+)px', style)
            courses[num] = {
                "top":      top_m.group(1)  if top_m  else "",
                "left":     left_m.group(1) if left_m else "",
                "category": BALL_CAT_DICT.get(cat_key, ""),
            }
    return courses


def scrape_all_pitches_of_game(game_id: str) -> list[dict]:
    print(f"\n>>> 試合ID: {game_id} の全投球データ取得を開始します")
    pitches        = []
    current_index  = "0110100"
    consec_fail    = 0   # 連続失敗カウンタ
    MAX_FAIL       = 3   # 連続N回失敗したら中断

    while current_index:
        url = f"{BASE_URL}/game/{game_id}/score?index={current_index}"
        sys.stdout.write(f"\r  取得中... Index: {current_index}")
        sys.stdout.flush()

        soup = get_soup(url)
        if not soup:
            consec_fail += 1
            if consec_fail >= MAX_FAIL:
                print(f"\n  [WARN] 試合ID {game_id}: 連続{MAX_FAIL}回失敗のため中断します（取得済み: {len(pitches)}球）")
                break
            # 失敗してもnext_indexが不明なため中断せざるを得ない
            break

        situation      = _parse_situation(soup)
        pitcher        = _parse_pitcher_info(soup)
        batter         = _parse_batter_info(soup)
        courses        = _parse_courses(soup)
        next_batter    = get_text(soup, ".nextBatter dd a p")
        result_summary = get_text(soup, "#result span")

        target_table = next(
            (t for t in soup.select("table.bb-splitsTable")
             if t.select_one("th") and "投球数" in t.select_one("th").text),
            None,
        )

        if target_table:
            for tr in target_table.select("tbody tr"):
                tds = tr.select("td")
                if len(tds) < 5: continue
                ball_num = tds[0].text.strip()
                c_info   = courses.get(ball_num, {})
                pitches.append({
                    **situation, **pitcher, **batter,
                    "試合ID": game_id, "次打者": next_batter,
                    "打席内球数": ball_num, "通算投球数": tds[1].text.strip(),
                    "球種": tds[2].text.strip(), "球速": tds[3].text.strip(),
                    "コース(Top)":  c_info.get("top", ""),
                    "コース(Left)": c_info.get("left", ""),
                    "判定カテゴリ": c_info.get("category", ""),
                    "1球結果": clean_text(tds[4].text), "打席完了結果": result_summary,
                })

        next_btn      = soup.select_one("a#btn_next")
        current_index = next_btn["index"] if (next_btn and "index" in next_btn.attrs) else None
        time.sleep(1)

    print(f"\n  取得完了（{len(pitches)}球）")
    return pitches


def run_pitch_scraper(target_game_ids: list[str] | None = None) -> str:
    """
    投球データを取得して daily_pitch_data_{TARGET_DATE}.xlsx に保存。

    Args:
        target_game_ids: 取得対象の試合IDリスト。
                         None の場合はスケジュールから全試合を取得。
                         指定した場合は既存 xlsx にマージ（再取得モード）。
    """
    output_path = os.path.join(RAW_DIR, f"daily_pitch_data_{TARGET_DATE}.xlsx")

    if target_game_ids:
        # 再取得モード: 指定した試合IDのみ取得して既存データにマージ
        game_ids   = target_game_ids
        retry_mode = True
        print(f"\n再取得モード: {len(game_ids)} 試合の投球データを再取得します。")
        print(f"  対象試合ID: {game_ids}")
    else:
        # 通常モード: スケジュールから全試合を取得
        game_ids = get_game_ids()
        if not game_ids:
            print("試合が見つかりませんでした。")
            return ""
        retry_mode = False
        print(f"本日（{TARGET_DATE}）の全 {len(game_ids)} 試合の投球データを取得します。")
        print("※完了まで長時間かかります。そのままお待ちください。")

    new_pitches = []
    failed_ids  = []
    for gid in game_ids:
        result = scrape_all_pitches_of_game(gid)
        if result:
            new_pitches.extend(result)
        else:
            failed_ids.append(gid)
        time.sleep(3)

    if not new_pitches:
        print("\nデータが取得できませんでした。")
        if failed_ids:
            print(f"失敗した試合ID: {failed_ids}")
            print(f"再実行: python run.py --steps pitch --game_ids {' '.join(failed_ids)}")
        # 既存ファイルがあれば返す（試合なし日でも後続ステップを継続）
        if os.path.exists(output_path):
            print(f"  既存ファイルを使用: {output_path}")
            return output_path
        return ""

    make_output_dir()

    if retry_mode and os.path.exists(output_path):
        # 既存データを読み込み、再取得した試合IDの行を置き換えてマージ
        df_existing = pd.read_excel(output_path, engine="openpyxl")
        # 再取得対象の試合IDを既存データから除去
        df_existing = df_existing[~df_existing["試合ID"].astype(str).isin(
            [str(g) for g in game_ids]
        )]
        df_new    = pd.DataFrame(new_pitches)
        df_merged = pd.concat([df_existing, df_new], ignore_index=True)
        # 試合IDの出現順でソート（既存の順序を維持）
        all_gids  = get_game_ids()
        if all_gids:
            gid_order = {str(g): i for i, g in enumerate(all_gids)}
            df_merged["_sort"] = df_merged["試合ID"].astype(str).map(
                lambda x: gid_order.get(x, 999)
            )
            df_merged = df_merged.sort_values("_sort").drop(columns=["_sort"])
        df = df_merged.reset_index(drop=True)
        print(f"\nマージ完了: 合計 {len(df)} 球（再取得: {len(pd.DataFrame(new_pitches))} 球）")
    else:
        df = pd.DataFrame(new_pitches)

    df.to_excel(output_path, index=False, engine="openpyxl")
    print(f"完了: 合計 {len(df)} 球のデータを '{output_path}' に保存しました。")

    if failed_ids:
        print(f"\n⚠️  取得失敗した試合ID: {failed_ids}")
        print(f"再実行コマンド: python run.py --steps pitch --game_ids {' '.join(failed_ids)}")

    return output_path


# %%
# ==================================================
# Section 7. Step 4 ── データマート生成
#   出力: npb_dashboard_datamart_{TARGET_DATE}.xlsx
# ==================================================

def calc_batted_stats(series: pd.Series) -> dict:
    """
    打席完了結果（テキスト）から打球を分類・集計する。

    【分類定義】
    GB  : ゴロアウト + 内野安打 + エラー（犠打・捕犠打野選は除外）
    LD  : ライナーアウト
    FB  : フライアウト + ファウルフライ + 本塁打 + 犠飛
    IFFB: 内野フライアウト + ファウルフライ（FBの部分集合）
    HR  : 本塁打
    判断不可打球: 右安打・左安打・中安打・各方向の長打（軌道不明）

    除外: 犠打, 捕犠打野選, 守備妨害, 打撃妨害, ノイズ（ボール/見逃し/けん制）

    【計算式】
    全打球数 = GB + LD + FB + 判断不可打球
    GB%   = GB   / (GB + LD + FB)
    LD%   = LD   / (GB + LD + FB)
    FB%   = FB   / (GB + LD + FB)
    HR%   = HR   / (GB + LD + FB)
    IFFB% = IFFB / FB
    """
    s = series.fillna("").astype(str)

    # ── GB：ゴロアウト + 内野安打 + エラー ──
    # ゴロアウト（犠打・捕犠打野選は除外）
    is_gb_out = (
        s.str.contains(r"[一二三遊投捕]ゴロ|[一二三遊投捕]併殺打|[一二三遊投捕]野選", regex=True) &
        ~s.str.contains(r"安打|犠打野選", regex=True)
    )
    # 内野安打（一安打・三安打・遊安打・投安打）
    is_gb_hit = s.str.contains(r"^[一二三遊投]安打$", regex=True)
    # エラー（打球は飛んでいるのでGB扱い）
    is_gb_err = s.str.contains(r"[一二三遊投捕]エラー", regex=True)
    is_gb = is_gb_out | is_gb_hit | is_gb_err

    # ── LD：ライナーアウト ──
    is_ld = (
        s.str.contains(r"[左中右一二三遊投捕][邪]?ライナー|[左中右一二三遊投捕][邪]?直", regex=True) &
        ~s.str.contains(r"安打", regex=True)
    )

    # ── FB：フライアウト + ファウルフライ + 本塁打 + 犠飛 ──
    is_fb_out = (
        s.str.contains(r"[左中右一二三遊投捕][邪]?フライ", regex=True) &
        ~s.str.contains(r"ファウルフライ|本塁打|安打|犠飛", regex=True)
    )
    is_fb_fo  = s.str.contains(r"[左中右一二三遊投捕]ファウルフライ", regex=True)
    is_hr     = s.str.contains(r"本塁打", regex=True)
    is_sf     = s.str.contains(r"犠飛", regex=True)
    is_fb     = is_fb_out | is_fb_fo | is_hr | is_sf

    # ── IFFB：内野フライアウト + ファウルフライ（FBの部分集合）──
    is_iffb = (
        (
            s.str.contains(r"[一二三遊投捕][邪]?フライ", regex=True) &
            ~s.str.contains(r"本塁打|安打|犠飛", regex=True)
        ) | is_fb_fo
    )

    # ── 判断不可打球（ヒット）：外野安打・長打（軌道不明）──
    is_unknown = (
        s.str.contains(r"^[左中右]安打|^[左中右][0-9２３]塁打|^[左中右].*安打", regex=True) |
        s.str.contains(r"[0-9]塁打|[２３]塁打|左3塁打|左中.*塁打|右中.*塁打", regex=True)
    ) & ~is_hr  # 本塁打は除く

    # 実数集計
    gb_n      = int(is_gb.sum())
    ld_n      = int(is_ld.sum())
    fb_n      = int(is_fb.sum())
    iffb_n    = int(is_iffb.sum())
    hr_n      = int(is_hr.sum())
    unknown_n = int(is_unknown.sum())

    bip_total = gb_n + ld_n + fb_n + unknown_n  # 全打球数
    bip_known = gb_n + ld_n + fb_n              # 判断不可除く（%の分母）

    def _pct(n, d):
        return round(n / d * 100, 1) if d > 0 else np.nan

    return {
        "全打球数":     bip_total,
        "GB":           gb_n,
        "LD":           ld_n,
        "FB":           fb_n,
        "IFFB":         iffb_n,
        "HR":           hr_n,
        "判断不可打球": unknown_n,
        "GB%":          _pct(gb_n,   bip_known),
        "LD%":          _pct(ld_n,   bip_known),
        "FB%":          _pct(fb_n,   bip_known),
        "HR%":          _pct(hr_n,   bip_known),
        "IFFB%":        _pct(iffb_n, fb_n),
    }

# --- 打者成績ヘルパー（モジュールレベル） ---
def bat_batted_stats(res) -> dict:
    """打者視点の打球アウト分類"""
    import pandas as pd
    s = pd.Series(res).fillna("").astype(str)
    fly_out = (
        s.str.contains(r"[左中右一二三遊投捕][邪]?フライ", regex=True) &
        ~s.str.contains(r"本塁打|安打", regex=True)
    ) | s.str.contains(r"[左中右一二三遊投捕]ファウルフライ", regex=True)
    line_out = (
        s.str.contains(r"[左中右一二三遊投捕][邪]?ライナー|[左中右一二三遊投捕][邪]?直", regex=True) &
        ~s.str.contains(r"安打", regex=True)
    )
    goro_out = (
        s.str.contains(r"[一二三遊投捕]ゴロ|[一二三遊投捕]併殺打", regex=True) &
        ~s.str.contains(r"安打|野選|犠打", regex=True)
    )
    fly_n = int(fly_out.sum()); line_n = int(line_out.sum()); goro_n = int(goro_out.sum())
    total = fly_n + line_n + goro_n
    return {
        "フライアウト(犠牲フライ含む)": fly_n,
        "ライナーアウト":                line_n,
        "ゴロアウト":                    goro_n,
        "フライアウト割合":   round(fly_n  / total * 100, 1) if total > 0 else "",
        "ライナーアウト割合": round(line_n / total * 100, 1) if total > 0 else "",
        "ゴロアウト割合":     round(goro_n / total * 100, 1) if total > 0 else "",
    }


def bat_counts(res) -> dict:
    """打席完了結果から全打数統計を返す"""
    import pandas as pd
    s = pd.Series(res).fillna("").astype(str)
    # 本塁打：正式名（〇本塁打）＋略称（右中本・左中本・中本など）
    hr_n  = int(s.str.contains(r"本塁打|右中本$|左中本$|^中本$", na=False, regex=True).sum())
    dbl_n = int(s.str.contains(r"2塁打",  na=False).sum())
    tpl_n = int(s.str.contains(r"3塁打",  na=False).sum())
    # 単打：「安打」含む行（外野安打＋内野安打）
    # ※2塁打・3塁打・本塁打は「安打」を含まないため別途加算
    sgl_n = int(s.str.contains(r"安打",   na=False).sum())
    h_n   = hr_n + dbl_n + tpl_n + sgl_n
    extra = dbl_n + tpl_n
    k_n   = int(s.str.contains("三振",             na=False).sum())
    bb_n  = int(s.str.contains(r"^四球$|敬遠",     na=False, regex=True).sum())
    hbp_n = int(s.str.contains(r"^死球$",           na=False, regex=True).sum())
    sf_n  = int(s.str.contains(r"犠飛",             na=False).sum())
    sac_n = int(s.str.contains(r"[投捕]犠打",       na=False, regex=True).sum())
    err_n = int(s.str.contains(r"エラー",           na=False).sum())
    fc_n  = int(s.str.contains(r"野選",             na=False).sum())
    pa    = len(s)
    ab    = pa - bb_n - hbp_n - sac_n - sf_n
    obp_d = ab + bb_n + hbp_n + sf_n
    obp_v = round((h_n + bb_n + hbp_n) / obp_d, 3) if obp_d > 0 else ""
    tb    = sgl_n + dbl_n * 2 + tpl_n * 3 + hr_n * 4
    slg_v = round(tb / ab, 3) if ab > 0 else ""
    ops_v = round(float(obp_v or 0) + float(slg_v or 0), 3) if obp_v != "" and slg_v != "" else ""
    ba_v  = round(h_n / ab, 3) if ab > 0 else ""
    return {
        "pa": pa, "ab": ab, "h": h_n, "hr": hr_n, "extra": extra, "sgl": sgl_n,
        "k": k_n, "bb": bb_n, "hbp": hbp_n, "sf": sf_n,
        "sac": sac_n, "err": err_n, "fc": fc_n, "other": sac_n + err_n + fc_n,
        "obp": obp_v, "slg": slg_v, "ops": ops_v, "ba": ba_v,
    }


def swing_counts(sw_g) -> dict:
    """投球DataFrameからスイング実数を返す"""
    if sw_g is None or (hasattr(sw_g, "empty") and sw_g.empty) or "is_swing" not in sw_g.columns:
        return {"z_sw": 0, "in_z": 0, "o_sw": 0, "out_z": 0,
                "swing": 0, "contact": 0, "swstr": 0}
    swing  = int(sw_g["is_swing"].sum())
    swstr  = int(sw_g["is_swstr"].sum())  if "is_swstr"  in sw_g.columns else 0
    in_z   = int(sw_g["in_zone"].sum())   if "in_zone"   in sw_g.columns else 0
    out_z  = int(sw_g["out_zone"].sum())  if "out_zone"  in sw_g.columns else 0
    z_sw   = int((sw_g["is_swing"] & sw_g["in_zone"]).sum())  if "in_zone"  in sw_g.columns else 0
    o_sw   = int((sw_g["is_swing"] & sw_g["out_zone"]).sum()) if "out_zone" in sw_g.columns else 0
    return {"z_sw": z_sw, "in_z": in_z, "o_sw": o_sw, "out_z": out_z,
            "swing": swing, "contact": swing - swstr, "swstr": swstr}


def swing_pcts(sc: dict) -> dict:
    """スイング実数→%と実数カラムを返す"""
    return {
        "Z-Swing%":  round(sc["z_sw"]    / sc["in_z"]  * 100, 1) if sc["in_z"]  > 0 else "",
        "O-Swing%":  round(sc["o_sw"]    / sc["out_z"] * 100, 1) if sc["out_z"] > 0 else "",
        "Contact%":  round(sc["contact"] / sc["swing"] * 100, 1) if sc["swing"] > 0 else "",
        "whiff%":    round(sc["swstr"]   / sc["swing"] * 100, 1) if sc["swing"] > 0 else "",
        "ゾーン内SW数":  sc["z_sw"],   "ゾーン内投球数": sc["in_z"],
        "ゾーン外SW数":  sc["o_sw"],   "ゾーン外投球数": sc["out_z"],
        "SW数":          sc["swing"],  "コンタクト数":   sc["contact"],
        "空振り数":      sc["swstr"],
    }


# --- 投球前処理 ---
def preprocess_pitch(df: pd.DataFrame) -> pd.DataFrame:
    """投球データに各種フラグ列を追加する。"""
    df = df.copy()

    # ── 重複行除去（スクレイピング時の重複取得を排除） ──
    # 通算投球数（試合内のユニーク番号）があれば投手名+通算投球数で厳密に除去
    # なければコース座標まで含めたキーで除去
    if "通算投球数" in df.columns and "投手名" in df.columns:
        dedup_keys = ["試合ID", "投手名", "通算投球数"]  # 試合IDを含める（複数試合混在時の誤除去防止）
    else:
        dedup_keys = [c for c in [
            "イニング", "表/裏", "アウト数", "打席内球数",
            "球種", "球速", "コース(Top)", "コース(Left)", "1球結果"
        ] if c in df.columns]
    if dedup_keys:
        before = len(df)
        df = df.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)
        removed = before - len(df)
        if removed > 0:
            print(f"    [dedup] 重複行 {removed} 件除去（{before} → {len(df)}球）")

    # 打席番号（打席内球数=1 が打席の開始）
    df["打席番号"] = (df["打席内球数"] == 1).cumsum()

    # 球速を数値化
    df["球速_num"] = df["球速"].apply(parse_vel)

    result = df["1球結果"].str.strip()
    cat    = df["判定カテゴリ"]

    # 空振り: 「空振り」または「空三振」で始まる1球結果
    df["is_swstr"] = result.str.startswith("空振り") | result.str.startswith("空三振")

    # スイング: 空振り + ファウル + インプレー系（アウト/出塁/犠打）
    df["is_swing"] = (
        df["is_swstr"] |
        result.str.startswith("ファウル") |
        cat.isin(["アウト系", "出塁/ヒット系", "犠打/犠飛系"])
    )

    # ゾーン判定（実データ検証済み境界）
    df["in_zone"]  = (
        df["コース(Top)"].between(ZONE_TOP_MIN, ZONE_TOP_MAX) &
        df["コース(Left)"].between(ZONE_LEFT_MIN, ZONE_LEFT_MAX)
    )
    df["out_zone"] = ~df["in_zone"]

    # ── カウント(B-S)を投球順に再構築 ──
    df_sorted = df.sort_values(["試合ID", "通算投球数"])
    balls_col, strikes_col = [], []
    balls, strikes = 0, 0
    for _, row in df_sorted.iterrows():
        if row["打席内球数"] == 1:
            balls, strikes = 0, 0
        balls_col.append(balls)
        strikes_col.append(strikes)
        c = str(row["判定カテゴリ"])
        r = str(row["1球結果"])
        if c == "ボール系":
            balls = min(balls + 1, 3)
        elif c == "ストライク/ファウル系":
            if "ファウル" in r and strikes >= 2:
                pass  # 2S時ファウルはカウント変わらず
            else:
                strikes = min(strikes + 1, 2)
        elif c in ("アウト系", "出塁/ヒット系", "犠打/犠飛系"):
            balls, strikes = 0, 0  # 打席完了でリセット
    df_sorted["pitch_balls"]   = balls_col
    df_sorted["pitch_strikes"] = strikes_col
    df = df_sorted  # ソート済みに差し替え

    # ストライク率フラグ（打者がバットを出した結果=アウト/ヒット/犠打は除外）
    df["is_sf"]   = cat == "ストライク/ファウル系"
    df["is_ball"] = cat == "ボール系"

    # Heart/Shadow/Chase/Waste フラグ
    top  = df["コース(Top)"]
    left = df["コース(Left)"]
    in_heart  = (top.between(HEART_T_MIN, HEART_T_MAX) &
                 left.between(HEART_L_MIN, HEART_L_MAX))
    in_shadow = (top.between(ZONE_TOP_MIN  - SHADOW_EXT, ZONE_TOP_MAX  + SHADOW_EXT) &
                 left.between(ZONE_LEFT_MIN - SHADOW_EXT, ZONE_LEFT_MAX + SHADOW_EXT))
    in_chase  = (top.between(ZONE_TOP_MIN  - SHADOW_EXT - CHASE_EXT,
                              ZONE_TOP_MAX  + SHADOW_EXT + CHASE_EXT) &
                 left.between(ZONE_LEFT_MIN - SHADOW_EXT - CHASE_EXT,
                              ZONE_LEFT_MAX + SHADOW_EXT + CHASE_EXT))
    df["zone_heart"]  = in_heart
    df["zone_shadow"] = in_shadow & ~in_heart
    df["zone_chase"]  = in_chase  & ~in_shadow
    df["zone_waste"]  = ~in_chase

    return df

# --- 球種キー変換 ---
PITCH_KEY_MAP = {
    "ストレート": "FF", "フォーシーム": "FF", "ツーシーム": "SI", "シンカー": "SI",
    "カット": "CT", "カットボール": "CT",
    "スライダー": "SL", "スイーパー": "SL", "縦スライダー": "SL",
    "カーブ": "CU", "フォーク": "FK", "スプリット": "FS",
    "チェンジアップ": "CH", "シュート": "SH",
}
def to_pitch_key(name: str) -> str:
    for k, v in PITCH_KEY_MAP.items():
        if k in str(name): return v
    return str(name)[:2].upper()

# --- OPS計算用イニング解析 ---
INN_COLS    = [f"{i}回" for i in range(1, 10)]
SINGLE_PAT  = r"[左中右三遊投二]安$|遊野選"
DOUBLE_PAT  = r"[左中右二]２$|中２$"
TRIPLE_PAT  = r"[左中右三]３$|中３$|右３$"
HR_PAT      = r"[左中右]本$|左中本$|中本$"
BB_PAT      = r"^四球$|^死球$"
SF_PAT      = r"[右左中]犠飛"
BUNT_PAT    = r"[投捕]犠打"
K_PAT       = r"^空三振$|^見三振$"

def parse_batting_line(row) -> dict:
    ab = bb = h = double = triple = hr = sf = 0
    for c in INN_COLS:
        val = str(row.get(c, "")).strip()
        if not val or val in ("nan", ""): continue
        if re.search(BB_PAT,    val): bb += 1;                              continue
        if re.search(SF_PAT,    val): sf += 1;                              continue
        if re.search(BUNT_PAT,  val):                                        continue
        if re.search(K_PAT,     val): ab += 1;                              continue
        if re.search(HR_PAT,    val): ab += 1; h += 1; hr     += 1;        continue
        if re.search(TRIPLE_PAT,val): ab += 1; h += 1; triple += 1;        continue
        if re.search(DOUBLE_PAT,val): ab += 1; h += 1; double += 1;        continue
        if re.search(SINGLE_PAT,val): ab += 1; h += 1;                     continue
        ab += 1
    single = h - double - triple - hr
    tb     = single + double * 2 + triple * 3 + hr * 4
    obp_d  = ab + bb + sf
    obp    = (h + bb) / obp_d if obp_d > 0 else 0.0
    slg    = tb / ab           if ab   > 0 else 0.0
    return {"ab": ab, "bb": bb, "h": h, "hr": hr,
            "obp": round(obp, 3), "slg": round(slg, 3), "ops": round(obp + slg, 3)}

# --- K%, BB%, K-BB% 計算（投手マスタ用） ---
def calc_pitcher_pcts(row) -> dict:
    ip  = innings_to_float(row.get("投球回"))
    k   = to_num(row.get("奪三振"))
    bb  = to_num(row.get("与四球"))
    hbp = to_num(row.get("与死球"))
    h   = to_num(row.get("被安打"))
    if any(pd.isna(v) for v in [ip, k, bb, hbp, h]) or ip == 0:
        return {"k_pct": 0.0, "bb_pct": 0.0, "k_bb_pct": 0.0}
    bf = ip * 3 + h + bb + hbp
    if bf == 0: return {"k_pct": 0.0, "bb_pct": 0.0, "k_bb_pct": 0.0}
    kp  = round(k  / bf * 100, 1)
    bbp = round(bb / bf * 100, 1)
    return {"k_pct": kp, "bb_pct": bbp, "k_bb_pct": round(kp - bbp, 1)}


def run_datamart(
    path_all_games: str,
    path_pitch: str,
    highlights_path: str | None = None,
) -> str:
    """中間ファイルからデータマートを生成する。出力パスを返す。"""
    print("\n" + "=" * 50)
    print("Step 4: データマート&JSON作成中...")
    print("=" * 50)

    # --- 読み込み ---
    all_games     = pd.read_excel(path_all_games, sheet_name=None)
    df_game_info  = all_games.get("試合基本情報", pd.DataFrame())
    df_scoreboard = all_games.get("スコアボード",  pd.DataFrame())
    df_batters    = all_games.get("打撃成績",      pd.DataFrame())
    df_pitchers   = all_games.get("投手成績",      pd.DataFrame())
    df_stamen     = all_games.get("スタメン",      pd.DataFrame())

    # 試合IDを全シートで文字列に統一（Excel読み込みでint64になるため）
    for df in [df_game_info, df_scoreboard, df_batters, df_pitchers, df_stamen]:
        if not df.empty and "試合ID" in df.columns:
            df["試合ID"] = df["試合ID"].astype(str)

    df_pitch = preprocess_pitch(pd.read_excel(path_pitch))
    if "試合ID" in df_pitch.columns:
        df_pitch["試合ID"] = df_pitch["試合ID"].astype(str)

    home_team_map     = dict(zip(df_game_info["試合ID"], df_game_info["ホームチーム"]))
    last_pitch_per_ab = df_pitch.groupby("打席番号").last().reset_index()

    if not df_stamen.empty and "打順/投手" in df_stamen.columns:
        df_stamen_norm = df_stamen.copy()
        df_stamen_norm["試合ID"] = df_stamen_norm["試合ID"].astype(str)
        starter_map = (
            df_stamen_norm[df_stamen_norm["打順/投手"] == "先発"]
            .set_index(["試合ID", "チーム"])["選手名"].to_dict()
        )
        batting_order_map = (
            df_stamen_norm[df_stamen_norm["打順/投手"].str.match(r"^\d+$", na=False)]
            .assign(order=lambda d: d["打順/投手"].astype(int),
                    選手名=lambda d: d["選手名"].str.strip())
            .set_index(["試合ID", "チーム", "選手名"])["order"].to_dict()
        )
    else:
        starter_map       = {}
        batting_order_map = {}

    # --- 試合ディメンション ---
    print("  試合ディメンションを生成中...")
    rows_dim = []
    for _, g in df_game_info.iterrows():
        gid  = g["試合ID"]
        home = g.get("ホームチーム", "")
        away = g.get("アウェイチーム", "")
        def sb(team, col):
            r = df_scoreboard[(df_scoreboard["試合ID"] == gid) & (df_scoreboard["チーム"] == team)]
            return r.iloc[0][col] if not r.empty else ""
        rows_dim.append({
            "試合ID": gid, "試合日": TARGET_DATE,
            "ホームチーム": home, "アウェイチーム": away,
            "ホーム得点": g.get("ホーム得点", ""), "アウェイ得点": g.get("アウェイ得点", ""),
            "球場": clean_str(g.get("球場", "")), "開始時間": clean_str(g.get("開始時間", "")),
            "試合時間": clean_str(g.get("試合時間", "")),
            "観客数": g.get("観客数", ""),
            "ホームイニング得点": build_innings_str(df_scoreboard, gid, home),
            "アウェイイニング得点": build_innings_str(df_scoreboard, gid, away),
            "ホーム安打": sb(home, "安"), "アウェイ安打": sb(away, "安"),
            "ホームエラー": sb(home, "失"), "アウェイエラー": sb(away, "失"),
            "MVP選手名": g.get("エキサイティングプレーヤー", ""),
            "MVP成績": g.get("MEP_本日の成績", ""),
            "ハイライト": g.get("戦評", ""),
        })
    dim_game = pd.DataFrame(rows_dim)

    # --- 投手ディメンション（pitchデータから生成） ---
    print("  投手ディメンションを生成中...")
    if not df_pitchers.empty and "選手名" in df_pitchers.columns:
        dim_pitcher = df_pitchers[["選手名", "チーム"]].drop_duplicates().reset_index(drop=True)
    else:
        dim_pitcher = pd.DataFrame(columns=["選手名", "チーム"])
    dim_pitcher.insert(0, "投手ID", make_seq_id("P", len(dim_pitcher)))
    pitcher_id_map = {(r["選手名"], r["チーム"]): r["投手ID"] for _, r in dim_pitcher.iterrows()}

    # --- 打者ディメンション（打撃成績から生成） ---
    print("  打者ディメンションを生成中...")
    if not df_batters.empty and "選手名" in df_batters.columns:
        dim_batter = df_batters[["選手名", "チーム"]].drop_duplicates().reset_index(drop=True)
    else:
        dim_batter = pd.DataFrame(columns=["選手名", "チーム"])
    dim_batter.insert(0, "打者ID", make_seq_id("B", len(dim_batter)))
    batter_id_map = {(r["選手名"], r["チーム"]): r["打者ID"] for _, r in dim_batter.iterrows()}

    # --- 投手×試合 集計（打席単位） ---
    # 定義:
    #   対戦打者数    = 全打球数 + 奪三振 + 与四球 + 与死球
    #   アウト数      = GB(アウト) + LD + FB(本塁打・犠飛除く) + 奪三振
    #   全打球数 = GB + LD + FB + 判断不可打球（軌道不明ヒット）
    #   全打球数_除判断不可 = GB + LD + FB
    gp_ab_rows = []
    for (gid, pname), sg in last_pitch_per_ab.groupby(["試合ID", "投手名"]):
        res  = sg["打席完了結果"].fillna("")
        k    = int(res.str.contains("三振",           na=False).sum())
        bb   = int(res.str.contains("^四球$|敬遠",    na=False).sum())
        hbp  = int(res.str.contains("^死球$",          na=False).sum())
        # 犠打：打球分類からは除外しているが対戦打者数・アウト数にはカウント
        # 捕犠打野選は出塁なのでアウトにはならないが打席としてはカウント
        sac      = int(res.str.contains(r"[投捕]犠打$", na=False, regex=True).sum())
        sac_wild = int(res.str.contains("犠打野選",      na=False).sum())

        bstats = calc_batted_stats(res)
        bip_total = bstats["全打球数"]

        # 対戦打者数 = 全打球数 + 奪三振 + 与四球 + 与死球 + 犠打 + 犠打野選
        tbf = bip_total + k + bb + hbp + sac + sac_wild

        # アウト数 = GB(アウト) + LD + FB(本塁打・犠飛除く) + 奪三振 + 犠打（アウト）
        fb_outs = bstats["FB"] - bstats["HR"]
        outs = bstats["GB"] + bstats["LD"] + fb_outs + k + sac  # 犠打野選はアウトでない

        k_pct  = round(k  / tbf * 100, 1) if tbf > 0 else 0.0
        bb_pct = round(bb / tbf * 100, 1) if tbf > 0 else 0.0

        def _pct(v):
            return "" if (v is None or (isinstance(v, float) and np.isnan(v))) else v

        gp_ab_rows.append({
            "試合ID":       gid, "投手名": pname,
            "対戦打者数":   tbf,
            "奪三振":       k,
            "与四球":       bb,
            "与死球":       hbp,
            "全打球数":     bip_total,
            "GB":           bstats["GB"],
            "LD":           bstats["LD"],
            "FB":           bstats["FB"],
            "IFFB":         bstats["IFFB"],
            "HR":           bstats["HR"],
            "判断不可打球": bstats["判断不可打球"],
            "K%":           k_pct,
            "BB%":          bb_pct,
            "K-BB%":        round(k_pct - bb_pct, 1),
            "GB%":          _pct(bstats["GB%"]),
            "LD%":          _pct(bstats["LD%"]),
            "FB%":          _pct(bstats["FB%"]),
            "HR%":          _pct(bstats["HR%"]),
            "IFFB%":        _pct(bstats["IFFB%"]),
        })
    gp_ab = pd.DataFrame(gp_ab_rows)

    # 投手の利き手マップ
    pit_hand_map = df_pitch.groupby("投手名")["投左右"].first().to_dict()

    def _calc_pit_stats(name, g):
        """投手×試合グループの投球指標を計算する（9分割含む）。"""
        n        = len(g)
        pit_left = (pit_hand_map.get(name, "") == "左投")
        g_sf     = g["is_sf"]
        g_ball   = g["is_ball"]
        # ストライク率 = SF系 / (SF系 + ボール系)
        # アウト系・ヒット系（インプレー打球）は分母から除外
        g_sf_cnt   = g["is_sf"].sum()
        g_ball_cnt = g["is_ball"].sum()
        sf_ball    = int(g_sf_cnt + g_ball_cnt)  # SF系+ボール系を分母に

        # 全打者合算の In%/中央%/Out%/高め%/真中%/低め%
        # 打者左右ごとに In/Out 方向が異なるため、左右別に計算して合算する
        all_left  = g["コース(Left)"]
        all_top   = g["コース(Top)"]
        col_in_all  = pd.array([False] * len(g), dtype=bool)
        col_out_all = pd.array([False] * len(g), dtype=bool)
        idx_map     = {idx: i for i, idx in enumerate(g.index)}
        for bh in ["右打", "左打"]:
            mask    = g["打左右"] == bh
            bat_l   = (bh == "左打")
            sm_in   = pit_left ^ bat_l
            sl      = all_left[mask]
            pos     = [idx_map[i] for i in mask[mask].index]
            vals_in  = (sl < ZONE_LE1).values if sm_in else (sl >= ZONE_LE2).values
            vals_out = (sl >= ZONE_LE2).values if sm_in else (sl < ZONE_LE1).values
            for p, vi, vo in zip(pos, vals_in, vals_out):
                col_in_all[p]  = bool(vi)
                col_out_all[p] = bool(vo)
        col_in_all  = pd.Series(col_in_all,  index=g.index, dtype=bool)
        col_out_all = pd.Series(col_out_all, index=g.index, dtype=bool)
        col_ctr_all = ~col_in_all & ~col_out_all

        result = {
            "空振り率":           round(g["is_swstr"].sum() / n * 100, 1) if n > 0 else 0.0,
            "ゾーン内スイング率":  round((g["is_swing"] & g["in_zone"]).sum() / g["in_zone"].sum() * 100, 1)
                                  if g["in_zone"].sum() > 0 else 0.0,
            "ゾーン外スイング率":  round((g["is_swing"] & g["out_zone"]).sum() / g["out_zone"].sum() * 100, 1)
                                  if g["out_zone"].sum() > 0 else 0.0,
            "ゾーン率":           round(g["in_zone"].sum()  / n * 100, 1) if n > 0 else 0.0,
            "ストライク率":       round(g_sf_cnt / sf_ball * 100, 1) if sf_ball > 0 else 0.0,
            # SW系実数
            "ゾーン内SW数":      int((g["is_swing"] & g["in_zone"]).sum()),
            "ゾーン内投球数":    int(g["in_zone"].sum()),
            "ゾーン外SW数":      int((g["is_swing"] & g["out_zone"]).sum()),
            "ゾーン外投球数":    int(g["out_zone"].sum()),
            "SW数":              int(g["is_swing"].sum()),
            "コンタクト数":      int(g["is_swing"].sum()) - int(g["is_swstr"].sum()),
            "空振り数":          int(g["is_swstr"].sum()),
            # 全打者合算 9分割
            "In%":   round(col_in_all.sum()  / n * 100, 1) if n > 0 else 0.0,
            "中央%":  round(col_ctr_all.sum() / n * 100, 1) if n > 0 else 0.0,
            "Out%":  round(col_out_all.sum() / n * 100, 1) if n > 0 else 0.0,
            "高め%":  round((all_top < ZONE_TE1).sum()                       / n * 100, 1) if n > 0 else 0.0,
            "真中%":  round(all_top.between(ZONE_TE1, ZONE_TE2).sum()         / n * 100, 1) if n > 0 else 0.0,
            "低め%":  round((all_top >= ZONE_TE2).sum()                       / n * 100, 1) if n > 0 else 0.0,
        }

        # 9分割（右打/左打別）
        for bat_hand, label in [("右打", "vs右"), ("左打", "vs左")]:
            sg = g[g["打左右"] == bat_hand]
            sn = len(sg)
            if sn == 0:
                for key in ["In%", "中央%", "Out%", "高め%", "真中%", "低め%"]:
                    result[f"{label}_{key}"] = 0.0
                continue
            bat_left = (bat_hand == "左打")
            small_in = pit_left ^ bat_left  # True → Left小=In
            sleft = sg["コース(Left)"]
            stop  = sg["コース(Top)"]
            col_in  = (sleft < ZONE_LE1) if small_in else (sleft >= ZONE_LE2)
            col_out = (sleft >= ZONE_LE2) if small_in else (sleft < ZONE_LE1)
            col_ctr = ~col_in & ~col_out
            result[f"{label}_In%"]   = round(col_in.sum()  / sn * 100, 1)
            result[f"{label}_中央%"]  = round(col_ctr.sum() / sn * 100, 1)
            result[f"{label}_Out%"]  = round(col_out.sum() / sn * 100, 1)
            result[f"{label}_高め%"]  = round((stop < ZONE_TE1).sum()               / sn * 100, 1)
            result[f"{label}_真中%"]  = round(stop.between(ZONE_TE1, ZONE_TE2).sum() / sn * 100, 1)
            result[f"{label}_低め%"]  = round((stop >= ZONE_TE2).sum()               / sn * 100, 1)
        return result

    # groupby キー列の除外問題を回避するためループで集計
    gp_pit_rows = []
    for (gid, name), g in df_pitch.groupby(["試合ID", "投手名"]):
        row = _calc_pit_stats(name, g)
        row["試合ID"]  = gid
        row["投手名"]  = name
        gp_pit_rows.append(row)
    gp_pit = pd.DataFrame(gp_pit_rows)

    # 役割列がない場合は追加して先頭行を先発に設定
    if not df_pitchers.empty and "役割" not in df_pitchers.columns:
        df_pitchers = df_pitchers.copy()
        df_pitchers["役割"] = "中継ぎ"
    # ホーム/アウェイ列がない場合はフォールバック
    if not df_pitchers.empty and "ホーム/アウェイ" not in df_pitchers.columns:
        df_pitchers = df_pitchers.copy()
        df_pitchers["ホーム/アウェイ"] = "home"
    # 役割が全て「中継ぎ」の場合（古いデータ等）、試合内最初の投手を「先発」に補正
    if not df_pitchers.empty and "役割" in df_pitchers.columns and set(df_pitchers["役割"].dropna().unique()) <= {"中継ぎ", ""}:
        grp_cols = [c for c in ["試合ID", "チーム"] if c in df_pitchers.columns]
        if grp_cols:
            first_idx = df_pitchers.groupby(grp_cols).head(1).index
            df_pitchers = df_pitchers.copy()
            df_pitchers.loc[first_idx, "役割"] = "先発"
            print("  [INFO] 役割を自動補正: 各チーム最初の投手を「先発」に設定")

    # --- 試合別投手成績 ---
    print("  試合別投手成績を生成中...")
    fp_rows = []
    for _, p in df_pitchers.iterrows():
        gid  = p["試合ID"]
        name = str(p.get("選手名", "")).strip()
        team = str(p.get("チーム", "")).strip()
        ab_r  = gp_ab[(gp_ab["試合ID"] == gid) & (gp_ab["投手名"] == name)]
        pit_r = gp_pit[(gp_pit["試合ID"] == gid) & (gp_pit["投手名"] == name)]
        def _ab(c, default=""):
            if ab_r.empty or c not in ab_r.columns: return default
            v = ab_r.iloc[0][c]
            return default if (isinstance(v, float) and np.isnan(v)) else v
        def _pit(c): return pit_r.iloc[0][c] if not pit_r.empty else 0.0
        ip_v = innings_to_float(p.get("投球回", ""))

        # 役割判定
        st_name = starter_map.get((gid, team))
        role = ("先発" if st_name == name else "中継ぎ") if st_name else p.get("役割", "中継ぎ")

        fp_rows.append({
            "選手名":            name,
            "チーム":            team,
            "試合ID":            gid,
            "試合日":            TARGET_DATE,
            "投球回":            round(ip_v, 2) if ip_v is not None else "",
            "投球数":            p.get("投球数", ""),
            "対戦打者数":        _ab("対戦打者数", 0),
            "失点":              p.get("失点", ""),
            "自責点":            p.get("自責点", ""),
            "被安打":            p.get("被安打", ""),
            "被本塁打":          p.get("被本塁打", _ab("HR", "")),  # all_gamesにあればそちら優先
            "与四球":            p.get("与四球", ""),
            "与死球":            p.get("与死球", ""),
            "奪三振":            p.get("奪三振", ""),
            "K%":                _ab("K%",    0.0),
            "BB%":               _ab("BB%",   0.0),
            "K-BB%":             _ab("K-BB%", 0.0),
            "全打球数":          _ab("全打球数", ""),
            "GB":                _ab("GB",       ""),
            "LD":                _ab("LD",       ""),
            "FB":                _ab("FB",       ""),
            "IFFB":              _ab("IFFB",     ""),
            "判断不可打球":      _ab("判断不可打球", ""),
            "HR":                _ab("HR",       ""),
            "H":                 p.get("被安打", ""),  # 被安打と同値
            "GB%":               _ab("GB%",   ""),
            "LD%":               _ab("LD%",   ""),
            "FB%":               _ab("FB%",   ""),
            "IFFB%":             _ab("IFFB%", ""),
            "HR%":               _ab("HR%",   ""),
            "空振り率":           _pit("空振り率"),
            "ゾーン内スイング率":  _pit("ゾーン内スイング率"),
            "ゾーン外スイング率":  _pit("ゾーン外スイング率"),
            "ゾーン率":           _pit("ゾーン率"),
            "ゾーン外スイング率": _pit("ゾーン外スイング率"),
            "ストライク率":      _pit("ストライク率"),
            "ゾーン内SW数":      _pit("ゾーン内SW数"),
            "ゾーン内投球数":    _pit("ゾーン内投球数"),
            "ゾーン外SW数":      _pit("ゾーン外SW数"),
            "ゾーン外投球数":    _pit("ゾーン外投球数"),
            "SW数":              _pit("SW数"),
            "コンタクト数":      _pit("コンタクト数"),
            "空振り数":          _pit("空振り数"),
            "In%":               _pit("In%"),
            "中央%":             _pit("中央%"),
            "Out%":              _pit("Out%"),
            "高め%":             _pit("高め%"),
            "真中%":             _pit("真中%"),
            "低め%":             _pit("低め%"),
            # 内部用（playerシートでは除外）
            "投手ID":            pitcher_id_map.get((name, team), ""),
            "役割":              role,
            "勝敗成績":          clean_str(p.get("勝敗成績", "")),
            "ホーム/アウェイ":   get_side(gid, team, home_team_map),
        })
    fact_game_pitcher = pd.DataFrame(fp_rows)
    fact_game_pitcher.insert(0, "投手試合ID", make_seq_id("FP", len(fact_game_pitcher)))

    # --- 試合別投球配球 ---
    print("  試合別投球配球を生成中...")
    def _pitch_metrics(g):
        """球種グループの指標を計算する共通関数。"""
        n = len(g)
        g_sf_cnt_pm   = g["is_sf"].sum()
        g_ball_cnt_pm = g["is_ball"].sum()
        sf_ball       = int(g_sf_cnt_pm + g_ball_cnt_pm)  # SF系+ボール系を分母に
        # g_strike_total_lr は削除（不使用）
        return pd.Series({
            "投球数":            n,
            "平均球速":          round(g["球速_num"].mean(), 1) if g["球速_num"].notna().any() else None,
            "最高球速":          g["球速_num"].max() if g["球速_num"].notna().any() else None,
            "空振り率":          round(g["is_swstr"].sum() / n * 100, 1) if n > 0 else 0.0,
            "ゾーン率":          round(g["in_zone"].sum()  / n * 100, 1) if n > 0 else 0.0,
            "ゾーン外スイング率": round((g["is_swing"] & g["out_zone"]).sum() / g["out_zone"].sum() * 100, 1)
                                if g["out_zone"].sum() > 0 else 0.0,
            "ストライク率":      round(g_sf_cnt_pm / sf_ball * 100, 1) if sf_ball > 0 else 0.0,
        })

    pm_agg = (
        df_pitch.groupby(["試合ID", "投手名", "球種"])
        .apply(_pitch_metrics, include_groups=False)
        .reset_index()
    )
    pm_agg["試合ID"] = pm_agg["試合ID"].astype(str)  # gp_ref との型統一
    pm_agg["投球割合%"] = (pm_agg["投球数"] / pm_agg.groupby(["試合ID","投手名"])["投球数"].transform("sum") * 100).round(1)

    # 打席完了結果から球種別の打球分類・被安打・被本塁打を集計
    batted_by_pitch = {}
    for (gid, pname, pitch), g in last_pitch_per_ab.groupby(["試合ID", "投手名", "球種"]):
        res = g["打席完了結果"].fillna("")
        bstats = calc_batted_stats(res)
        batted_by_pitch[(gid, pname, pitch)] = {
            "被安打数":   int(res.str.contains("安打|2塁打|3塁打|本塁打", na=False).sum()),
            "被本塁打数": int(res.str.contains("本塁打", na=False).sum()),
            "全打球数":   bstats["全打球数"],
            "GB":         bstats["GB"],   "LD":         bstats["LD"],
            "FB":         bstats["FB"],   "IFFB":       bstats["IFFB"],
            "HR":         bstats["HR"],   "判断不可打球": bstats["判断不可打球"],
            "GB%":        bstats["GB%"],  "LD%":        bstats["LD%"],
            "FB%":        bstats["FB%"],  "HR%":        bstats["HR%"],
            "IFFB%":      bstats["IFFB%"],
        }

    # 9分割（投手×試合×球種）
    pit_hand_map_pm = df_pitch.groupby("投手名")["投左右"].first().to_dict()
    zone9_by_pitch = {}
    for (gid, pname, pitch), g in df_pitch.groupby(["試合ID", "投手名", "球種"]):
        n = len(g)
        if n == 0: continue
        pit_left = (pit_hand_map_pm.get(pname, "") == "左投")
        all_left = g["コース(Left)"]
        all_top  = g["コース(Top)"]
        col_in  = pd.array([False]*n, dtype=bool)
        col_out = pd.array([False]*n, dtype=bool)
        idx_map = {idx: i for i, idx in enumerate(g.index)}
        for bh in ["右打", "左打"]:
            mask   = g["打左右"] == bh
            bat_l  = (bh == "左打")
            sm_in  = pit_left ^ bat_l
            sl     = all_left[mask]
            pos    = [idx_map[i] for i in mask[mask].index]
            vi = (sl < ZONE_LE1).values if sm_in else (sl >= ZONE_LE2).values
            vo = (sl >= ZONE_LE2).values if sm_in else (sl < ZONE_LE1).values
            for p, i_, o_ in zip(pos, vi, vo):
                col_in[p]  = bool(i_)
                col_out[p] = bool(o_)
        col_in  = pd.Series(col_in,  index=g.index, dtype=bool)
        col_out = pd.Series(col_out, index=g.index, dtype=bool)
        col_ctr = ~col_in & ~col_out
        zone9_by_pitch[(gid, pname, pitch)] = {
            "In%":  round(col_in.sum()  / n * 100, 1),
            "中央%": round(col_ctr.sum() / n * 100, 1),
            "Out%": round(col_out.sum() / n * 100, 1),
            "高め%": round((all_top < ZONE_TE1).sum()                       / n * 100, 1),
            "真中%": round(all_top.between(ZONE_TE1, ZONE_TE2).sum()         / n * 100, 1),
            "低め%": round((all_top >= ZONE_TE2).sum()                       / n * 100, 1),
        }

    gp_ref = fact_game_pitcher[["投手試合ID","試合ID","選手名","投手ID","チーム","役割","ホーム/アウェイ"]].copy()
    pm_agg = pm_agg.merge(gp_ref, left_on=["試合ID","投手名"], right_on=["試合ID","選手名"], how="left")
    pm_rows = []
    for _, r in pm_agg.iterrows():
        pname = str(r["球種"]).strip()
        gid   = r["試合ID"]
        iname = r["投手名"]
        bd    = batted_by_pitch.get((gid, iname, r["球種"]), {})
        z9    = zone9_by_pitch.get((gid, iname, r["球種"]), {})

        def _bd(k): return bd.get(k, "")
        def _z9(k): return z9.get(k, 0.0)
        def _pct(v): return "" if (v is None or (isinstance(v, float) and np.isnan(v))) else v

        # SW系実数をpitch groupから計算
        gp = df_pitch[(df_pitch["試合ID"] == gid) & (df_pitch["投手名"] == iname) & (df_pitch["球種"] == r["球種"])]
        gp_n  = len(gp)
        sw_n  = int(gp["is_swing"].sum())  if gp_n > 0 else 0
        swstr_n = int(gp["is_swstr"].sum()) if gp_n > 0 else 0
        in_z_n  = int(gp["in_zone"].sum())  if gp_n > 0 else 0
        out_z_n = int(gp["out_zone"].sum()) if gp_n > 0 else 0
        z_sw_n  = int((gp["is_swing"] & gp["in_zone"]).sum())  if gp_n > 0 else 0
        o_sw_n  = int((gp["is_swing"] & gp["out_zone"]).sum()) if gp_n > 0 else 0

        pm_rows.append({
            "投手試合ID":      r.get("投手試合ID",""), "試合ID": gid, "試合日": TARGET_DATE,
            "投手ID":          r.get("投手ID",""), "選手名": iname,
            "チーム":          r.get("チーム",""), "役割": r.get("役割",""),
            "ホーム/アウェイ": r.get("ホーム/アウェイ",""),
            "球種コード":      to_pitch_key(pname), "球種名": pname,
            "投球数":          int(r["投球数"]), "投球割合%": r["投球割合%"],
            "平均球速":        r["平均球速"] if pd.notna(r["平均球速"]) else "",
            "最高球速":        r["最高球速"] if pd.notna(r["最高球速"]) else "",
            "全打球数":        _bd("全打球数"),
            "GB":              _bd("GB"),   "LD":    _bd("LD"),
            "FB":              _bd("FB"),   "IFFB":  _bd("IFFB"),
            "判断不可打球":    _bd("判断不可打球"),
            "HR":              _bd("HR"),
            "H":               _bd("被安打数"),
            "GB%":             _pct(_bd("GB%")),  "LD%":   _pct(_bd("LD%")),
            "FB%":             _pct(_bd("FB%")),  "HR%":   _pct(_bd("HR%")),
            "IFFB%":           _pct(_bd("IFFB%")),
            "空振り率":            round(swstr_n / gp_n * 100, 1) if gp_n > 0 else 0.0,
            "ゾーン内スイング率":   round(z_sw_n / in_z_n * 100, 1) if in_z_n > 0 else 0.0,
            "ゾーン外スイング率":   round(o_sw_n / out_z_n * 100, 1) if out_z_n > 0 else 0.0,
            "ゾーン率":            round(in_z_n / gp_n * 100, 1) if gp_n > 0 else 0.0,
            "ストライク率":        r["ストライク率"],
            "ゾーン内SW数":        z_sw_n,
            "ゾーン内投球数":      in_z_n,
            "ゾーン外SW数":        o_sw_n,
            "ゾーン外投球数":      out_z_n,
            "SW数":                sw_n,
            "コンタクト数":        sw_n - swstr_n,
            "空振り数":            swstr_n,
            "In%":  _z9("In%"),  "中央%": _z9("中央%"), "Out%": _z9("Out%"),
            "高め%": _z9("高め%"), "真中%": _z9("真中%"), "低め%": _z9("低め%"),
        })
    fact_pitch_mix = pd.DataFrame(pm_rows)
    fact_pitch_mix.insert(0, "配球試合ID", make_seq_id("PM", len(fact_pitch_mix), width=6))

    # --- 試合別打者成績 ---
    print("  試合別打者成績を生成中...")

    # 投球データから打者スイング指標を集計（実数も保持してシーズン集計に使う）
    bat_swing = {}
    for (gid, bname), g in df_pitch.groupby(["試合ID", "打者名"]):
        swing   = int(g["is_swing"].sum())
        swstr   = int(g["is_swstr"].sum())
        in_z    = int(g["in_zone"].sum())
        out_z   = int(g["out_zone"].sum())
        z_sw    = int((g["is_swing"] & g["in_zone"]).sum())
        o_sw    = int((g["is_swing"] & g["out_zone"]).sum())
        contact = swing - swstr
        bat_swing[(str(gid), bname)] = {
            # 出力用%
            "Z-Swing%": round(z_sw   / in_z  * 100, 1) if in_z  > 0 else 0.0,
            "O-Swing%": round(o_sw   / out_z * 100, 1) if out_z > 0 else 0.0,
            "Contact%": round(contact / swing * 100, 1) if swing > 0 else 0.0,
            "whiff%":   round(swstr  / swing * 100, 1) if swing > 0 else 0.0,
            # シーズン集計用の実数（出力シートには含めない）
            "ゾーン内SW数":    z_sw,    "ゾーン内投球数":  in_z,
            "ゾーン外SW数":    o_sw,    "ゾーン外投球数": out_z,
            "SW数":   swing,   "コンタクト数": contact,
            "空振り数":   swstr,
        }

    # 打席完了結果から打者視点の打球アウト分類
    def _bat_batted(res: pd.Series) -> dict:
        """打者視点の打球アウト集計（安打・HR・犠飛はアウトでないので除外）"""
        s = res.fillna("").astype(str)
        # フライアウト = フライアウト + ファウルフライ + 犠飛
        fly_out = (
            s.str.contains(r"[左中右一二三遊投捕][邪]?フライ", regex=True) &
            ~s.str.contains(r"本塁打|安打", regex=True)
        ) | s.str.contains(r"[左中右一二三遊投捕]ファウルフライ", regex=True)
        # ライナーアウト
        line_out = (
            s.str.contains(r"[左中右一二三遊投捕][邪]?ライナー|[左中右一二三遊投捕][邪]?直", regex=True) &
            ~s.str.contains(r"安打", regex=True)
        )
        # ゴロアウト（内野安打・野選出塁・犠打は除く）
        goro_out = (
            s.str.contains(r"[一二三遊投捕]ゴロ|[一二三遊投捕]併殺打", regex=True) &
            ~s.str.contains(r"安打|野選|犠打", regex=True)
        )
        fly_n  = int(fly_out.sum())
        line_n = int(line_out.sum())
        goro_n = int(goro_out.sum())
        total  = fly_n + line_n + goro_n
        return {
            "フライアウト": fly_n, "ライナーアウト": line_n, "ゴロアウト": goro_n,
            "フライアウト割合":   round(fly_n  / total * 100, 1) if total > 0 else "",
            "ライナーアウト割合": round(line_n / total * 100, 1) if total > 0 else "",
            "ゴロアウト割合":     round(goro_n / total * 100, 1) if total > 0 else "",
        }

    # 打者×試合の打席完了結果を取得
    bat_last = last_pitch_per_ab.copy()
    # 試合IDを文字列に統一（型不一致防止）
    if "試合ID" in bat_last.columns:
        bat_last["試合ID"] = bat_last["試合ID"].astype(str)

    gb_rows = []
    for _, b in df_batters.iterrows():
        gid  = str(b["試合ID"])   # 文字列に統一
        name = str(b.get("選手名", "")).strip()
        team = str(b.get("チーム", "")).strip()
        sw   = bat_swing.get((gid, name), bat_swing.get((b["試合ID"], name), {
            "Z-Swing%": 0.0, "O-Swing%": 0.0, "Contact%": 0.0, "whiff%": 0.0
        }))

        # 打席完了結果から全統計を正確に計算
        bat_res = bat_last[
            (bat_last["試合ID"] == gid) & (bat_last["打者名"] == name)
        ]["打席完了結果"].fillna("")
        res_s = bat_res.astype(str)

        # 各結果のカウント
        hr_cnt  = int(res_s.str.contains(r"本塁打|右中本$|左中本$|^中本$", na=False, regex=True).sum())
        dbl_cnt = int(res_s.str.contains(r"2塁打", na=False).sum())
        tpl_cnt = int(res_s.str.contains(r"3塁打", na=False).sum())
        # 単打：「安打」含む行（外野安打＋内野安打）
        # ※2塁打・3塁打・本塁打は「安打」を含まないため別途加算
        sgl_cnt = int(res_s.str.contains(r"安打", na=False).sum())
        h_cnt   = hr_cnt + dbl_cnt + tpl_cnt + sgl_cnt
        extra_hit = dbl_cnt + tpl_cnt   # 長打 = 2塁打 + 3塁打
        k_cnt   = int(res_s.str.contains("三振",    na=False).sum())
        bb_cnt  = int(res_s.str.contains(r"^四球$|敬遠", na=False, regex=True).sum())
        hbp_cnt = int(res_s.str.contains(r"^死球$",      na=False, regex=True).sum())
        sf_cnt  = int(res_s.str.contains(r"犠飛",        na=False).sum())
        sac_cnt = int(res_s.str.contains(r"[投捕]犠打",  na=False, regex=True).sum())
        err_cnt = int(res_s.str.contains(r"エラー",      na=False).sum())
        fc_cnt  = int(res_s.str.contains(r"野選",        na=False).sum())
        other_cnt = sac_cnt + err_cnt + fc_cnt

        # 打席数・打数
        pa = len(bat_res)                                   # 全打席数
        # 打数 = 打席 - 四球 - 死球 - 犠打 - 犠飛
        ab = pa - bb_cnt - hbp_cnt - sac_cnt - sf_cnt

        # 出塁率・長打率・OPS
        obp_d = ab + bb_cnt + hbp_cnt + sf_cnt
        obp   = round((h_cnt + bb_cnt + hbp_cnt) / obp_d, 3) if obp_d > 0 else 0.0
        tb    = sgl_cnt + dbl_cnt * 2 + tpl_cnt * 3 + hr_cnt * 4
        slg   = round(tb / ab, 3) if ab > 0 else 0.0
        ops   = round(obp + slg, 3)
        ba    = round(h_cnt / ab, 3) if ab > 0 else 0.0

        # 割合
        bb_pct = round(bb_cnt / pa * 100, 1) if pa > 0 else 0.0
        k_pct  = round(k_cnt  / pa * 100, 1) if pa > 0 else 0.0
        hr_pct = round(hr_cnt    / h_cnt * 100, 1) if h_cnt > 0 else ""
        xh_pct = round(extra_hit / h_cnt * 100, 1) if h_cnt > 0 else ""
        sg_pct = round(sgl_cnt   / h_cnt * 100, 1) if h_cnt > 0 else ""

        # 打球アウト分類
        bd = bat_batted_stats(bat_res)

        gb_rows.append({
            "試合ID":          gid,
            "試合日":          TARGET_DATE,
            "選手名":          name,
            "チーム":          team,
            "ホーム/アウェイ": get_side(gid, team, home_team_map),
            "打順":            batting_order_map.get((gid, team, name), ""),
            "守備位置":        b.get("位置", ""),
            "打席別結果":      ",".join([
                str(b[c]).strip() for c in INN_COLS
                if c in b.index and pd.notna(b[c]) and str(b[c]).strip() not in ("", "nan")
            ]),
            "OPS": ops, "出塁率": obp, "長打率": slg, "打率": ba,
            "打点":   to_num(b.get("打点", 0), default=0),
            "盗塁":   to_num(b.get("盗塁", 0), default=0),
            "打席": pa, "打数": ab,
            "安打": h_cnt, "本塁打": hr_cnt, "長打": extra_hit, "単打": sgl_cnt,
            "四球":   bb_cnt, "死球": hbp_cnt,
            "三振":   k_cnt,
            "フライアウト(犠牲フライ含む)": bd.get("フライアウト(犠牲フライ含む)", bd.get("フライアウト（犠牲フライ含む）", 0)),
            "ライナーアウト":              bd["ライナーアウト"],
            "ゴロアウト":                  bd["ゴロアウト"],
            "その他(犠打、失策、野選)": other_cnt,
            "本塁打割合": hr_pct, "長打割合": xh_pct, "単打割合": sg_pct,
            "フライアウト割合":   bd["フライアウト割合"],
            "ライナーアウト割合": bd["ライナーアウト割合"],
            "ゴロアウト割合":     bd["ゴロアウト割合"],
            "K%": k_pct, "BB%": bb_pct,
            "Z-Swing%": sw["Z-Swing%"], "O-Swing%": sw["O-Swing%"],
            "Contact%": sw["Contact%"], "whiff%":    sw["whiff%"],
            "球数":          sw.get("ゾーン内投球数", 0) + sw.get("ゾーン外投球数", 0),
            "ゾーン内SW数":  sw.get("ゾーン内SW数",   0), "ゾーン内投球数": sw.get("ゾーン内投球数", 0),
            "ゾーン外SW数":  sw.get("ゾーン外SW数",   0), "ゾーン外投球数": sw.get("ゾーン外投球数", 0),
            "SW数":          sw.get("SW数",           0), "コンタクト数":   sw.get("コンタクト数",   0),
            "空振り数":      sw.get("空振り数",        0),
        })
    fact_game_batter = pd.DataFrame(gb_rows)
    fact_game_batter.insert(0, "打者試合ID", make_seq_id("FB", len(fact_game_batter)))
    fact_game_batter_out = fact_game_batter

    # --- 試合別投手成績(左右別) ---
    print("  試合別投手成績(左右別)を生成中...")
    # last_pitch_per_ab から左右別の打席完了結果を集計するための辞書
    # 被安打・与四球・奪三振・失点・自責点は打席単位（last_pitch_per_ab）で集計
    lr_ab_dict = {}  # (gid, name, hand) → stats dict
    for (gid, pname, hand), sg_ab in last_pitch_per_ab.groupby(["試合ID", "投手名", "打左右"]):
        res  = sg_ab["打席完了結果"].fillna("")
        h    = int(res.str.contains("安打|2塁打|3塁打|本塁打", na=False).sum())
        bb   = int(res.str.contains("^四球$",  na=False).sum())
        hbp  = int(res.str.contains("^死球$",  na=False).sum())
        k    = int(res.str.contains("三振",    na=False).sum())
        sac      = int(res.str.contains(r"[投捕]犠打$", na=False, regex=True).sum())
        sac_wild = int(res.str.contains("犠打野選",      na=False).sum())

        bstats    = calc_batted_stats(res)
        bip_total = bstats["全打球数"]
        fb_outs   = bstats["FB"] - bstats["HR"]
        outs      = bstats["GB"] + bstats["LD"] + fb_outs + k + sac
        tbf       = bip_total + k + bb + hbp + sac + sac_wild
        k_p  = round(k  / tbf * 100, 1) if tbf > 0 else 0.0
        bb_p = round(bb / tbf * 100, 1) if tbf > 0 else 0.0

        def _p(v):
            return "" if (v is None or (isinstance(v, float) and np.isnan(v))) else v

        lr_ab_dict[(gid, pname, hand)] = {
            "被安打": h, "与四球": bb, "与死球": hbp, "奪三振": k,
            "対戦打者数":   tbf,
            "全打球数":     bip_total,
            "GB":    bstats["GB"],   "LD":    bstats["LD"],
            "FB":    bstats["FB"],   "IFFB":  bstats["IFFB"],
            "HR":    bstats["HR"],   "判断不可打球": bstats["判断不可打球"],
            "K%":    k_p, "BB%": bb_p, "K-BB%": round(k_p - bb_p, 1),
            "GB%":   _p(bstats["GB%"]),   "LD%":   _p(bstats["LD%"]),
            "FB%":   _p(bstats["FB%"]),   "HR%":   _p(bstats["HR%"]),
            "IFFB%": _p(bstats["IFFB%"]),
        }

    # fact_game_pitcher から失点・自責点・投手メタ情報を引く辞書
    fp_dict = {}
    for _, row in fact_game_pitcher.iterrows():
        fp_dict[(row["試合ID"], row["選手名"])] = row

    pit_lr_rows = []
    for (gid, name), g in df_pitch.groupby(["試合ID", "投手名"]):
        pit_left = (pit_hand_map.get(name, "") == "左投")
        fp = fp_dict.get((gid, name))

        pit_trial_id = fp["投手試合ID"]    if fp is not None else ""
        pit_id       = fp["投手ID"]        if fp is not None else ""
        team         = fp["チーム"]        if fp is not None else ""
        ha           = fp["ホーム/アウェイ"] if fp is not None else ""
        role         = fp["役割"]          if fp is not None else ""
        # 失点・自責点は試合全体の値を按分せず、左右別に打席データから取れないので
        # 試合合計を左右合計投球数で按分して概算値として付与
        total_pitches = len(g)
        er_total  = fp["失点"]   if fp is not None else ""
        era_total = fp["自責点"] if fp is not None else ""

        for bat_hand in ["右打", "左打"]:
            sg = g[g["打左右"] == bat_hand]
            sn = len(sg)
            if sn == 0:
                continue

            bat_left = (bat_hand == "左打")
            small_in = pit_left ^ bat_left
            sleft = sg["コース(Left)"]
            stop  = sg["コース(Top)"]
            col_in  = (sleft < ZONE_LE1) if small_in else (sleft >= ZONE_LE2)
            col_out = (sleft >= ZONE_LE2) if small_in else (sleft < ZONE_LE1)
            col_ctr = ~col_in & ~col_out
            sf_ball = int(sg["is_sf"].sum() + sg["is_ball"].sum())  # SF系+ボール系を分母に
            # sg_strike_total は削除（不使用）
            tai = "右" if bat_hand == "右打" else "左"

            ab_stats = lr_ab_dict.get((gid, name, bat_hand), {})
            n_ab_h  = ab_stats.get("被安打", "")
            n_ab_bb = ab_stats.get("与四球", "")
            n_ab_hbp= ab_stats.get("与死球", "")
            n_ab_k  = ab_stats.get("奪三振", "")

            # K%/BB% は対戦打者数（lr_ab_dictから）ベースで算出
            n_tbf   = ab_stats.get("対戦打者数", 0)
            k_pct   = round(n_ab_k  / n_tbf * 100, 1) if (isinstance(n_ab_k, int) and n_tbf > 0) else 0.0
            bb_pct  = round(n_ab_bb / n_tbf * 100, 1) if (isinstance(n_ab_bb, int) and n_tbf > 0) else 0.0
            kbb_pct = round(k_pct - bb_pct, 1)

            pit_lr_rows.append({
                "投手試合ID":    pit_trial_id,
                "試合ID":        gid,
                "試合日":        TARGET_DATE,
                "投手ID":        pit_id,
                "選手名":        name,
                "チーム":        team,
                "ホーム/アウェイ": ha,
                "役割":          role,
                "対打者":        tai,
                "投球数":        sn,
                "対戦打者数":    n_tbf,
                "失点":          er_total,
                "自責点":        era_total,
                "被安打":        n_ab_h,
                "被本塁打":      ab_stats.get("HR", ""),
                "与四球":        n_ab_bb,
                "与死球":        n_ab_hbp,
                "奪三振":        n_ab_k,
                "K%":            k_pct,
                "BB%":           bb_pct,
                "K-BB%":         kbb_pct,
                "全打球数":      ab_stats.get("全打球数", ""),
                "GB":            ab_stats.get("GB",    ""),
                "LD":            ab_stats.get("LD",    ""),
                "FB":            ab_stats.get("FB",    ""),
                "IFFB":          ab_stats.get("IFFB",  ""),
                "判断不可打球":  ab_stats.get("判断不可打球", ""),
                "HR":            ab_stats.get("HR",    ""),
                "H":             n_ab_h,
                "GB%":           ab_stats.get("GB%",   ""),
                "LD%":           ab_stats.get("LD%",   ""),
                "FB%":           ab_stats.get("FB%",   ""),
                "IFFB%":         ab_stats.get("IFFB%", ""),
                "HR%":           ab_stats.get("HR%",   ""),
                # ゾーン系
                "空振り率":          round(sg["is_swstr"].sum() / sn * 100, 1) if sn > 0 else 0.0,
                "ゾーン内スイング率": round((sg["is_swing"] & sg["in_zone"]).sum() / sg["in_zone"].sum() * 100, 1)
                                     if sg["in_zone"].sum() > 0 else 0.0,
                "ゾーン外スイング率": round((sg["is_swing"] & sg["out_zone"]).sum() / sg["out_zone"].sum() * 100, 1)
                                     if sg["out_zone"].sum() > 0 else 0.0,
                "ゾーン率":          round(sg["in_zone"].sum() / sn * 100, 1),
                "ストライク率":      round(sg["is_sf"].sum() / sf_ball * 100, 1) if sf_ball > 0 else 0.0,
                "ゾーン内SW数":      int((sg["is_swing"] & sg["in_zone"]).sum()),
                "ゾーン内投球数":    int(sg["in_zone"].sum()),
                "ゾーン外SW数":      int((sg["is_swing"] & sg["out_zone"]).sum()),
                "ゾーン外投球数":    int(sg["out_zone"].sum()),
                "SW数":              int(sg["is_swing"].sum()),
                "コンタクト数":      int(sg["is_swing"].sum()) - int(sg["is_swstr"].sum()),
                "空振り数":          int(sg["is_swstr"].sum()),
                "In%":   round(col_in.sum()  / sn * 100, 1),
                "中央%":  round(col_ctr.sum() / sn * 100, 1),
                "Out%":  round(col_out.sum() / sn * 100, 1),
                "高め%":  round((stop < ZONE_TE1).sum() / sn * 100, 1),
                "真中%":  round(stop.between(ZONE_TE1, ZONE_TE2).sum() / sn * 100, 1),
                "低め%":  round((stop >= ZONE_TE2).sum() / sn * 100, 1),
            })

    _FP_LR_COLS = [
        "投手試合ID", "試合ID", "試合日", "投手ID", "選手名", "チーム",
        "ホーム/アウェイ", "役割", "対打者", "投球数", "対戦打者数",
        "失点", "自責点", "被安打", "与四球", "与死球", "奪三振",
        "K%", "BB%", "K-BB%",
        "全打球数",
        "GB", "LD", "FB", "IFFB", "HR", "判断不可打球",
        "GB%", "LD%", "FB%", "HR%", "IFFB%",
        "ゾーン率", "ゾーン外スイング率", "ストライク率",
        "In%", "中央%", "Out%", "高め%", "真中%", "低め%",
    ]
    if pit_lr_rows:
        fact_game_pitcher_lr = pd.DataFrame(pit_lr_rows)
    else:
        fact_game_pitcher_lr = pd.DataFrame(columns=_FP_LR_COLS)

    # --- 試合別投球配球(左右別) ---
    print("  試合別投球配球(左右別)を生成中...")

    def _pitch_metrics_lr(g):
        """左右別球種グループの指標計算"""
        n = len(g)
        g_sf_cnt_lr   = g["is_sf"].sum()
        g_ball_cnt_lr = g["is_ball"].sum()
        sf_ball       = int(g_sf_cnt_lr + g_ball_cnt_lr)  # SF系+ボール系を分母に
        # g_st は削除（不使用）
        return pd.Series({
            "投球数":            n,
            "平均球速":          round(g["球速_num"].mean(), 1) if g["球速_num"].notna().any() else None,
            "最高球速":          g["球速_num"].max() if g["球速_num"].notna().any() else None,
            "空振り率":          round(g["is_swstr"].sum() / n * 100, 1) if n > 0 else 0.0,
            "ゾーン率":          round(g["in_zone"].sum()  / n * 100, 1) if n > 0 else 0.0,
            "ゾーン外スイング率": round((g["is_swing"] & g["out_zone"]).sum() / g["out_zone"].sum() * 100, 1)
                               if g["out_zone"].sum() > 0 else 0.0,
            "ストライク率":      round(g_sf_cnt_lr / sf_ball * 100, 1) if sf_ball > 0 else 0.0,
        })

    # 左右別：被安打・打球分類の集計
    batted_by_pitch_lr = {}
    for (gid, pname, hand, pitch), g in last_pitch_per_ab.groupby(["試合ID","投手名","打左右","球種"]):
        res = g["打席完了結果"].fillna("")
        bstats = calc_batted_stats(res)
        batted_by_pitch_lr[(gid, pname, hand, pitch)] = {
            "被安打数":   int(res.str.contains("安打|2塁打|3塁打|本塁打", na=False).sum()),
            "被本塁打数": int(res.str.contains("本塁打", na=False).sum()),
            "全打球数":   bstats["全打球数"],
            "GB":         bstats["GB"],   "LD":         bstats["LD"],
            "FB":         bstats["FB"],   "IFFB":       bstats["IFFB"],
            "HR":         bstats["HR"],   "判断不可打球": bstats["判断不可打球"],
            "GB%":        bstats["GB%"],  "LD%":        bstats["LD%"],
            "FB%":        bstats["FB%"],  "HR%":        bstats["HR%"],
            "IFFB%":      bstats["IFFB%"],
        }

    # 左右別小計（投球割合%の分母用）
    total_lr = {}
    for (gid, pname, hand), g in df_pitch.groupby(["試合ID","投手名","打左右"]):
        total_lr[(gid, pname, hand)] = len(g)

    # gp_ref を辞書化
    gp_ref_dict = {}
    for _, row in gp_ref.iterrows():
        gp_ref_dict[(row["試合ID"], row["選手名"])] = row

    pm_lr_rows = []
    for (gid, pname, hand, pitch), g in df_pitch.groupby(["試合ID","投手名","打左右","球種"]):
        g = g.reset_index(drop=True)
        n = len(g)
        sf_ball  = int(g["is_sf"].sum() + g["is_ball"].sum())  # SF系+ボール系を分母に
        # g_st2 は削除（不使用）
        vel_mean = round(g["球速_num"].mean(), 1) if g["球速_num"].notna().any() else ""
        vel_max  = g["球速_num"].max() if g["球速_num"].notna().any() else ""
        total    = total_lr.get((gid, pname, hand), n)
        pct      = round(n / total * 100, 1) if total > 0 else 0.0
        bd       = batted_by_pitch_lr.get((gid, pname, hand, pitch), {})
        ref      = gp_ref_dict.get((gid, pname), None)
        pitch_key = to_pitch_key(str(pitch).strip())

        # 9分割（左右別）
        pit_left = (pit_hand_map_pm.get(pname, "") == "左投")
        bat_left = (hand == "左打")
        sm_in    = pit_left ^ bat_left
        sleft    = g["コース(Left)"]
        stop     = g["コース(Top)"]
        col_in   = (sleft < ZONE_LE1) if sm_in else (sleft >= ZONE_LE2)
        col_out  = (sleft >= ZONE_LE2) if sm_in else (sleft < ZONE_LE1)
        col_ctr  = ~col_in & ~col_out

        def _bd(k): return bd.get(k, "")
        def _pct(v): return "" if (v is None or (isinstance(v, float) and np.isnan(v))) else v

        pm_lr_rows.append({
            "投手試合ID":      ref["投手試合ID"] if ref is not None else "",
            "試合ID":          gid, "試合日": TARGET_DATE,
            "投手ID":          ref["投手ID"]  if ref is not None else "",
            "選手名":          pname,
            "チーム":          ref["チーム"]  if ref is not None else "",
            "ホーム/アウェイ": ref["ホーム/アウェイ"] if ref is not None else "",
            "役割":            ref["役割"]    if ref is not None else "",
            "対打者":          "右" if hand == "右打" else "左",
            "球種コード":      pitch_key, "球種名": str(pitch).strip(),
            "投球数":          n, "投球割合%": pct,
            "平均球速":        vel_mean, "最高球速": vel_max,
            "全打球数":        _bd("全打球数"),
            "GB":              _bd("GB"),   "LD":    _bd("LD"),
            "FB":              _bd("FB"),   "IFFB":  _bd("IFFB"),
            "判断不可打球":    _bd("判断不可打球"),
            "HR":              _bd("HR",),
            "H":               _bd("被安打数"),
            "GB%":             _pct(_bd("GB%")),  "LD%":   _pct(_bd("LD%")),
            "FB%":             _pct(_bd("FB%")),  "HR%":   _pct(_bd("HR%")),
            "IFFB%":           _pct(_bd("IFFB%")),
            # ゾーン系
            "空振り率":            round(g["is_swstr"].sum() / n * 100, 1) if n > 0 else 0.0,
            "ゾーン内スイング率":   round((g["is_swing"] & g["in_zone"]).sum() / g["in_zone"].sum() * 100, 1)
                               if g["in_zone"].sum() > 0 else 0.0,
            "ゾーン外スイング率":   round((g["is_swing"] & g["out_zone"]).sum() / g["out_zone"].sum() * 100, 1)
                               if g["out_zone"].sum() > 0 else 0.0,
            "ゾーン率":            round(g["in_zone"].sum() / n * 100, 1) if n > 0 else 0.0,
            "ストライク率":        round(g["is_sf"].sum() / sf_ball * 100, 1) if sf_ball > 0 else 0.0,
            "ゾーン内SW数":        int((g["is_swing"] & g["in_zone"]).sum()),
            "ゾーン内投球数":      int(g["in_zone"].sum()),
            "ゾーン外SW数":        int((g["is_swing"] & g["out_zone"]).sum()),
            "ゾーン外投球数":      int(g["out_zone"].sum()),
            "SW数":                int(g["is_swing"].sum()),
            "コンタクト数":        int(g["is_swing"].sum()) - int(g["is_swstr"].sum()),
            "空振り数":            int(g["is_swstr"].sum()),
            "In%":   round(col_in.sum()  / n * 100, 1),
            "中央%":  round(col_ctr.sum() / n * 100, 1),
            "Out%":  round(col_out.sum() / n * 100, 1),
            "高め%":  round((stop < ZONE_TE1).sum()                       / n * 100, 1),
            "真中%":  round(stop.between(ZONE_TE1, ZONE_TE2).sum()         / n * 100, 1),
            "低め%":  round((stop >= ZONE_TE2).sum()                       / n * 100, 1),
        })
    _PM_LR_COLS = [
        "配球試合ID", "投手試合ID", "試合ID", "試合日", "投手ID", "選手名", "チーム", "役割",
        "対打者", "球種コード", "球種名",
        "投球数", "投球割合%", "被安打数", "被本塁打数", "平均球速", "最高球速",
        "空振り率", "ゾーン内スイング率", "ゾーン外スイング率", "ゾーン率", "ストライク率",
        "全打球数", "GB", "LD", "FB", "IFFB", "判断不可打球",
        "GB%", "LD%", "FB%", "HR%", "IFFB%",
        "In%", "中央%", "Out%", "高め%", "真中%", "低め%",
    ]
    if pm_lr_rows:
        fact_pitch_mix_lr = pd.DataFrame(pm_lr_rows)
        fact_pitch_mix_lr.insert(0, "配球試合ID", make_seq_id("PL", len(fact_pitch_mix_lr), width=6))
    else:
        fact_pitch_mix_lr = pd.DataFrame(columns=_PM_LR_COLS)

    # --- 出力 ---
    make_output_dir()
    output_path = os.path.join(GAMES_DM_DIR, f"{TARGET_DATE}.xlsx")
    if highlights_path and os.path.exists(highlights_path):
        df_highlights = pd.read_excel(highlights_path, engine="openpyxl")
        print(f"  活躍選手データ読み込み: {highlights_path} ({len(df_highlights)} 件)")
    else:
        df_highlights = pd.DataFrame(columns=["date", "no", "player", "team", "detail"])

    # ── datamart出力カラム定義 ──
    _DM_ZONE = ['空振り率','ゾーン内スイング率','ゾーン外スイング率','ゾーン率','ストライク率',
                'ゾーン内SW数','ゾーン内投球数','ゾーン外SW数','ゾーン外投球数','SW数','コンタクト数','空振り数',
                'In%','中央%','Out%','高め%','真中%','低め%']
    _DM_BIP  = ['全打球数','GB','LD','FB','IFFB','判断不可打球','HR','H','GB%','LD%','FB%','IFFB%','HR%']
    DM_COL_PIT = (
        ['試合ID','試合日','選手名','チーム','ホーム/アウェイ','役割','勝敗成績',
         '投球回','投球数','対戦打者数','失点','自責点','被安打','被本塁打','与四球','与死球','奪三振',
         'K%','BB%','K-BB%'] + _DM_BIP + _DM_ZONE
    )
    DM_COL_PIT_LR = (
        ['試合ID','試合日','選手名','チーム','ホーム/アウェイ','役割','対打者',
         '投球数','対戦打者数','失点','自責点','被安打','被本塁打','与四球','与死球','奪三振',
         'K%','BB%','K-BB%'] + _DM_BIP + _DM_ZONE
    )
    DM_COL_MIX = (
        ['試合ID','試合日','選手名','チーム','ホーム/アウェイ','役割',
         '球種名','球種コード','投球数','投球割合%','平均球速','最高球速'] + _DM_BIP + _DM_ZONE
    )
    DM_COL_MIX_LR = (
        ['試合ID','試合日','選手名','チーム','ホーム/アウェイ','役割',
         '球種名','球種コード','対打者','投球数','投球割合%','平均球速','最高球速'] + _DM_BIP + _DM_ZONE
    )
    DM_COL_BAT = (
        ['試合ID','試合日','選手名','チーム','ホーム/アウェイ','打順','守備位置','打席別結果',
         'OPS','出塁率','長打率','打率','打点','盗塁','打席','打数','安打','本塁打','長打','単打',
         '四球','死球','三振','フライアウト(犠牲フライ含む)','ライナーアウト','ゴロアウト',
         'その他(犠打、失策、野選)','本塁打割合','長打割合','単打割合',
         'フライアウト割合','ライナーアウト割合','ゴロアウト割合','K%','BB%',
         '球数','ゾーン内SW数','ゾーン内投球数','ゾーン外SW数','ゾーン外投球数',
         'SW数','コンタクト数','空振り数','Z-Swing%','O-Swing%','Contact%','whiff%']
    )

    def _dm_reorder(df, cols):
        """datamartシート用：指定カラム順に整列、不足は空補完、余分は除外"""
        if df is None or (hasattr(df, 'empty') and df.empty): return df
        df = df.copy()
        for c in cols:
            if c not in df.columns: df[c] = ""
        return df[cols]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dim_game.to_excel(writer,              sheet_name="試合概要",              index=False)
        _dm_reorder(fact_game_batter_out, DM_COL_BAT).to_excel(writer, sheet_name="試合別打者成績",        index=False)
        _dm_reorder(fact_game_pitcher,    DM_COL_PIT).to_excel(writer, sheet_name="試合別投手成績",        index=False)
        _dm_reorder(fact_game_pitcher_lr, DM_COL_PIT_LR).to_excel(writer, sheet_name="試合別投手成績_左右別", index=False)
        _dm_reorder(fact_pitch_mix,       DM_COL_MIX).to_excel(writer, sheet_name="試合別投球配球",        index=False)
        _dm_reorder(fact_pitch_mix_lr,    DM_COL_MIX_LR).to_excel(writer, sheet_name="試合別投球配球_左右別", index=False)
        df_highlights.to_excel(writer,         sheet_name="活躍選手",              index=False)

    print(f"\n完了: '{output_path}'")
    for name, df in [
        ("試合概要",              dim_game),
        ("試合別打者成績",        fact_game_batter_out),
        ("試合別投手成績",        fact_game_pitcher),
        ("試合別投手成績_左右別", fact_game_pitcher_lr),
        ("試合別投球配球",        fact_pitch_mix),
        ("試合別投球配球_左右別", fact_pitch_mix_lr),
        ("活躍選手",              df_highlights),
    ]:
        print(f"  {name:18s}: {len(df):>5} rows")

    return output_path


# %%
# ==================================================
# Section 7b. Step 4b ── ダッシュボードJSON生成
# ==================================================

# 打席結果 日本語 → HTML略称
_ABS_MAP = {
    "空三振": "K",   "見三振": "K",
    "左安": "H",  "中安": "H",  "右安": "H",  "遊安": "H",
    "二安": "H",  "三安": "H",  "投安": "H",  "一安": "H",
    "左２": "2B", "中２": "2B", "右２": "2B",
    "左３": "3B", "中３": "3B", "右３": "3B",
    "左本": "HR", "中本": "HR", "右本": "HR", "左中本": "HR",
    "遊ゴロ": "GO", "二ゴロ": "GO", "三ゴロ": "GO",
    "投ゴロ": "GO", "捕ゴロ": "GO", "一ゴロ": "GO",
    "左飛": "FO",  "中飛": "FO",  "右飛": "FO",
    "遊飛": "FO",  "二飛": "FO",  "三飛": "FO",
    "一飛": "FO",  "投飛": "FO",  "捕飛": "FO",
    "左邪飛": "FO", "右邪飛": "FO", "三邪飛": "FO",
    "一邪飛": "FO", "捕邪飛": "FO",
    "三直": "FO",  "左直": "FO",  "中直": "FO",
    "右直": "FO",  "二直": "FO",
    "四球": "BB",  "死球": "HBP",
    "遊併打": "DP", "二併打": "DP", "三併打": "DP",
    "遊野選": "E",  "遊失": "E",   "投失": "E",
    "一失": "E",   "三失": "E",   "二失": "E",
    "投犠打": "SAC", "捕犠打": "SAC", "三犠打": "SAC",
    "右犠飛": "SF",  "左犠飛": "SF",  "中犠飛": "SF",
}


def _nv(v, default=None):
    """None / NaN → default"""
    if v is None:
        return default
    if isinstance(v, float) and math.isnan(v):
        return default
    return v


def _fv(v, decimals=1, default=0.0):
    v = _nv(v)
    if v is None:
        return default
    try:
        return round(float(v), decimals)
    except Exception:
        return default


def _iv(v, default=0):
    v = _nv(v)
    if v is None:
        return default
    try:
        return int(v)
    except Exception:
        return default


def _ip_str(ip):
    """投球回 float → '4.2' 形式文字列"""
    if ip is None:
        return "0.0"
    if isinstance(ip, int):
        return f"{ip}.0"
    whole = int(ip)
    frac = round((float(ip) - whole) * 3)
    return f"{whole}.{frac}"


def _parse_abs(result_str):
    """打席別結果文字列 → リスト"""
    if not result_str:
        return []
    return [_ABS_MAP.get(p.strip(), p.strip()) for p in str(result_str).split(",") if p.strip()]


def _lr_stat(r: dict | None) -> dict | None:
    """試合別投手成績_左右別の1行 → JSON用dict"""
    if not r:
        return None
    return {
        "pitches": _iv(r.get("投球数")),
        "tbf":     _iv(r.get("対戦打者数")),
        "k":       _iv(r.get("奪三振")),
        "bb":      _iv(r.get("与四球")),
        "hbp":     _iv(r.get("与死球")),
        "h":       _iv(r.get("被安打")),
        "kpct":    _fv(r.get("K%")),
        "bbpct":   _fv(r.get("BB%")),
        "kbbpct":  _fv(r.get("K-BB%")),
        "gbpct":   _fv(r.get("GB%")),
        "ldpct":   _fv(r.get("LD%")),
        "fbpct":   _fv(r.get("FB%")),
        "iffbpct": _fv(r.get("IFFB%")),
        "zone":    _fv(r.get("ゾーン率")),
        "oSwing":  _fv(r.get("ゾーン外スイング率")),
        "strike":  _fv(r.get("ストライク率")),
        "swstr":   _fv(r.get("空振り率")),
    }


def _mix_row(row) -> dict:
    """投球配球行 → mix要素辞書"""
    return {
        "name":    _nv(row.get("球種名"), ""),
        "key":     _nv(row.get("球種コード"), ""),
        "count":   _iv(row.get("投球数")),
        "pct":     _fv(row.get("投球割合%")),
        "hits":    _iv(row.get("H")),
        "hr":      _iv(row.get("HR")),
        "vel":     _fv(row.get("平均球速")),
        "maxVel":  _fv(row.get("最高球速")),
        "swstr":   _fv(row.get("空振り率")),
        "oSwing":  _fv(row.get("ゾーン外スイング率")),
        "zone":    _fv(row.get("ゾーン率")),
        "strike":  _fv(row.get("ストライク率")),
        "gbpct":   _fv(row.get("GB%")),
        "heart":   _fv(row.get("Heart%")),
        "shadow":  _fv(row.get("Shadow%")),
        "chase":   _fv(row.get("Chase%")),
        "waste":   _fv(row.get("Waste%")),
        # 9分割コース分布（投手視点: In=内角、Out=外角）
        "zIn":   _fv(row.get("In%")),
        "zCtr":  _fv(row.get("中央%")),
        "zOut":  _fv(row.get("Out%")),
        "zHi":   _fv(row.get("高め%")),
        "zMid":  _fv(row.get("真中%")),
        "zLo":   _fv(row.get("低め%")),
    }


def _build_dashboard_data(datamart_path: str, pitch_locs: dict | None = None, cbs_idx: dict | None = None, hand_map: dict | None = None) -> dict:
    """データマートExcel → HTMLのDATA形式 dict
    pitch_locs: {(game_id, pitcher_name, pitch_key, bat_hand): [(x, y), ...]}  bat_hand: "R"/"L"/"ALL"
    """
    import openpyxl

    wb = openpyxl.load_workbook(datamart_path)
    def to_dicts(ws):
        headers = [c.value for c in ws[1]]
        return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

    sheets = {s: to_dicts(wb[s]) for s in wb.sheetnames}

    # インデックス構築
    pit_master_idx = {}

    mix_game_idx    = defaultdict(list)
    for r in sheets.get("試合別投球配球", []):
        mix_game_idx[(r["試合ID"], r["選手名"])].append(r)

    # 試合別投球配球_左右別 → mixVsR / mixVsL
    mix_lr_idx = defaultdict(list)
    for r in sheets.get("試合別投球配球_左右別", []):
        tai = r.get("対打者", "")
        mix_lr_idx[(r["試合ID"], r["選手名"], tai)].append(r)

    # 試合別投手成績_左右別 → pitStatVsR / pitStatVsL
    pit_lr_idx = {}
    for r in sheets.get("試合別投手成績_左右別", []):
        tai = r.get("対打者", "")
        pit_lr_idx[(r["試合ID"], r["選手名"], tai)] = r

    # pitch_locs インデックス（座標データ）
    locs_idx = pitch_locs or {}
    # cbs_idx（全投球ベースのカウント別集計）
    _cbs_idx = cbs_idx or {}

    # 投手利き手マップ（hand_map: {(gid, name): "R"/"L"} or 名前→利き手dict）
    _pit_hand_map_db = {}
    if hand_map:
        for k, v in hand_map.items():
            # k は投手名（str）
            _pit_hand_map_db[k] = v

    pit_game_idx    = defaultdict(lambda: {"home": [], "away": []})
    for r in sheets.get("試合別投手成績", []):
        pit_game_idx[r["試合ID"]][r.get("ホーム/アウェイ", "home")].append(r)

    bat_game_idx    = defaultdict(lambda: {"home": [], "away": []})
    for r in sheets.get("試合別打者成績", []):
        bat_game_idx[r["試合ID"]][r.get("ホーム/アウェイ", "home")].append(r)

    def build_pitcher(pit_row):
        name    = _nv(pit_row["選手名"], "")
        game_id = pit_row["試合ID"]

        def _mix_with_locs(r, bat_hand="ALL"):
            pk         = _nv(r.get("球種コード"), "")
            pitch_name = _nv(r.get("球種名"), pk)  # 球種名キーで検索（重複防止）
            m  = _mix_row(r)
            key_all = (str(game_id), name, pitch_name, "ALL")
            key_lr  = (str(game_id), name, pitch_name, bat_hand)
            locs = locs_idx.get(key_lr) or locs_idx.get(key_all) or []
            if locs:
                m["locs"] = locs
            # cbsは全投球ベース（座標なし含む）で設定
            cbs = _cbs_idx.get(key_lr) or _cbs_idx.get(key_all)
            if cbs:
                m["cbs"] = cbs
            return m

        mix = [_mix_with_locs(r, "ALL") for r in sorted(
            mix_game_idx[(game_id, name)], key=lambda r: -_iv(r.get("投球数")))]
        season_mix = []

        m = pit_master_idx.get(name, {})
        season = {
            "gs":     _iv(_nv(m.get("先発登板数"))),
            "ip":     _nv(m.get("投球回"), "0.0"),
            "era":    _fv(m.get("防御率")),
            "whip":   _fv(m.get("WHIP")),
            "k9":     None, "bb9": None, "kBB": None, "war": None,
            "kpct":   _fv(m.get("K%")),
            "bbpct":  _fv(m.get("BB%")),
            "kbbpct": _fv(m.get("K-BB%")),
        }

        zs = pit_row
        zone_stats = {
            "swstr":  _fv(zs.get("空振り率")),
            "zone":   _fv(zs.get("ゾーン率")),
            "oSwing": _fv(zs.get("ゾーン外スイング率")),
            "strike": _fv(zs.get("ストライク率")),
            "heart":  _fv(zs.get("Heart%")),
            "shadow": _fv(zs.get("Shadow%")),
            "chase":  _fv(zs.get("Chase%")),
            "waste":  _fv(zs.get("Waste%")),
            "vsR": {
                "in":  _fv(zs.get("vs右_In%")),  "ctr": _fv(zs.get("vs右_中央%")),
                "out": _fv(zs.get("vs右_Out%")), "hi":  _fv(zs.get("vs右_高め%")),
                "mid": _fv(zs.get("vs右_真中%")), "lo":  _fv(zs.get("vs右_低め%")),
            },
            "vsL": {
                "in":  _fv(zs.get("vs左_In%")),  "ctr": _fv(zs.get("vs左_中央%")),
                "out": _fv(zs.get("vs左_Out%")), "hi":  _fv(zs.get("vs左_高め%")),
                "mid": _fv(zs.get("vs左_真中%")), "lo":  _fv(zs.get("vs左_低め%")),
            },
        }

        return {
            "name":      name,
            "role":      _nv(pit_row.get("役割"), ""),
            "hand":      "L" if _pit_hand_map_db.get(name, "") == "左投" else "R",
            "result":    _nv(pit_row.get("勝敗成績"), "ND"),
            "ip":        _ip_str(pit_row.get("投球回")),
            "pitches":   _iv(pit_row.get("投球数")),
            "h":         _iv(pit_row.get("被安打")),
            "bb":        _iv(pit_row.get("与四球")),
            "hbp":       _iv(pit_row.get("与死球")),
            "k":         _iv(pit_row.get("奪三振")),
            "r":         _iv(pit_row.get("失点")),
            "er":        _iv(pit_row.get("自責点")),
            "season":    season,
            "mix":       mix,
            "seasonMix": season_mix,
            "mixVsR":    [_mix_with_locs(r, "R") for r in sorted(
                mix_lr_idx[(game_id, name, "右")], key=lambda r: -_iv(r.get("投球数")))],
            "mixVsL":    [_mix_with_locs(r, "L") for r in sorted(
                mix_lr_idx[(game_id, name, "左")], key=lambda r: -_iv(r.get("投球数")))],
            "pitStatVsR": _lr_stat(pit_lr_idx.get((game_id, name, "右"))),
            "pitStatVsL": _lr_stat(pit_lr_idx.get((game_id, name, "左"))),
            "prevGame":  None,
            "tbf":       _iv(pit_row.get("対戦打者数")),
            "kpct":      _fv(pit_row.get("K%")),
            "gbpct":     _fv(pit_row.get("GB%")),
            "bbpct":     _fv(pit_row.get("BB%")),
            "kbbpct":    _fv(pit_row.get("K-BB%")),
            "swstr":     _fv(pit_row.get("空振り率")),
            "oSwing":    _fv(pit_row.get("ゾーン外スイング率")),
            "strike":    _fv(pit_row.get("ストライク率")),
            "zone":      _fv(pit_row.get("ゾーン率")),
            "zoneStats": zone_stats,
        }

    def build_batter(bat_row):
        return {
            "order":   _iv(bat_row.get("打順")),
            "name":    _nv(bat_row.get("選手名"), ""),
            "pos":     _nv(bat_row.get("守備位置"), ""),
            "abs":     _parse_abs(bat_row.get("打席別結果")),
            "chase":   _fv(bat_row.get("O-Swing%")),    # ボール球スイング率
            "whiff":   _fv(bat_row.get("whiff%")),        # 空振り率（whiff）
            "contact": _fv(bat_row.get("Z-Swing%")),      # ゾーン内スイング率
            "hardHit": None,
            "bbPct":   _fv(bat_row.get("BB%")),
            "kPct":    _fv(bat_row.get("K%")),
            "swStr":   _fv(bat_row.get("whiff%")),        # swStr = whiff% （空振り率）
            "ops":     _fv(bat_row.get("OPS")),
            "obp":     _fv(bat_row.get("出塁率")),
            "slg":     _fv(bat_row.get("長打率")),
            "h":       _iv(bat_row.get("安打")),
            "hr":      _iv(bat_row.get("本塁打")),
            "bb":      _iv(bat_row.get("四死球")),
            "k":       _iv(bat_row.get("三振")),
            "pa":      _iv(bat_row.get("打席数")),
            "ab":      _iv(bat_row.get("打数")),
            "rbi":     _iv(bat_row.get("打点")),
            "sb":      _iv(bat_row.get("盗塁")),
        }

    DATA = {}
    for g in sheets.get("試合概要", []):
        game_id  = g["試合ID"]
        date_str = str(g["試合日"])[:10]

        def parse_inn(s):
            return [_iv(x) for x in str(s).split(",")] if s else []

        pit_sides = pit_game_idx[game_id]
        bat_sides = bat_game_idx[game_id]

        game_obj = {
            "gameId":    game_id,
            "home":      _nv(g.get("ホームチーム"), ""),
            "away":      _nv(g.get("アウェイチーム"), ""),
            "homeScore": _iv(g.get("ホーム得点")),
            "awayScore": _iv(g.get("アウェイ得点")),
            "homeRecord":_nv(g.get("ホーム勝敗成績"), ""),
            "awayRecord":_nv(g.get("アウェイ勝敗成績"), ""),
            "stadium":   _nv(g.get("球場"), ""),
            "time":      _nv(g.get("開始時間"), ""),
            "gameTime":  _nv(g.get("試合時間"), ""),
            "att":       _iv(str(g.get("観客数") or "0").replace(",", "").replace("人", "")),
            "status":    _nv(g.get("試合状態"), ""),  # 試合中止など
            "innings": {
                "home": parse_inn(g.get("ホームイニング得点")),
                "away": parse_inn(g.get("アウェイイニング得点")),
            },
            "rhe": {
                "home": [_iv(g.get("ホーム得点")),  _iv(g.get("ホーム安打")),  _iv(g.get("ホームエラー"))],
                "away": [_iv(g.get("アウェイ得点")), _iv(g.get("アウェイ安打")), _iv(g.get("アウェイエラー"))],
            },
            "pitchers": {
                "home": [build_pitcher(r) for r in sorted(
                    pit_sides["home"], key=lambda r: (r.get("役割") != "先発", r.get("投手試合ID", "")))],
                "away": [build_pitcher(r) for r in sorted(
                    pit_sides["away"], key=lambda r: (r.get("役割") != "先発", r.get("投手試合ID", "")))],
            },
            "batters": {
                "home": [build_batter(r) for r in sorted(bat_sides["home"], key=lambda r: _iv(r.get("打順")))],
                "away": [build_batter(r) for r in sorted(bat_sides["away"], key=lambda r: _iv(r.get("打順")))],
            },
            "mvp": {
                "name": _nv(g.get("MVP選手名"), ""),
                "team": "",
                "stat": _nv(g.get("MVP成績"), ""),
            },
            "highlight": _nv(g.get("ハイライト"), ""),
        }

        DATA.setdefault(date_str, []).append(game_obj)

    # 活躍選手シートを highlights キーに格納
    hl_rows = sheets.get("活躍選手", [])
    highlights_by_date = {}
    for r in hl_rows:
        d = str(r.get("date", ""))[:10]
        if not d:
            continue
        highlights_by_date.setdefault(d, []).append({
            "no":     _iv(r.get("no")),
            "player": _nv(r.get("player"), ""),
            "team":   _nv(r.get("team"), ""),
            "detail": _nv(r.get("detail"), ""),
        })
    DATA["highlights"] = highlights_by_date

    return DATA


def _inject_into_html(html_path: str, data: dict) -> str:
    """HTMLの const DATA={...}; を data で置き換えた文字列を返す"""
    html = Path(html_path).read_text(encoding="utf-8")
    json_str = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    new_html, n = re.subn(
        r"const DATA\s*=\s*\{.*?\};",
        f"const DATA={json_str};",
        html, count=1, flags=re.DOTALL,
    )
    if n == 0:
        print("[WARN] HTMLの 'const DATA=' が見つかりませんでした", file=sys.stderr)
    return new_html


def run_dashboard(datamart_path: str, html_template: str | None = None,
                  path_pitch: str | None = None) -> str:
    """
    Step 7: データマート → ダッシュボードJSON (+ HTML) 生成

    Args:
        datamart_path     : npb_dashboard_datamart_{date}.xlsx のパス
        html_template     : 埋め込み先HTMLテンプレートのパス（None の場合はJSONのみ）
        path_pitch        : 投球データExcelのパス（座標ヒートマップ用、省略可）

    Returns:
        出力JSONのパス
    """
    # 座標データの集約（path_pitch が指定された場合）
    pitch_locs = None
    cbs_idx    = None  # カウント別集計（全投球ベース）
    if path_pitch:
        try:
            df_locs = preprocess_pitch(pd.read_excel(path_pitch))
            if "試合ID" in df_locs.columns:
                df_locs["試合ID"] = df_locs["試合ID"].astype(str)
            pitch_locs = {}
            # 座標を数値化してフィルタ（座標なし行は除外）
            df_locs["_top"]  = pd.to_numeric(df_locs["コース(Top)"],  errors="coerce")
            df_locs["_left"] = pd.to_numeric(df_locs["コース(Left)"], errors="coerce")
            df_valid = df_locs.dropna(subset=["_top", "_left"])

            # カウント別集計（全投球・座標なし含む）
            # キー: (gid, pname, pitch_name, bat_hand) → {countKey: {c,sw,fo,lo,ba,ou,hi}}
            cbs_idx = {}
            for _, crow in df_locs.iterrows():
                _cgid  = str(crow["試合ID"])
                _cpn   = str(crow.get("投手名", ""))
                _cpitch = str(crow.get("球種名", crow.get("球種", "")))
                _cbath = str(crow.get("打左右", ""))
                _chand = "R" if _cbath == "右打" else ("L" if _cbath == "左打" else "ALL")
                _cb = int(crow["pitch_balls"])   if "pitch_balls"   in crow.index and not (isinstance(crow["pitch_balls"],   float) and math.isnan(crow["pitch_balls"]))   else None
                _cs = int(crow["pitch_strikes"]) if "pitch_strikes" in crow.index and not (isinstance(crow["pitch_strikes"], float) and math.isnan(crow["pitch_strikes"])) else None
                if _cb is None or _cs is None: continue
                _ckey = f"{min(_cb,3)}-{min(_cs,2)}"
                _ccat = str(crow.get("判定カテゴリ", ""))
                _cres = str(crow.get("打席完了結果", ""))
                _cb1  = str(crow.get("1球結果", ""))
                # 結果分類
                _csw = _cb1.startswith("空振り") or _cb1.startswith("空三振")
                _cfo = "ファウル" in _cb1 and not _csw
                _clo = _ccat == "ストライク/ファウル系" and not _csw and not _cfo
                _cba = _ccat == "ボール系"
                _cou = _ccat == "アウト系"
                _chi = _ccat in ("出塁/ヒット系", "犠打/犠飛系") and ("安打" in _cres or "ヒット" in _cres)
                for _ck in [(_cgid, _cpn, _cpitch, "ALL"), (_cgid, _cpn, _cpitch, _chand)]:
                    if _ck not in cbs_idx: cbs_idx[_ck] = {}
                    if _ckey not in cbs_idx[_ck]: cbs_idx[_ck][_ckey] = {"c":0,"sw":0,"fo":0,"lo":0,"ba":0,"ou":0,"hi":0}
                    d2 = cbs_idx[_ck][_ckey]
                    d2["c"]  += 1
                    d2["sw"] += int(_csw)
                    d2["fo"] += int(_cfo)
                    d2["lo"] += int(_clo)
                    d2["ba"] += int(_cba)
                    d2["ou"] += int(_cou)
                    d2["hi"] += int(_chi)
            # デバッグ: 特定投手の集計数確認
            for _dk, _dv in list(cbs_idx.items())[:3]:
                _dtotal = sum(v["c"] for v in _dv.values())
                print(f"  [cbs_debug] {_dk}: {_dtotal}球 ({list(_dv.keys())})")
            print(f"  カウント別集計: {len(cbs_idx)} キー")
            # 投手視点の正規化: ストライクゾーンを [-1, 1] x [-1, 1] に変換
            # Left: ZONE_LEFT_MIN〜ZONE_LEFT_MAX → [-1, 1]（小=内角方向は打者左右で反転済み）
            # Top:  ZONE_TOP_MIN〜ZONE_TOP_MAX   → [1, -1]（小=高め）
            def _norm_coord(top, left, is_swstr, result):
                # x: 左右位置（-1=外角Left, +1=外角Right）
                cx = (left - (ZONE_LEFT_MIN + ZONE_LEFT_MAX) / 2) / ((ZONE_LEFT_MAX - ZONE_LEFT_MIN) / 2)
                # y: 高低位置（-1=低め, +1=高め）
                cy = -((top  - (ZONE_TOP_MIN  + ZONE_TOP_MAX)  / 2) / ((ZONE_TOP_MAX  - ZONE_TOP_MIN)  / 2))
                return round(float(cx), 3), round(float(cy), 3)
            pit_hand_map_d = df_valid.groupby("投手名")["投左右"].first().to_dict()
            for _, row in df_valid.iterrows():
                gid        = str(row["試合ID"])
                pname      = str(row.get("投手名", ""))
                pitch      = str(row.get("球種", ""))
                pk         = to_pitch_key(pitch)
                pitch_name = str(row.get("球種名", pitch))  # 球種名（スライダー/スイーパー等）
                bat_h      = row.get("打左右", "")
                hand       = "R" if bat_h == "右打" else ("L" if bat_h == "左打" else "ALL")
                top_v  = float(row["_top"])
                left_v = float(row["_left"])
                # 打者視点でx座標を反転（対左打者は内外角が逆）
                pit_left = (pit_hand_map_d.get(pname, "") == "左投")
                bat_left = (bat_h == "左打")
                flip_x = pit_left ^ bat_left
                cx = (left_v - (ZONE_LEFT_MIN + ZONE_LEFT_MAX) / 2) / ((ZONE_LEFT_MAX - ZONE_LEFT_MIN) / 2)
                if flip_x: cx = -cx
                cy = -((top_v - (ZONE_TOP_MIN + ZONE_TOP_MAX) / 2) / ((ZONE_TOP_MAX - ZONE_TOP_MIN) / 2))
                cx = round(float(cx), 3)
                cy = round(float(cy), 3)
                result_str = str(row.get("1球結果", ""))
                is_swstr   = bool(row.get("is_swstr", False))
                is_swing   = bool(row.get("is_swing", False))
                in_zone    = bool(row.get("in_zone", False))
                _cat = str(row.get("判定カテゴリ", ""))
                _res = str(row.get("打席完了結果", ""))
                _b1  = str(row.get("1球結果", ""))
                # result_type: 0=その他, 1=ゴロアウト, 2=フライ/ライナーアウト, 3=安打, 4=長打(2B/3B), 5=HR
                if   _cat=="アウト系"     and "ゴロ"  in _res: _rtype = 1
                elif _cat=="アウト系"     and ("フライ" in _res or "ライナー" in _res): _rtype = 2
                elif _cat=="出塁/ヒット系" and "本塁打" in _res: _rtype = 5
                elif _cat=="出塁/ヒット系" and ("2塁打" in _res or "3塁打" in _res): _rtype = 4
                elif _cat=="出塁/ヒット系" and ("安打" in _res or "ヒット" in _res): _rtype = 3
                else: _rtype = 0
                # is_strike: 1=ストライク（空振り/ファウル/見逃しストライク）, 0=ボール
                _is_strike = 1 if _cat == "ストライク/ファウル系" or is_swstr else 0
                # [x, y, flag, in_zone, result_type, is_strike, balls, strikes]
                _b_raw   = row["pitch_balls"]   if "pitch_balls"   in row.index else 0
                _s_raw   = row["pitch_strikes"] if "pitch_strikes" in row.index else 0
                _balls   = int(_b_raw) if (_b_raw is not None and not (isinstance(_b_raw, float) and math.isnan(_b_raw))) else 0
                _strikes = int(_s_raw) if (_s_raw is not None and not (isinstance(_s_raw, float) and math.isnan(_s_raw))) else 0
                _entry = [cx, cy, 1 if is_swstr else (0 if is_swing else -1), 1 if in_zone else 0, _rtype, _is_strike, _balls, _strikes]
                # キーは球種名で集約（同じpkでも別球種を区別: スライダー/スイーパー等）
                _pitch_name = str(row.get("球種名", pitch))  # 球種名カラムを使用
                key_all = (gid, pname, _pitch_name, "ALL")
                if key_all not in pitch_locs: pitch_locs[key_all] = []
                pitch_locs[key_all].append(_entry)
                # 左右別
                if hand in ("R", "L"):
                    key_lr = (gid, pname, _pitch_name, hand)
                    if key_lr not in pitch_locs: pitch_locs[key_lr] = []
                    pitch_locs[key_lr].append(_entry)
            print(f"  座標データ集約: {len(pitch_locs)} キー")
        except Exception as e:
            print(f"  [WARN] 座標データ集約失敗: {e}")
            pitch_locs = None

    # 投手利き手マップを構築（path_pitch の 投左右 カラムから）
    hand_map_local: dict = {}
    if path_pitch:
        try:
            _df_hand = preprocess_pitch(pd.read_excel(path_pitch))
            if "投左右" in _df_hand.columns and "投手名" in _df_hand.columns:
                hand_map_local = _df_hand.groupby("投手名")["投左右"].first().to_dict()
        except Exception:
            pass

    print(f"  データマート読み込み: {datamart_path}")
    DATA = _build_dashboard_data(datamart_path, pitch_locs=pitch_locs, cbs_idx=cbs_idx, hand_map=hand_map_local)
    # league情報をJSONに付与（HTMLでの1軍/2軍タブ切り替え用）
    DATA["_league"]    = _current_league
    DATA["_game_type"] = _current_game_type

    dates = [k for k in DATA.keys() if not k.startswith("_") and k != "highlights"]
    dates = sorted(dates)
    total = sum(len(v) for k, v in DATA.items() if not k.startswith("_") and k != "highlights")
    print(f"  日付: {dates}  試合数: {total}")

    # JSON 出力（games/json/{TARGET_DATE}.json）
    json_path = os.path.join(GAMES_JSON_DIR, f"{TARGET_DATE}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(DATA, f, ensure_ascii=False, indent=2)
    print(f"  JSON出力: {json_path}  ({os.path.getsize(json_path) // 1024} KB)")

    # index.json を更新（games/json/ 内の全 YYYY-MM-DD.json を列挙）
    index_path = os.path.join(GAMES_JSON_DIR, "index.json")
    existing = sorted(
        f for f in os.listdir(GAMES_JSON_DIR)
        if f != "index.json" and f.endswith(".json")
        and not f.startswith("season_")
    )
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump({"files": existing}, f, ensure_ascii=False, indent=2)
    print(f"  index.json 更新: {len(existing)}件")

    # HTML 埋め込み（テンプレートが指定されている場合）
    tpl = html_template
    if tpl and os.path.exists(tpl):
        new_html = _inject_into_html(tpl, DATA)
        html_out = os.path.join(JSON_DIR, f"game_dashboard_{TARGET_DATE}.html")
        Path(html_out).write_text(new_html, encoding="utf-8")
        print(f"  HTML出力: {html_out}  ({os.path.getsize(html_out) // 1024} KB)")
    elif tpl:
        print(f"  [WARN] HTMLテンプレートが見つかりません: {tpl}")

    return json_path


# %%
# ==================================================
# Section 7c. Step 3 ── 活躍選手選出（Gemini）
# ==================================================

def _hl_get_soup(url: str) -> "BeautifulSoup | None":
    """活躍選手ステップ用 get_soup（リトライあり）"""
    for attempt in range(5):
        try:
            res = requests.get(url, headers=HEADERS, timeout=15)
            res.raise_for_status()
            return BeautifulSoup(res.content, "html.parser")
        except Exception as e:
            wait = 2.0 ** attempt
            if attempt < 4:
                print(f"  [WARN] 通信エラー ({url}): {e}  → {wait:.0f}秒後にリトライ ({attempt+1}/4回)")
                time.sleep(wait)
            else:
                print(f"  [WARN] 通信エラー ({url}): {e}  → リトライ上限に達しました")
    return None


def _hl_fetch_game_links() -> tuple[list[str], list[str]]:
    """スケジュールページから stats/top URL 一覧を取得"""
    soup = _hl_get_soup(SCHEDULE_URL)
    if not soup:
        return [], []
    SITE_ROOT = "https://baseball.yahoo.co.jp"
    def abs_url(href):
        return href if href.startswith("http") else SITE_ROOT + href
    elms = soup.find_all("a", class_="bb-score__content")
    stats_urls   = [abs_url(a["href"].replace("index", "stats")) for a in elms]
    summary_urls = [abs_url(a["href"].replace("index", "top"))   for a in elms]
    return stats_urls, summary_urls


def _hl_fetch_summary(url: str) -> dict:
    """対戦カード・スコアプレーを取得"""
    import re as _re
    soup = _hl_get_soup(url)
    if not soup:
        return {"対戦カード": "取得失敗", "スコアプレー": []}
    teams = soup.select(".bb-gameDetail .bb-gameTeam__name a")
    title = (f"{teams[0].text.strip()} vs {teams[1].text.strip()}"
             if len(teams) >= 2 else "対戦カード不明")
    plays = []
    scor_ply = soup.select_one("#scor_ply")
    if scor_ply:
        for tr in scor_ply.select("tbody tr"):
            inn  = tr.select_one("th")
            td   = tr.select_one("td")
            if not (inn and td): continue
            team   = td.select_one(".bb-gameTable__team")
            player = td.select_one(".bb-gameTable__player")
            sums   = td.select(".bb-gameTable__summary")
            detail = _re.sub(
                r"投手交代:.*|守備交代:.*|リプレー検証後判定変わらず|打球が.*リクエスト",
                "", " / ".join(p.text.strip().replace("\n", " ") for p in sums)
            ).strip()
            parts = [f"{inn.get_text(strip=True)}:",
                     team.get_text(strip=True)   if team   else "",
                     player.get_text(strip=True) if player else "",
                     detail]
            plays.append(" ".join(filter(None, parts)))
    return {"対戦カード": title, "スコアプレー": plays}


def _hl_fetch_batter_stats(url: str) -> pd.DataFrame:
    """野手成績を取得"""
    import re as _re
    soup = _hl_get_soup(url)
    if not soup: return pd.DataFrame()
    rows = []
    for row in soup.find_all("tr", class_="bb-statsTable__row"):
        tds = row.find_all("td", class_=lambda x: x and "bb-statsTable__data" in x)
        if not tds: continue
        stats = [td.text.strip() for td in tds][:14]
        if len(stats) < 14: continue
        pos_raw = _re.sub(r"[()\d打走]", "", stats[0])
        stats[0] = pos_raw[0] if pos_raw else "指"
        for i in range(3, 14):
            try:    stats[i] = int(stats[i])
            except: stats[i] = 0
        rows.append(stats)
    df = pd.DataFrame(rows, columns=[
        "位置","選手名","打率","打数","得点","安打","打点",
        "三振","四球","死球","犠打","盗塁","失策","本塁打"
    ])
    return df[df["打数"] > 0].reset_index(drop=True)


def _hl_fetch_pitcher_stats(url: str) -> pd.DataFrame:
    """投手成績を取得"""
    soup = _hl_get_soup(url)
    if not soup: return pd.DataFrame()
    rows = []
    for row in soup.find_all("tr", class_="bb-scoreTable__row"):
        tds = row.find_all("td", class_=lambda x: x and "bb-scoreTable__data" in x)
        if not tds: continue
        stats = [td.text.strip() for td in tds][:14]
        if len(stats) < 14: continue
        stats[0] = "投"
        stats[1] = stats[1].replace("\n", "").strip()
        rows.append(stats)
    return pd.DataFrame(rows, columns=[
        "守備","選手名","防御率","投球回","投球数","打者数","被安打",
        "被本塁打","奪三振","与四球","与死球","ボーク","失点","自責点"
    ])


def _hl_build_input_text() -> str:
    """全試合のスクレイピング結果をテキスト化"""
    stats_urls, summary_urls = _hl_fetch_game_links()
    if not stats_urls:
        print(f"  [WARN] {TARGET_DATE} の試合リンクが取得できませんでした")
        return ""
    buf = io.StringIO()
    for stat_url, summary_url in zip(stats_urls, summary_urls):
        summary    = _hl_fetch_summary(summary_url)
        df_batter  = _hl_fetch_batter_stats(stat_url)
        df_pitcher = _hl_fetch_pitcher_stats(stat_url)
        print(f"\n==============================", file=buf)
        print(f"【{summary['対戦カード']}】", file=buf)
        print("\n▼スコアプレー", file=buf)
        for play in summary["スコアプレー"]:
            print(play, file=buf)
        print("\n▼野手成績", file=buf)
        print(df_batter.to_string(index=False), file=buf)
        print("\n▼投手成績", file=buf)
        print(df_pitcher.to_string(index=False), file=buf)
        time.sleep(1)
    text = buf.getvalue()
    buf.close()
    return text


def run_highlights() -> str | None:
    """
    Step 4: 活躍選手をスクレイピング + Gemini で選出し
            RAW_DIR に xlsx 保存、パスを返す。
            Gemini が使えない場合は None を返す。
    """
    if not _GENAI_AVAILABLE:
        print("  [SKIP] google-generativeai 未インストールのためスキップ")
        return None

    # .env 読み込み（存在する場合）
    env_file = _SCRIPT_DIR / ".env"
    if env_file.exists():
        from dotenv import load_dotenv
        load_dotenv(env_file)
    api_key = os.getenv("GEMINI_API_KEY", "")
    if not api_key:
        print("  [SKIP] GEMINI_API_KEY が未設定のためスキップ")
        return None

    # プロンプト・スキーマ読み込み
    prompt_path = Path(HIGHLIGHTS_PROMPT_FILE)
    schema_path = Path(HIGHLIGHTS_SCHEMA_FILE)
    if not prompt_path.exists():
        print(f"  [WARN] プロンプトファイルが見つかりません: {prompt_path}")
        return None
    if not schema_path.exists():
        print(f"  [WARN] スキーマファイルが見つかりません: {schema_path}")
        return None

    prompt = prompt_path.read_text(encoding="utf-8")
    import json as _json
    response_schema = _json.loads(schema_path.read_text(encoding="utf-8"))

    print("  試合データをスクレイピング中...")
    input_text = _hl_build_input_text()
    if not input_text.strip():
        print("  [WARN] 入力テキストが空のためスキップ")
        return None

    print("  Gemini API で活躍選手を選出中...")
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(
        model_name=GEMINI_MODEL,
        generation_config=genai.GenerationConfig(
            temperature=GEMINI_TEMPERATURE,
            max_output_tokens=GEMINI_MAX_OUTPUT_TOKENS,
            response_mime_type="application/json",
            response_schema=response_schema,
        )
    )

    all_players = []
    response = None  # ★追加: responseを事前に初期化しておく
    try:
        response = model.generate_content(contents=[
            {"role": "user",  "parts": [prompt]},
            {"role": "model", "parts": ["わかりました。データを送ってください。"]},
            {"role": "user",  "parts": [input_text]},
        ])
        
        if response and response.candidates[0].finish_reason.name != "STOP":
            print(f"  [WARN] 生成が中断されました。理由: {response.candidates[0].finish_reason.name}")
            
        players = _json.loads(response.text)
        all_players.extend(players)
        print(f"  活躍選手 {len(all_players)} 件取得")
    except Exception as e:
        print(f"  [WARN] Gemini API制限またはエラー: {e}")
        # response が定義されており、かつ中身がある場合のみ表示
        if response is not None:
            try:
                print("  ▼ Geminiの出力内容:")
                print(response.text)
            except:
                pass
        print("  [INFO] 活躍選手の選出をスキップして続行します。")

    if all_players:
        df = pd.DataFrame(all_players)
        for col in ["no", "player", "team", "detail"]:
            if col not in df.columns:
                df[col] = ""
        df = df[["no", "player", "team", "detail"]]
    else:
        df = pd.DataFrame(columns=["no", "player", "team", "detail"])

    df.insert(0, "date", TARGET_DATE)

    # RAW_DIR に xlsx として保存
    make_output_dir()
    date_nodash = TARGET_DATE.replace("-", "")
    output_path = os.path.join(RAW_DIR, f"highlights_{date_nodash}.xlsx")
    df.to_excel(output_path, index=False, engine="openpyxl")
    print(f"  xlsx出力: {output_path}  ({len(df)} 件)")
    return output_path


# %%
# ==================================================
# Section 8. メイン実行（4 Step を順番に実行）
# ==================================================

def run_all():
    print("=" * 50)
    print(f"NPB データパイプライン開始: {TARGET_DATE}")
    print("=" * 50)

    print("\n--- Step 1: 試合データ取得 ---")
    path_all_games = run_game_scraper()
    if not path_all_games:
        print("Step 1 失敗。処理を中断します。")
        return

    print("\n--- Step 2: 投球データ取得 ---")
    path_pitch = run_pitch_scraper()
    if not path_pitch:
        print("Step 2 失敗。処理を中断します。")
        return

    print("\n--- Step 3: 活躍選手選出 ---")
    path_highlights = run_highlights()

    print("\n--- Step 4: データマート&JSON作成 ---")
    path_datamart = run_datamart(path_all_games, path_pitch, path_highlights)
    path_json = run_dashboard(path_datamart, path_pitch=path_pitch)

    print("\n--- Step 5: シーズン成績取得 ---")

    print("\n--- Step 5: 選手個人データ出力 ---")
    try:
        _export_player_data(path_datamart)
    except Exception as e:
        print(f"  [WARN] 選手個人データ出力失敗: {e}")

    print("\n" + "=" * 50)
    print("✅ 全工程完了!")
    print(f"  試合データ       : {path_all_games}")
    print(f"  投球データ       : {path_pitch}")
    print(f"  活躍選手         : {path_highlights or 'スキップ'}")
    print(f"  データマート     : {path_datamart}")
    print(f"  ダッシュボードJSON: {path_json}")
    print("=" * 50)


# %%
# ==================================================
# Section 9. CLI（引数でステップを選択して実行）
# ==================================================
# 使い方:
#
#   python scripts/run.py                                     # 1軍&2軍：全ステップ
#   python scripts/run.py --steps games                       # 試合データ取得 → データマート&JSON
#   python scripts/run.py --steps pitch                       # 投球データ取得 → データマート&JSON
#   python scripts/run.py --steps pitch 2021040109,2021040110 # 特定試合の投球取得 → データマート&JSON
#   python scripts/run.py --steps highlights                  # 活躍選手選出 → データマート&JSON
#   python scripts/run.py --steps datamart                    # データマート&JSON作成のみ
#   python scripts/run.py --steps pitch highlights            # 投球 + 活躍選手 → データマート&JSON
#
# ▼ --1軍 / --2軍: リーグ指定
#   python scripts/run.py --1軍                               # 1軍：全ステップ
#   python scripts/run.py --2軍 --steps pitch                 # 2軍：投球取得 → データマート&JSON
#
# オプション:
#   --date 2026-03-20      対象日付を上書き

import argparse

# 使用できるステップ（json は datamart に自動付与されるため単独指定不要）
STEP_CHOICES = ["all", "games", "pitch", "highlights", "datamart"]

STEP_LABELS = {
    "games"     : "Step1 試合データ取得",
    "pitch"     : "Step2 投球データ取得",
    "highlights": "Step3 活躍選手選出",
    "datamart"  : "Step4 データマート&JSON作成",
    "json"      : "Step4 ダッシュボードJSON生成",

}

def parse_args():
    parser = argparse.ArgumentParser(
        description="NPB データ収集 → データマート 生成パイプライン",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--steps",
        nargs="+",
        choices=STEP_CHOICES,
        default=["all"],
        metavar="STEP",
        help=(
            "実行するステップ（スペース区切りで複数指定）。\n"
            "\n"
            "  games      : 試合データ取得\n"
            "  pitch      : 投球データ取得\n"
            "  highlights : 活躍選手選出\n"
            "  datamart   : データマート&JSON作成のみ（RAW再利用）\n"
            "  all        : 全ステップ実行（デフォルト）\n"
            "\n"
            "複数指定例:\n"
            "  --steps games pitch           : 試合+投球データ取得\n"
            "  --steps games pitch highlights: 全RAW取得+活躍選手\n"
        ),
    )
    parser.add_argument(
        "--date",
        default=None,
        metavar="DATE",
        help=(
            "対象日付の指定（省略時はスクリプト内の TARGET_DATE を使用）\n"
            "\n"
            "  特定の1日  : --date 2026-03-27\n"
            "  複数日     : --date 2026-03-27,2026-03-28,2026-03-30\n"
            "  期間       : --date 2026-03-20:2026-03-28\n"
        ),
    )
    parser.add_argument(
        "--html",
        default=None,
        metavar="PATH",
        help="ダッシュボードHTML埋め込み先テンプレート（例: game_dashboard_v111.html）",
    )

    return parser.parse_args()


def _check_datamart_quality(datamart_path: str) -> None:
    """
    データマートの内容を確認し、欠損・異常をログ出力する。
    """
    if not os.path.exists(datamart_path):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(datamart_path)
        issues = []

        def sheet_rows(name):
            ws = wb[name]
            headers = [c.value for c in ws[1]]
            return headers, [
                dict(zip(headers, row))
                for row in ws.iter_rows(min_row=2, values_only=True)
            ]

        # ① 試合シート確認
        _, games = sheet_rows("試合概要")
        game_ids = [g["試合ID"] for g in games]
        if not games:
            issues.append("⚠️  試合シート: データなし")

        # ② 投手成績: 試合IDごとの投球数
        _, fp = sheet_rows("試合別投手成績")
        fp_by_game = {}
        for r in fp:
            fp_by_game.setdefault(r["試合ID"], []).append(r)

        for gid in game_ids:
            if gid not in fp_by_game:
                issues.append(f"⚠️  試合別投手成績: 試合ID {gid} のデータなし")
            else:
                zero_pitch = [r["選手名"] for r in fp_by_game[gid]
                               if not r.get("投球数") or r["投球数"] == 0]
                if zero_pitch:
                    issues.append(f"⚠️  試合別投手成績: 試合ID {gid} 投球数=0 → {zero_pitch}")

        # ③ 投球配球: 試合IDごとの球種数
        _, pm = sheet_rows("試合別投球配球")
        pm_game_ids = set(r["試合ID"] for r in pm)
        for gid in game_ids:
            if gid not in pm_game_ids:
                issues.append(f"⚠️  試合別投球配球: 試合ID {gid} のデータなし（投球データ未取得の可能性）")

        # ④ 打者成績
        _, gb = sheet_rows("試合別打者成績")
        gb_by_game = {}
        for r in gb:
            gb_by_game.setdefault(r["試合ID"], []).append(r)
        for gid in game_ids:
            if gid not in gb_by_game:
                issues.append(f"⚠️  試合別打者成績: 試合ID {gid} のデータなし")

        # ⑥ 活躍選手
        if "活躍選手" in wb.sheetnames:
            _, hl = sheet_rows("活躍選手")
            if not hl:
                issues.append("ℹ️  活躍選手: データなし（highlights ステップ未実行の可能性）")

        print("\n" + "=" * 50)
        if issues:
            print("📋 データ品質チェック結果 ── 以下の問題が検出されました")
            for issue in issues:
                print(f"  {issue}")
            print()
            print("  投球データが欠損している試合がある場合は以下で再取得できます:")
            missing_games = [
                str(gid) for gid in game_ids
                if gid not in pm_game_ids
            ]
            if missing_games:
                print(f"  python scripts/run.py --steps pitch {','.join(missing_games)}")
        else:
            print("✅ データ品質チェック: 問題なし")
        print("=" * 50)

    except Exception as e:
        print(f"\n[WARN] データ品質チェック中にエラー: {e}")


def run_steps(steps: list[str], target_game_ids: list[str] | None = None,
              league: str = "ichi", date: str | None = None):
    """
    ステップ順:
      Step1: games     → raw/all_games_{date}.xlsx
      Step2: pitch     → raw/daily_pitch_data_{date}.xlsx
      Step3: highlights→ raw/highlights_{date}.xlsx
      Step4: datamart  → games/datamart/{date}.xlsx + games/json/{date}.json


    自動付与ルール:
      games/pitch/highlights → datamart を自動付与
      season                 → players を自動付与
      players 単独           → 他ステップ付与なし
    """
    _run_date = date if date else TARGET_DATE
    set_league_dirs(league, _run_date)
    league_label = "1軍" if league == "ichi" else "2軍"
    type_label   = _current_game_type

    # FULL_ORDER（新しいステップ順）
    FULL_ORDER = ["games", "pitch", "highlights", "datamart"]

    # "all" → 全ステップ
    if "all" in steps:
        steps = list(FULL_ORDER)

    # 重複除去・定義順を維持
    steps = [s for s in FULL_ORDER if s in set(steps)]

    # 自動付与ルール


    # games/pitch/highlights がある → datamart を自動付与
    if any(s in steps for s in ["games", "pitch", "highlights"]):
        if "datamart" not in steps:
            steps.append("datamart")
    # 並び替え
    steps = [s for s in FULL_ORDER if s in set(steps)]

    full_steps = steps

    print("=" * 50)
    print(f"NPB データパイプライン開始: {TARGET_DATE} [{league_label} / {type_label}]")
    label_str = " → ".join(STEP_LABELS[s] for s in full_steps)
    if target_game_ids:
        label_str = label_str.replace(
            STEP_LABELS["pitch"],
            f"{STEP_LABELS['pitch']}（{','.join(target_game_ids)}）"
        )
    print(f"実行ステップ: {label_str}")
    print("=" * 50)

    results = {}

    def resolve(key: str) -> str:
        date_nd = TARGET_DATE.replace("-", "")
        year    = TARGET_DATE[:4]
        paths = {
            "games"     : os.path.join(RAW_DIR, f"all_games_{TARGET_DATE}.xlsx"),
            "pitch"     : os.path.join(RAW_DIR, f"daily_pitch_data_{TARGET_DATE}.xlsx"),
            "highlights": os.path.join(RAW_DIR, f"highlights_{date_nd}.xlsx"),
            "datamart"  : os.path.join(GAMES_DM_DIR, f"{TARGET_DATE}.xlsx"),
        }
        p = paths.get(key, "")
        return p if os.path.exists(p) else ""

    for step in full_steps:
        print(f"\n--- {STEP_LABELS[step]} ---")

        if step == "games":
            path = run_game_scraper()
            if not path:
                print(f"[WARN] {STEP_LABELS[step]}: 試合なし or 取得失敗。後続ステップは既存ファイルを使用します。")
            else:
                results["games"] = path

        elif step == "pitch":
            path = run_pitch_scraper(target_game_ids=target_game_ids)
            if not path:
                print(f"[WARN] {STEP_LABELS[step]} に失敗しました。datamartをスキップして続行します。")
                # pitch失敗時はdatamartもスキップするが、playersは継続可能
                results["pitch"] = ""
            else:
                results["pitch"] = path

        elif step == "highlights":
            path = run_highlights()
            if path:
                results["highlights"] = path
            # 失敗してもパイプライン継続

        elif step == "datamart":
            path_games = results.get("games",  resolve("games"))
            path_pitch = results.get("pitch",  resolve("pitch"))

            if not path_games or not os.path.exists(path_games):
                print(f"[WARN] all_games ファイルが見つかりません。datamartをスキップします。")
            elif not path_pitch or not os.path.exists(path_pitch):
                print(f"[WARN] daily_pitch_data ファイルが見つかりません。datamartをスキップします。")
            else:
                date_nd = TARGET_DATE.replace("-", "")
                path_hl = results.get("highlights") or os.path.join(RAW_DIR, f"highlights_{date_nd}.xlsx")
                if not os.path.exists(path_hl):
                    path_hl = None

                path = run_datamart(path_games, path_pitch, path_hl)
                results["datamart"] = path

                print(f"\n--- {STEP_LABELS['json']} ---")
                if path and os.path.exists(path):
                    path_json = run_dashboard(path, path_pitch=path_pitch)
                    results["json"] = path_json



    print("\n" + "=" * 50)
    print("✅ 完了!")
    for s in full_steps:
        if s in results:
            key = s if s != "datamart" else "datamart"
            print(f"  {STEP_LABELS[s]:24s}: {results[s]}")
        if s == "datamart" and "json" in results:
            print(f"  {STEP_LABELS['json']:24s}: {results['json']}")
    print("=" * 50)

    # データ品質チェック
    dm_path = results.get("datamart") or resolve("datamart")
    if dm_path:
        _check_datamart_quality(dm_path)


def parse_date_arg(date_str: str) -> list[str]:
    """
    --date 引数を日付リストに変換する。

    形式:
        1日    : "2026-03-27"           → ["2026-03-27"]
        複数日 : "2026-03-27,2026-03-28" → ["2026-03-27", "2026-03-28"]
        期間   : "2026-03-20:2026-03-28" → ["2026-03-20", "2026-03-21", ..., "2026-03-28"]
    """
    from datetime import date, timedelta

    date_str = date_str.strip()

    if ":" in date_str:
        # 期間指定
        parts = date_str.split(":")
        if len(parts) != 2:
            raise ValueError(f"期間指定は 'YYYY-MM-DD:YYYY-MM-DD' の形式で入力してください: {date_str}")
        start = date.fromisoformat(parts[0].strip())
        end   = date.fromisoformat(parts[1].strip())
        if start > end:
            raise ValueError(f"開始日が終了日より後になっています: {start} > {end}")
        dates = []
        cur = start
        while cur <= end:
            dates.append(cur.isoformat())
            cur += timedelta(days=1)
        return dates

    elif "," in date_str:
        # 複数日指定
        dates = [d.strip() for d in date_str.split(",") if d.strip()]
        # フォーマット検証
        for d in dates:
            date.fromisoformat(d)
        return sorted(set(dates))  # 重複除去・日付順

    else:
        # 1日指定
        date.fromisoformat(date_str)  # フォーマット検証
        return [date_str]


if __name__ == "__main__":
    import sys as _sys

    _raw_args    = _sys.argv[1:]
    _game_id_str = None
    _filtered    = []
    _in_steps    = False
    _league_flag = None  # None=both / "ichi" / "ni"

    for _a in _raw_args:
        # --1軍 / --2軍 を事前に解釈（argparse に渡さない）
        if _a == "--1軍":
            _league_flag = "ichi"
            continue
        if _a == "--2軍":
            _league_flag = "ni"
            continue
        if _a == "--steps":
            _in_steps = True
            _filtered.append(_a)
        elif _in_steps and not _a.startswith("-"):
            if _a not in STEP_CHOICES and _a != "all":
                _game_id_str = _a  # カンマ区切り試合ID
            else:
                _filtered.append(_a)
        else:
            if _a.startswith("-"):
                _in_steps = False
            _filtered.append(_a)

    _sys.argv = [_sys.argv[0]] + _filtered
    args = parse_args()

    _date_arg  = args.date if args.date else TARGET_DATE
    _target_game_ids = _game_id_str.split(",") if _game_id_str else None
    _steps = list(args.steps)
    # 日付リストに展開
    try:
        _date_list = parse_date_arg(_date_arg)
    except ValueError as e:
        print(f"[ERROR] --date の指定が正しくありません: {e}")
        _sys.exit(1)

    # 複数日の場合はサマリー表示
    if len(_date_list) > 1:
        print("=" * 50)
        print(f"▶ 複数日実行モード: {len(_date_list)}日分")
        print(f"  {_date_list[0]} 〜 {_date_list[-1]}")
        print("=" * 50)
        print()

    for _date in _date_list:
        if len(_date_list) > 1:
            print(f"\n{'─' * 50}")
            print(f"▶ {_date} 処理開始")
            print(f"{'─' * 50}")

        if _league_flag is None:
            print("=" * 50)
            print("▶ 1軍 & 2軍 両方実行モード")
            print("=" * 50)
            print()
            run_steps(_steps, target_game_ids=_target_game_ids, league="ichi", date=_date)
            print()
            run_steps(_steps, target_game_ids=_target_game_ids, league="ni",   date=_date)
        else:
            run_steps(_steps, target_game_ids=_target_game_ids, league=_league_flag, date=_date)

    if len(_date_list) > 1:
        print(f"\n{'=' * 50}")
        print(f"▶ 全{len(_date_list)}日分の処理が完了しました")
        print(f"  {_date_list[0]} 〜 {_date_list[-1]}")
        print("=" * 50)